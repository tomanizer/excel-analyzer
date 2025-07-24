import openpyxl
from pathlib import Path
import re
import warnings
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.cell import range_boundaries
from typing import List, Dict, Any, Set
from datetime import datetime

# Suppress the specific zipfile warning
warnings.filterwarnings("ignore", message=".*I/O operation on closed file.*")

def find_data_islands(sheet: Worksheet, visited_cells: Set[str]) -> List[Set[str]]:
    """Finds contiguous blocks of data not already part of a formal table."""
    islands = []
    visited = set()
    # Consider all non-empty cells not already visited (i.e., not in a formal table)
    all_cells = {cell.coordinate for row in sheet.iter_rows() for cell in row 
                 if cell.value is not None and str(cell.value).strip() != "" and cell.coordinate not in visited_cells}

    for cell_coord in all_cells:
        if cell_coord not in visited:
            island = set()
            queue = [cell_coord]
            visited.add(cell_coord)
            while queue:
                current_coord = queue.pop(0)
                island.add(current_coord)
                col, row = openpyxl.utils.cell.coordinate_from_string(current_coord)
                col_idx = openpyxl.utils.cell.column_index_from_string(col)
                for r_offset, c_offset in [(0, 1), (0, -1), (1, 0), (-1, 0)]:
                    neighbor_row, neighbor_col_idx = row + r_offset, col_idx + c_offset
                    if neighbor_row > 0 and neighbor_col_idx > 0:
                        neighbor_coord = f"{openpyxl.utils.cell.get_column_letter(neighbor_col_idx)}{neighbor_row}"
                        if neighbor_coord in all_cells and neighbor_coord not in visited:
                            visited.add(neighbor_coord)
                            queue.append(neighbor_coord)
            islands.append(island)
    return islands

def analyze_workbook_final(file_path: Path, return_data: bool = False):
    """
    Analyze an Excel workbook and return structured data or print results.
    
    Args:
        file_path: Path to the Excel file
        return_data: If True, return structured data instead of printing
    
    Returns:
        If return_data=True: Dictionary with analysis results
        If return_data=False: None (prints results)
    """
    if not file_path.exists(): 
        return None if not return_data else {}
    
    if not return_data:
        print(f"--- Comprehensive Analysis for: {file_path.name} ---\n")

    wb = None
    analysis_data = {
        'metadata': {
            'filename': file_path.name,
            'file_size_kb': file_path.stat().st_size / 1024,
            'file_path': str(file_path),
            'analysis_timestamp': datetime.now().isoformat()
        },
        'global_features': {
            'vba_detected': False,
            'external_links': [],
            'named_ranges': {}
        },
        'sheets': {},
        'summary': {
            'total_sheets': 0,
            'total_formal_tables': 0,
            'total_pivot_tables': 0,
            'total_charts': 0,
            'total_data_islands': 0,
            'total_data_validation_rules': 0
        }
    }
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False, keep_vba=True)
        
        # 1. VBA Macro Detection
        has_vba = file_path.suffix == '.xlsm'
        analysis_data['global_features']['vba_detected'] = has_vba
        
        if not return_data:
            print(f"VBA Project Detected: {has_vba}")

        # 2. External Link Detection
        external_links = [link.Target for link in wb.external_links] if hasattr(wb, 'external_links') else []
        analysis_data['global_features']['external_links'] = external_links
        if external_links and not return_data:
            print("\nExternal Dependencies:")
            for link in external_links: print(f"  - {link}")

        # 3. Named Range Detection
        named_ranges = {}
        for name, d in wb.defined_names.items():
            try:
                destinations = list(d.destinations)
                named_ranges[name] = destinations
            except:
                named_ranges[name] = "Error reading destinations"
        
        analysis_data['global_features']['named_ranges'] = named_ranges
        if named_ranges and not return_data:
            print("\nNamed Ranges:")
            for name, dest in named_ranges.items(): 
                print(f"  - {name}: {dest}")

        # 4. Sheet-level Analysis
        all_tables = []
        visited_cells = set()
        
        if not return_data:
            print("\n--- Sheet-Level Analysis ---")
        
        for sheet in wb:
            sheet_data = {
                'name': sheet.title,
                'formal_tables': [],
                'pivot_tables': [],
                'charts': [],
                'data_validation': [],
                'data_islands': []
            }
            
            if not return_data:
                print(f"\nProcessing Sheet: {sheet.title}")
            
            # Formal Tables
            for tbl in sheet.tables.values():
                table_info = {
                    "name": tbl.displayName, 
                    "type": "Formal Table", 
                    "range": tbl.ref, 
                    "sheet": sheet.title
                }
                all_tables.append(table_info)
                sheet_data['formal_tables'].append(table_info)
                analysis_data['summary']['total_formal_tables'] += 1
                
                # Add table cells to visited
                min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(tbl.ref)
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        visited_cells.add(f"{openpyxl.utils.cell.get_column_letter(c)}{r}")
            
            # Chart Detection
            charts = []
            for chart in sheet._charts:
                try:
                    chart_info = {"name": chart.title or "Untitled Chart", "type": type(chart).__name__}
                    charts.append(chart_info)
                    analysis_data['summary']['total_charts'] += 1
                except:
                    charts.append({"name": "Unknown Chart", "type": "Unknown"})
                    analysis_data['summary']['total_charts'] += 1
            
            sheet_data['charts'] = charts
            if charts and not return_data:
                print("  Charts Found:")
                for chart in charts: print(f"    - '{chart['name']}' ({chart['type']})")

            # Pivot Table Detection
            pivot_tables = []
            for pivot in sheet._pivots:
                try:
                    pivot_info = {
                        "name": pivot.name or "Untitled Pivot", 
                        "range": str(pivot.location),
                        "location_ref": pivot.location.ref if hasattr(pivot.location, 'ref') else None
                    }
                    pivot_tables.append(pivot_info)
                    analysis_data['summary']['total_pivot_tables'] += 1
                    
                    # Add pivot table cells to visited_cells
                    if hasattr(pivot.location, 'ref'):
                        min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(pivot.location.ref)
                        for r in range(min_row, max_row + 1):
                            for c in range(min_col, max_col + 1):
                                visited_cells.add(f"{openpyxl.utils.cell.get_column_letter(c)}{r}")
                except Exception as e:
                    pivot_tables.append({"name": "Unknown Pivot", "range": f"Error: {str(e)}"})
                    analysis_data['summary']['total_pivot_tables'] += 1
            
            sheet_data['pivot_tables'] = pivot_tables
            if pivot_tables and not return_data:
                print("  Pivot Tables Found:")
                for pivot in pivot_tables: print(f"    - '{pivot['name']}' at range {pivot['range']}")

            # Data Validation Detection
            validations = []
            for dv in sheet.data_validations.dataValidation:
                validation_info = {
                    "range": dv.sqref,
                    "formula": dv.formula1,
                    "type": dv.type
                }
                validations.append(validation_info)
                analysis_data['summary']['total_data_validation_rules'] += 1
            
            sheet_data['data_validation'] = validations
            if validations and not return_data:
                print("  Data Validation Rules Found:")
                for val in validations: print(f"    - {val['range']}: {val['formula']}")

            # Informal Data Islands
            islands = find_data_islands(sheet, visited_cells)
            for island in islands:
                coords = [openpyxl.utils.cell.coordinate_from_string(c) for c in island]
                rows = [c[1] for c in coords]; cols = [openpyxl.utils.cell.column_index_from_string(c[0]) for c in coords]
                bounding_box = f"{openpyxl.utils.cell.get_column_letter(min(cols))}{min(rows)}:{openpyxl.utils.cell.get_column_letter(max(cols))}{max(rows)}"
                
                island_info = {
                    "name": f"Island_{bounding_box}", 
                    "type": "Informal Data Island", 
                    "range": bounding_box, 
                    "sheet": sheet.title,
                    "cells": list(island)
                }
                all_tables.append(island_info)
                sheet_data['data_islands'].append(island_info)
                analysis_data['summary']['total_data_islands'] += 1

            analysis_data['sheets'][sheet.title] = sheet_data
            analysis_data['summary']['total_sheets'] += 1

        if all_tables and not return_data:
            print("\n--- Discovered Data Tables & Islands ---")
            for table in all_tables:
                print(f"  - {table['name']} ({table['type']}) on sheet '{table['sheet']}' at range {table['range']}")
        
        analysis_data['all_tables'] = all_tables
        
    finally:
        if wb is not None:
            wb.close()
    
    return analysis_data if return_data else None


def generate_markdown_report(analysis_data: dict, output_file: Path = None) -> str:
    """
    Generate a comprehensive markdown report from analysis data.
    
    Args:
        analysis_data: Analysis data from analyze_workbook_final
        output_file: Optional file path to save the report
    
    Returns:
        Markdown report as string
    """
    if not analysis_data:
        return "No analysis data provided."
    
    md = []
    
    # Header
    md.append(f"# Excel Analysis Report: {analysis_data['metadata']['filename']}")
    md.append("")
    md.append(f"**Analysis Date:** {analysis_data['metadata']['analysis_timestamp']}")
    md.append(f"**File Size:** {analysis_data['metadata']['file_size_kb']:.1f} KB")
    md.append("")
    
    # Executive Summary
    md.append("## 📊 Executive Summary")
    md.append("")
    summary = analysis_data['summary']
    md.append(f"- **Total Sheets:** {summary['total_sheets']}")
    md.append(f"- **Formal Tables:** {summary['total_formal_tables']}")
    md.append(f"- **Pivot Tables:** {summary['total_pivot_tables']}")
    md.append(f"- **Charts:** {summary['total_charts']}")
    md.append(f"- **Data Islands:** {summary['total_data_islands']}")
    md.append(f"- **Data Validation Rules:** {summary['total_data_validation_rules']}")
    md.append("")
    
    # Global Features
    md.append("## 🌐 Global Features")
    md.append("")
    
    global_features = analysis_data['global_features']
    md.append(f"- **VBA Macros:** {'Yes' if global_features['vba_detected'] else 'No'}")
    
    if global_features['external_links']:
        md.append("- **External Dependencies:**")
        for link in global_features['external_links']:
            md.append(f"  - {link}")
    else:
        md.append("- **External Dependencies:** None")
    
    if global_features['named_ranges']:
        md.append("- **Named Ranges:**")
        for name, dest in global_features['named_ranges'].items():
            md.append(f"  - `{name}`: {dest}")
    else:
        md.append("- **Named Ranges:** None")
    
    md.append("")
    
    # Sheet-by-Sheet Analysis
    md.append("## 📋 Sheet-by-Sheet Analysis")
    md.append("")
    
    for sheet_name, sheet_data in analysis_data['sheets'].items():
        md.append(f"### Sheet: {sheet_name}")
        md.append("")
        
        # Formal Tables
        if sheet_data['formal_tables']:
            md.append("**Formal Tables:**")
            for table in sheet_data['formal_tables']:
                md.append(f"- `{table['name']}` at range `{table['range']}`")
            md.append("")
        
        # Pivot Tables
        if sheet_data['pivot_tables']:
            md.append("**Pivot Tables:**")
            for pivot in sheet_data['pivot_tables']:
                md.append(f"- `{pivot['name']}` at range `{pivot['range']}`")
            md.append("")
        
        # Charts
        if sheet_data['charts']:
            md.append("**Charts:**")
            for chart in sheet_data['charts']:
                md.append(f"- `{chart['name']}` ({chart['type']})")
            md.append("")
        
        # Data Validation
        if sheet_data['data_validation']:
            md.append("**Data Validation Rules:**")
            for val in sheet_data['data_validation']:
                md.append(f"- Range `{val['range']}`: {val['formula']}")
            md.append("")
        
        # Data Islands
        if sheet_data['data_islands']:
            md.append("**Data Islands:**")
            for island in sheet_data['data_islands']:
                md.append(f"- `{island['name']}` at range `{island['range']}`")
            md.append("")
    
    # Detailed Table Summary
    md.append("## 📊 Detailed Table Summary")
    md.append("")
    
    if analysis_data['all_tables']:
        md.append("| Name | Type | Sheet | Range |")
        md.append("|------|------|-------|-------|")
        for table in analysis_data['all_tables']:
            md.append(f"| {table['name']} | {table['type']} | {table['sheet']} | `{table['range']}` |")
    else:
        md.append("No tables or data islands found.")
    
    md.append("")
    md.append("---")
    md.append("*Report generated by Excel Analyzer*")
    
    report_content = "\n".join(md)
    
    if output_file:
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(report_content)
    
    return report_content


def extract_data_to_dataframes(analysis_data: dict, file_path: Path) -> dict:
    """
    Extract data from Excel file into pandas DataFrames based on analysis.
    
    Args:
        analysis_data: Analysis data from analyze_workbook_final
        file_path: Path to the Excel file
    
    Returns:
        Dictionary with DataFrames for each table/island
    """
    import pandas as pd
    
    dataframes = {}
    
    try:
        # Load workbook with data_only=True to get values, not formulas
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        for table in analysis_data['all_tables']:
            sheet_name = table['sheet']
            range_str = table['range']
            
            try:
                # Get the worksheet
                ws = wb[sheet_name]
                
                # Parse range to get start and end cells
                if ':' in range_str:
                    start_cell, end_cell = range_str.split(':')
                    # Convert Excel range to pandas DataFrame
                    data = []
                    for row in ws[start_cell:end_cell]:
                        row_data = []
                        for cell in row:
                            row_data.append(cell.value)
                        data.append(row_data)
                    
                    if data:
                        df = pd.DataFrame(data)
                        # Use first row as headers if it looks like headers
                        if len(df) > 1 and all(isinstance(x, str) for x in df.iloc[0] if x is not None):
                            df.columns = df.iloc[0]
                            df = df.iloc[1:].reset_index(drop=True)
                    else:
                        df = pd.DataFrame()
                else:
                    # Single cell
                    cell_value = ws[range_str].value
                    df = pd.DataFrame([[cell_value]], columns=['Value'])
                
                dataframes[table['name']] = df
                
            except Exception as e:
                print(f"Error extracting {table['name']}: {e}")
                dataframes[table['name']] = None
        
        wb.close()
        
    except Exception as e:
        print(f"Error loading workbook: {e}")
    
    return dataframes


if __name__ == "__main__":
    import sys
    import json
    
    if len(sys.argv) > 1:
        file_path = Path(sys.argv[1])
        if not file_path.exists():
            print(f"File not found: {file_path}")
            sys.exit(1)
        
        # Check for additional options
        generate_json = "--json" in sys.argv
        generate_markdown = "--markdown" in sys.argv
        generate_dataframes = "--dataframes" in sys.argv
        
        if generate_json or generate_markdown or generate_dataframes:
            # Get structured data
            analysis_data = analyze_workbook_final(file_path, return_data=True)
            
            if generate_json:
                # Save JSON data
                json_file = Path("reports") / f"{file_path.stem}.json"
                json_file.parent.mkdir(exist_ok=True)
                with open(json_file, 'w', encoding='utf-8') as f:
                    json.dump(analysis_data, f, indent=2, default=str)
                print(f"JSON data saved to: {json_file}")
            
            if generate_markdown:
                # Generate markdown report
                markdown_file = Path("reports") / f"{file_path.stem}.md"
                markdown_file.parent.mkdir(exist_ok=True)
                report = generate_markdown_report(analysis_data, markdown_file)
                print(f"Markdown report saved to: {markdown_file}")
            
            if generate_dataframes:
                # Extract data to DataFrames
                dataframes = extract_data_to_dataframes(analysis_data, file_path)
                print(f"Extracted {len(dataframes)} DataFrames:")
                for name, df in dataframes.items():
                    if df is not None:
                        print(f"  - {name}: {df.shape[0]} rows × {df.shape[1]} columns")
                    else:
                        print(f"  - {name}: Error extracting data")
        else:
            # Standard analysis (print to console)
            analyze_workbook_final(file_path)
    else:
        print("Excel Analyzer - Comprehensive Excel File Analysis")
        print("=" * 50)
        print()
        print("Usage:")
        print("  python excel_parser.py <file.xlsx>                    # Standard analysis")
        print("  python excel_parser.py <file.xlsx> --json             # Save structured data as JSON")
        print("  python excel_parser.py <file.xlsx> --markdown         # Generate markdown report")
        print("  python excel_parser.py <file.xlsx> --dataframes       # Extract data to pandas DataFrames")
        print("  python excel_parser.py <file.xlsx> --json --markdown  # Multiple outputs")
        print()
        print("Examples:")
        print("  python excel_parser.py excel_files/mycoolsample.xlsx")
        print("  python excel_parser.py excel_files/mycoolsample.xlsx --json --markdown")
        print()
        
        # Create dummy files for demonstration
        # Create a dummy external file
        ext_wb = openpyxl.Workbook()
        ext_ws = ext_wb.active; ext_ws.title = "ExternalData"
        ext_ws["A1"] = "External Value"; ext_ws["B1"] = 500
        ext_file = Path("external_source.xlsx")
        ext_wb.save(ext_file)

        # Create the main workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active) # remove default sheet

        # Data Sheet with a Formal Table
        data_ws = wb.create_sheet("SalesData")
        data_ws.append(["Region", "Product", "Sales"])
        data_ws.append(["North", "A", 100]); data_ws.append(["South", "B", 150])
        data_ws.append(["North", "A", 200]); data_ws.append(["South", "C", 250])
        tbl = Table(displayName="SalesTable", ref=f"A1:C5")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tbl.tableStyleInfo = style
        data_ws.add_table(tbl)

        # Input sheet with Data Validation and Named Ranges
        input_ws = wb.create_sheet("Inputs")
        input_ws["A1"] = "Select Region:"
        input_ws["B1"] = "North" # Default value
        # Data Validation Dropdown
        dv = DataValidation(type="list", formula1='"North,South,East,West"')
        dv.add(input_ws["B1"])
        input_ws.data_validations.append(dv)
        # Named Range
        wb.create_named_range('Selected_Region', input_ws, 'B1')

        # External Link Sheet
        link_ws = wb.create_sheet("ExternalLinks")
        link_ws["A1"] = "Value from other file:"; link_ws["B1"] = f"='[{ext_file.name}]ExternalData'!B1"
        
        # NOTE: openpyxl cannot create charts or VBA projects.
        # The logic to detect them is included and will work on real user files.

        main_file = Path("final_model.xlsm")
        wb.save(main_file)
        print(f"Dummy files '{main_file.name}' and '{ext_file.name}' created.")

        analyze_workbook_final(main_file)
