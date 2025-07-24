#!/usr/bin/env python3
"""
Test File Generator for Excel Parser

Creates a comprehensive set of Excel files with varying complexity to test
the excel parser's capabilities.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from pathlib import Path
import random
from typing import Dict, List, Any

def create_simple_model() -> Workbook:
    """Create a simple financial model with basic calculations."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Simple Model"
    
    # Basic assumptions
    ws["A1"] = "Simple Financial Model"
    ws["A3"] = "Assumptions"
    ws["A4"] = "Revenue"
    ws["B4"] = 1000
    ws["A5"] = "Costs"
    ws["B5"] = 600
    ws["A6"] = "Profit"
    ws["B6"] = "=B4-B5"
    
    return wb

def create_intermediate_model() -> Workbook:
    """Create an intermediate model with multiple sheets and VLOOKUPs."""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Data sheet
    data_ws = wb.create_sheet("Data")
    data_ws["A1"] = "Product ID"
    data_ws["B1"] = "Product Name"
    data_ws["C1"] = "Price"
    data_ws["D1"] = "Category"
    
    products = [
        (101, "Widget A", 25.50, "Electronics"),
        (102, "Widget B", 15.75, "Electronics"),
        (103, "Tool X", 45.00, "Hardware"),
        (104, "Tool Y", 32.25, "Hardware"),
        (105, "Service Z", 100.00, "Services")
    ]
    
    for i, (pid, name, price, cat) in enumerate(products, 2):
        data_ws[f"A{i}"] = pid
        data_ws[f"B{i}"] = name
        data_ws[f"C{i}"] = price
        data_ws[f"D{i}"] = cat
    
    # Create formal table
    tbl = Table(displayName="ProductTable", ref="A1:D6")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tbl.tableStyleInfo = style
    data_ws.add_table(tbl)
    
    # Analysis sheet
    analysis_ws = wb.create_sheet("Analysis")
    analysis_ws["A1"] = "Sales Analysis"
    analysis_ws["A3"] = "Order ID"
    analysis_ws["B3"] = "Product ID"
    analysis_ws["C3"] = "Quantity"
    analysis_ws["D3"] = "Product Name"
    analysis_ws["E3"] = "Unit Price"
    analysis_ws["F3"] = "Total"
    
    orders = [
        (1, 101, 5),
        (2, 103, 2),
        (3, 102, 10),
        (4, 105, 1),
        (5, 104, 3)
    ]
    
    for i, (order_id, prod_id, qty) in enumerate(orders, 4):
        analysis_ws[f"A{i}"] = order_id
        analysis_ws[f"B{i}"] = prod_id
        analysis_ws[f"C{i}"] = qty
        # VLOOKUP for product name
        analysis_ws[f"D{i}"] = f'=VLOOKUP(B{i},Data!A1:D6,2,FALSE)'
        # VLOOKUP for unit price
        analysis_ws[f"E{i}"] = f'=VLOOKUP(B{i},Data!A1:D6,3,FALSE)'
        # Calculate total
        analysis_ws[f"F{i}"] = f'=C{i}*E{i}'
    
    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws["A1"] = "Sales Summary"
    summary_ws["A3"] = "Total Revenue"
    summary_ws["B3"] = "=SUM(Analysis!F4:F8)"
    summary_ws["A4"] = "Average Order Value"
    summary_ws["B4"] = "=B3/COUNT(Analysis!A4:A8)"
    
    return wb

def create_advanced_model() -> Workbook:
    """Create an advanced model with external links, charts, and complex formulas."""
    wb = Workbook()
    wb.remove(wb.active)
    
    # External data reference (simulated)
    external_ws = wb.create_sheet("External_Data")
    external_ws["A1"] = "External Market Data"
    external_ws["A2"] = "Month"
    external_ws["B2"] = "Market Growth"
    external_ws["C2"] = "Inflation Rate"
    
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]
    for i, month in enumerate(months, 3):
        external_ws[f"A{i}"] = month
        external_ws[f"B{i}"] = round(random.uniform(0.02, 0.08), 4)
        external_ws[f"C{i}"] = round(random.uniform(0.01, 0.04), 4)
    
    # Main model sheet
    main_ws = wb.create_sheet("Main_Model")
    main_ws["A1"] = "Advanced Financial Model"
    
    # Assumptions with data validation
    main_ws["A3"] = "Model Assumptions"
    main_ws["A4"] = "Base Revenue"
    main_ws["B4"] = 10000
    
    # Data validation for growth rate
    main_ws["A5"] = "Growth Rate"
    main_ws["B5"] = 0.05
    dv = DataValidation(type="decimal", operator="between", formula1="0", formula2="0.2")
    dv.add(main_ws["B5"])
    main_ws.add_data_validation(dv)
    
    # External link reference
    main_ws["A6"] = "Market Growth (External)"
    main_ws["B6"] = "=External_Data!B3"  # Simulated external link
    
    # Projections
    main_ws["A8"] = "Revenue Projections"
    main_ws["A9"] = "Month"
    main_ws["B9"] = "Revenue"
    main_ws["C9"] = "Growth Adjusted"
    
    for i, month in enumerate(months, 10):
        main_ws[f"A{i}"] = month
        main_ws[f"B{i}"] = f"=B4*(1+B5)^{i-9}"
        main_ws[f"C{i}"] = f"=B{i}*(1+B6)"
    
    # Create chart
    chart = LineChart()
    chart.title = "Revenue Projections"
    chart.x_axis.title = "Month"
    chart.y_axis.title = "Revenue"
    
    data = Reference(main_ws, min_col=2, min_row=9, max_col=3, max_row=15)
    cats = Reference(main_ws, min_col=1, min_row=10, max_row=15)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    main_ws.add_chart(chart, "E2")
    
    return wb

def create_complex_model() -> Workbook:
    """Create a complex model with multiple sheets, pivot-like structures, and advanced features."""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Input sheet with named ranges
    input_ws = wb.create_sheet("Inputs")
    input_ws["A1"] = "Model Inputs"
    
    # Create named ranges
    wb.create_named_range('Discount_Rate', input_ws, 'B3')
    wb.create_named_range('Growth_Rate', input_ws, 'B4')
    wb.create_named_range('Tax_Rate', input_ws, 'B5')
    
    input_ws["A3"] = "Discount Rate"
    input_ws["B3"] = 0.08
    input_ws["A4"] = "Growth Rate"
    input_ws["B4"] = 0.05
    input_ws["A5"] = "Tax Rate"
    input_ws["B5"] = 0.25
    
    # Data validation for discount rate
    dv = DataValidation(type="list", formula1='"0.05,0.06,0.07,0.08,0.09,0.10"')
    dv.add(input_ws["B3"])
    input_ws.add_data_validation(dv)
    
    # Historical data sheet
    hist_ws = wb.create_sheet("Historical_Data")
    hist_ws["A1"] = "Historical Financial Data"
    
    # Create a formal table for historical data
    headers = ["Year", "Revenue", "COGS", "Operating_Expenses", "EBIT", "Taxes", "Net_Income"]
    for i, header in enumerate(headers, 1):
        hist_ws.cell(row=2, column=i, value=header)
    
    # Add historical data
    years = [2020, 2021, 2022, 2023]
    data = [
        [1000000, 600000, 200000, 200000, 50000, 12500, 37500],
        [1100000, 650000, 220000, 230000, 55000, 13750, 41250],
        [1210000, 715000, 242000, 253000, 60500, 15125, 45375],
        [1331000, 786500, 266200, 278300, 66550, 16638, 49913]
    ]
    
    for i, (year, row_data) in enumerate(zip(years, data), 3):
        hist_ws.cell(row=i, column=1, value=year)
        for j, value in enumerate(row_data, 2):
            hist_ws.cell(row=i, column=j, value=value)
    
    # Create formal table
    tbl = Table(displayName="HistoricalTable", ref="A2:G6")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tbl.tableStyleInfo = style
    hist_ws.add_table(tbl)
    
    # Calculations sheet
    calc_ws = wb.create_sheet("Calculations")
    calc_ws["A1"] = "Financial Calculations"
    
    # Projections using historical data and assumptions
    calc_ws["A3"] = "Projections"
    calc_ws["A4"] = "Year"
    calc_ws["B4"] = "Revenue"
    calc_ws["C4"] = "COGS"
    calc_ws["D4"] = "Gross_Margin"
    calc_ws["E4"] = "Operating_Expenses"
    calc_ws["F4"] = "EBIT"
    calc_ws["G4"] = "Taxes"
    calc_ws["H4"] = "Net_Income"
    calc_ws["I4"] = "NPV"
    
    # Project future years
    for i, year in enumerate(range(2024, 2029), 5):
        calc_ws[f"A{i}"] = year
        
        # Revenue projection using growth rate
        calc_ws[f"B{i}"] = f"=Historical_Data!B6*(1+Growth_Rate)^{i-5}"
        
        # COGS as percentage of revenue (using historical average)
        calc_ws[f"C{i}"] = f"=B{i}*AVERAGE(Historical_Data!C3:C6/Historical_Data!B3:B6)"
        
        # Gross margin
        calc_ws[f"D{i}"] = f"=B{i}-C{i}"
        
        # Operating expenses (using historical average percentage)
        calc_ws[f"E{i}"] = f"=B{i}*AVERAGE(Historical_Data!D3:D6/Historical_Data!B3:B6)"
        
        # EBIT
        calc_ws[f"F{i}"] = f"=D{i}-E{i}"
        
        # Taxes
        calc_ws[f"G{i}"] = f"=F{i}*Tax_Rate"
        
        # Net Income
        calc_ws[f"H{i}"] = f"=F{i}-G{i}"
        
        # NPV calculation
        calc_ws[f"I{i}"] = f"=H{i}/(1+Discount_Rate)^{i-5}"
    
    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws["A1"] = "Model Summary"
    
    summary_ws["A3"] = "Key Metrics"
    summary_ws["A4"] = "Total NPV (5 years)"
    summary_ws["B4"] = "=SUM(Calculations!I5:I9)"
    
    summary_ws["A5"] = "Average Annual Growth"
    summary_ws["B5"] = "=Growth_Rate"
    
    summary_ws["A6"] = "Payback Period (years)"
    summary_ws["B6"] = "=MATCH(0,Calculations!I5:I9,1)"
    
    # Create a chart
    chart = BarChart()
    chart.title = "Projected Net Income"
    chart.x_axis.title = "Year"
    chart.y_axis.title = "Net Income"
    
    data = Reference(calc_ws, min_col=8, min_row=4, max_col=8, max_row=9)
    cats = Reference(calc_ws, min_col=1, min_row=5, max_row=9)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    summary_ws.add_chart(chart, "D2")
    
    return wb

def create_enterprise_model() -> Workbook:
    """Create an enterprise-level model with multiple business units, consolidation, and complex scenarios."""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Business Unit 1 - Manufacturing
    mfg_ws = wb.create_sheet("Manufacturing")
    mfg_ws["A1"] = "Manufacturing Division"
    
    # Production data
    mfg_ws["A3"] = "Production Metrics"
    mfg_ws["A4"] = "Month"
    mfg_ws["B4"] = "Units_Produced"
    mfg_ws["C4"] = "Unit_Cost"
    mfg_ws["D4"] = "Total_Cost"
    mfg_ws["E4"] = "Efficiency_Rate"
    
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]
    for i, month in enumerate(months, 5):
        mfg_ws[f"A{i}"] = month
        mfg_ws[f"B{i}"] = random.randint(1000, 5000)
        mfg_ws[f"C{i}"] = round(random.uniform(50, 100), 2)
        mfg_ws[f"D{i}"] = f"=B{i}*C{i}"
        mfg_ws[f"E{i}"] = round(random.uniform(0.85, 0.98), 3)
    
    # Business Unit 2 - Sales
    sales_ws = wb.create_sheet("Sales")
    sales_ws["A1"] = "Sales Division"
    
    sales_ws["A3"] = "Sales Performance"
    sales_ws["A4"] = "Month"
    sales_ws["B4"] = "Units_Sold"
    sales_ws["C4"] = "Unit_Price"
    sales_ws["D4"] = "Revenue"
    sales_ws["E4"] = "Sales_Commission"
    
    for i, month in enumerate(months, 5):
        sales_ws[f"A{i}"] = month
        sales_ws[f"B{i}"] = random.randint(800, 4500)
        sales_ws[f"C{i}"] = round(random.uniform(120, 200), 2)
        sales_ws[f"D{i}"] = f"=B{i}*C{i}"
        sales_ws[f"E{i}"] = f"=D{i}*0.05"  # 5% commission
    
    # Consolidation sheet
    consol_ws = wb.create_sheet("Consolidation")
    consol_ws["A1"] = "Consolidated Financials"
    
    consol_ws["A3"] = "Monthly Consolidation"
    consol_ws["A4"] = "Month"
    consol_ws["B4"] = "Revenue"
    consol_ws["C4"] = "COGS"
    consol_ws["D4"] = "Gross_Margin"
    consol_ws["E4"] = "Operating_Expenses"
    consol_ws["F4"] = "EBIT"
    consol_ws["G4"] = "Taxes"
    consol_ws["H4"] = "Net_Income"
    
    for i, month in enumerate(months, 5):
        consol_ws[f"A{i}"] = month
        # Revenue from Sales
        consol_ws[f"B{i}"] = f"=Sales!D{i}"
        # COGS from Manufacturing
        consol_ws[f"C{i}"] = f"=Manufacturing!D{i}"
        # Gross Margin
        consol_ws[f"D{i}"] = f"=B{i}-C{i}"
        # Operating Expenses (assumed)
        consol_ws[f"E{i}"] = f"=B{i}*0.15"  # 15% of revenue
        # EBIT
        consol_ws[f"F{i}"] = f"=D{i}-E{i}"
        # Taxes
        consol_ws[f"G{i}"] = f"=F{i}*0.25"  # 25% tax rate
        # Net Income
        consol_ws[f"H{i}"] = f"=F{i}-G{i}"
    
    # Scenarios sheet
    scenario_ws = wb.create_sheet("Scenarios")
    scenario_ws["A1"] = "Scenario Analysis"
    
    # Scenario inputs
    scenario_ws["A3"] = "Scenario Parameters"
    scenario_ws["A4"] = "Growth Rate"
    scenario_ws["B4"] = 0.05
    scenario_ws["A5"] = "Price Increase"
    scenario_ws["B5"] = 0.02
    scenario_ws["A6"] = "Cost Reduction"
    scenario_ws["B6"] = 0.03
    
    # Scenario results
    scenario_ws["A8"] = "Scenario Results"
    scenario_ws["A9"] = "Metric"
    scenario_ws["B9"] = "Base Case"
    scenario_ws["C9"] = "Optimistic"
    scenario_ws["D9"] = "Pessimistic"
    
    metrics = [
        ("Total Revenue", "=SUM(Consolidation!B5:B10)", "=B10*(1+B4)", "=B10*(1-B4)"),
        ("Total COGS", "=SUM(Consolidation!C5:C10)", "=C10*(1-B6)", "=C10*(1+B6)"),
        ("Gross Margin", "=B10-C10", "=D10-E10", "=F10-G10"),
        ("Net Income", "=SUM(Consolidation!H5:H10)", "=H10*(1+B4+B5-B6)", "=H10*(1-B4-B5+B6)")
    ]
    
    for i, (metric, base, opt, pes) in enumerate(metrics, 10):
        scenario_ws[f"A{i}"] = metric
        scenario_ws[f"B{i}"] = base
        scenario_ws[f"C{i}"] = opt
        scenario_ws[f"D{i}"] = pes
    
    # Dashboard sheet
    dashboard_ws = wb.create_sheet("Dashboard")
    dashboard_ws["A1"] = "Executive Dashboard"
    
    # Key metrics
    dashboard_ws["A3"] = "Key Performance Indicators"
    dashboard_ws["A4"] = "Total Revenue (6 months)"
    dashboard_ws["B4"] = "=SUM(Consolidation!B5:B10)"
    
    dashboard_ws["A5"] = "Total Net Income (6 months)"
    dashboard_ws["B5"] = "=SUM(Consolidation!H5:H10)"
    
    dashboard_ws["A6"] = "Average Monthly Growth"
    dashboard_ws["B6"] = "=AVERAGE(Consolidation!B5:B10)/AVERAGE(Consolidation!B5:B10)-1"
    
    dashboard_ws["A7"] = "Profit Margin"
    dashboard_ws["B7"] = "=B5/B4"
    
    # Create multiple charts
    # Revenue chart
    revenue_chart = LineChart()
    revenue_chart.title = "Monthly Revenue Trend"
    data = Reference(consol_ws, min_col=2, min_row=4, max_col=2, max_row=10)
    cats = Reference(consol_ws, min_col=1, min_row=5, max_row=10)
    revenue_chart.add_data(data, titles_from_data=True)
    revenue_chart.set_categories(cats)
    dashboard_ws.add_chart(revenue_chart, "D2")
    
    # Profit chart
    profit_chart = BarChart()
    profit_chart.title = "Monthly Net Income"
    data = Reference(consol_ws, min_col=8, min_row=4, max_col=8, max_row=10)
    profit_chart.add_data(data, titles_from_data=True)
    profit_chart.set_categories(cats)
    dashboard_ws.add_chart(profit_chart, "D8")
    
    return wb

def create_test_files():
    """Create all test files in the excel_files directory."""
    test_dir = Path("excel_files")
    test_dir.mkdir(exist_ok=True)
    
    # Create test files with varying complexity
    test_cases = [
        ("simple_model.xlsx", create_simple_model, "Basic financial model with simple calculations"),
        ("intermediate_model.xlsx", create_intermediate_model, "Model with VLOOKUPs and multiple sheets"),
        ("advanced_model.xlsx", create_advanced_model, "Advanced model with external links and charts"),
        ("complex_model.xlsx", create_complex_model, "Complex model with named ranges and projections"),
        ("enterprise_model.xlsx", create_enterprise_model, "Enterprise model with multiple business units")
    ]
    
    print("Creating test files...")
    print("=" * 50)
    
    for filename, creator_func, description in test_cases:
        print(f"Creating {filename}: {description}")
        wb = creator_func()
        filepath = test_dir / filename
        wb.save(filepath)
        print(f"  ✓ Saved to {filepath}")
    
    print("\n" + "=" * 50)
    print("Test files created successfully!")
    print(f"Location: {test_dir.absolute()}")
    print("\nTest files created:")
    
    for filename, _, description in test_cases:
        print(f"  • {filename} - {description}")

if __name__ == "__main__":
    create_test_files() 