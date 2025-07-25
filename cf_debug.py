import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'].value = 1
ws['A2'].value = 2
ws.conditional_formatting.add('A1:A2', CellIsRule(operator='equal', formula=['1'], fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')))
ws.conditional_formatting.add('A2', CellIsRule(operator='equal', formula=['2'], fill=PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')))

print('cf._cf_rules:', ws.conditional_formatting._cf_rules)
for cf_range, rule_list in ws.conditional_formatting._cf_rules.items():
    print('cf_range:', cf_range, 'rules:', rule_list)
    for rule in rule_list:
        print('rule.sqref:', getattr(rule, 'sqref', None), 'rule.type:', getattr(rule, 'type', None), 'rule.dxf:', getattr(rule, 'dxf', None)) 