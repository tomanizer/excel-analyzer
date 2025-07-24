#!/usr/bin/env python3
"""
Excel Extractor

Extracts all data, text, formulas, and structural information from Excel files
and produces comprehensive markdown output suitable for LLM analysis.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.datavalidation import DataValidation
# Chart types are imported individually as needed
from pathlib import Path
from typing import Dict, List, Any, Set, Tuple
import json
import re
from datetime import datetime

class ExcelExtractor:
    """Comprehensive Excel file extractor for LLM analysis."""
    
    def __init__(self, file_path: Path):
        """Initialize the extractor with an Excel file path."""
        self.file_path = file_path
        self.workbook = None
        self.extracted_data = {
            'metadata': {},
            'sheets': {},
            'global_features': {},
            'relationships': {},
            'summary': {}
        }
    
    def extract_all(self) -> Dict[str, Any]:
        """Extract all information from the Excel file."""
        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {self.file_path}")
        
        print(f"Extracting data from: {self.file_path.name}")
        
        # Load workbook
        self.workbook = openpyxl.load_workbook(
            self.file_path, 
            data_only=False,  # Keep formulas
            keep_vba=True
        )
        
        # Extract metadata
        self._extract_metadata()
        
        # Extract global features
        self._extract_global_features()
        
        # Extract sheet-level data
        self._extract_sheets()
        
        # Extract relationships
        self._extract_relationships()
        
        # Generate summary
        self._generate_summary()
        
        return self.extracted_data
    
    def _extract_metadata(self):
        """Extract file metadata."""
        self.extracted_data['metadata'] = {
            'filename': self.file_path.name,
            'file_size': self.file_path.stat().st_size,
            'file_size_kb': round(self.file_path.stat().st_size / 1024, 2),
            'last_modified': datetime.fromtimestamp(self.file_path.stat().st_mtime).isoformat(),
            'file_extension': self.file_path.suffix,
            'has_vba': self.file_path.suffix == '.xlsm',
            'sheet_count': len(self.workbook.sheetnames),
            'sheet_names': self.workbook.sheetnames
        }
    
    def _extract_global_features(self):
        """Extract global workbook features."""
        global_features = {}
        
        # Named ranges
        named_ranges = {}
        for name, defined_name in self.workbook.defined_names.items():
            destinations = []
            for dest in defined_name.destinations:
                sheet_name, coord = dest
                destinations.append(f"{sheet_name}!{coord}")
            named_ranges[name] = destinations
        
        global_features['named_ranges'] = named_ranges
        
        # External links
        external_links = []
        if hasattr(self.workbook, 'external_links'):
            for link in self.workbook.external_links:
                external_links.append(str(link.Target))
        
        global_features['external_links'] = external_links
        
        # Properties
        properties = {}
        if self.workbook.properties:
            props = self.workbook.properties
            try:
                if hasattr(props, 'title') and props.title:
                    properties['title'] = props.title
                if hasattr(props, 'creator') and props.creator:
                    properties['author'] = props.creator
                if hasattr(props, 'subject') and props.subject:
                    properties['subject'] = props.subject
                if hasattr(props, 'keywords') and props.keywords:
                    properties['keywords'] = props.keywords
                if hasattr(props, 'comments') and props.comments:
                    properties['comments'] = props.comments
            except:
                pass  # Skip properties if there are issues
        
        global_features['properties'] = properties
        
        self.extracted_data['global_features'] = global_features
    
    def _extract_sheets(self):
        """Extract data from all sheets."""
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            self.extracted_data['sheets'][sheet_name] = self._extract_sheet_data(sheet)
    
    def _extract_sheet_data(self, sheet) -> Dict[str, Any]:
        """Extract comprehensive data from a single sheet."""
        sheet_data = {
            'dimensions': {},
            'data': {},
            'formulas': {},
            'tables': [],
            'charts': [],
            'data_validations': [],
            'merged_cells': [],
            'styles': {},
            'summary': {}
        }
        
        # Sheet dimensions
        sheet_data['dimensions'] = {
            'max_row': sheet.max_row,
            'max_column': sheet.max_column,
            'max_column_letter': openpyxl.utils.cell.get_column_letter(sheet.max_column) if sheet.max_column else 'A'
        }
        
        # Extract all cell data
        self._extract_cell_data(sheet, sheet_data)
        
        # Extract formal tables
        self._extract_tables(sheet, sheet_data)
        
        # Extract charts
        self._extract_charts(sheet, sheet_data)
        
        # Extract data validations
        self._extract_data_validations(sheet, sheet_data)
        
        # Extract merged cells
        self._extract_merged_cells(sheet, sheet_data)
        
        # Extract styles
        self._extract_styles(sheet, sheet_data)
        
        # Generate sheet summary
        self._generate_sheet_summary(sheet_data)
        
        return sheet_data
    
    def _extract_cell_data(self, sheet, sheet_data: Dict[str, Any]):
        """Extract all cell data, values, and formulas."""
        data = {}
        formulas = {}
        
        for row in sheet.iter_rows():
            for cell in row:
                is_formula = cell.data_type == 'f'
                if cell.value is not None or is_formula:
                    coord = cell.coordinate
                    # Extract value
                    if cell.value is not None:
                        data[coord] = {
                            'value': cell.value,
                            'data_type': type(cell.value).__name__,
                            'is_formula': is_formula
                        }
                    # Extract formula
                    if is_formula:
                        formulas[coord] = {
                            'formula': str(cell.value),
                            'calculated_value': cell.value
                        }
        
        sheet_data['data'] = data
        sheet_data['formulas'] = formulas
    
    def _extract_tables(self, sheet, sheet_data: Dict[str, Any]):
        """Extract formal Excel tables."""
        tables = []
        
        for table in sheet.tables.values():
            table_info = {
                'name': table.displayName,
                'range': table.ref,
                'style': table.tableStyleInfo.name if table.tableStyleInfo else None,
                'show_first_column': table.tableStyleInfo.showFirstColumn if table.tableStyleInfo else None,
                'show_last_column': table.tableStyleInfo.showLastColumn if table.tableStyleInfo else None,
                'show_row_stripes': table.tableStyleInfo.showRowStripes if table.tableStyleInfo else None,
                'show_column_stripes': table.tableStyleInfo.showColumnStripes if table.tableStyleInfo else None
            }
            
            # Extract table data
            table_data = {}
            min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(table.ref)
            
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    coord = cell.coordinate
                    is_formula = cell.data_type == 'f'
                    if cell.value is not None:
                        table_data[coord] = {
                            'value': cell.value,
                            'formula': str(cell.value) if is_formula else None
                        }
            
            table_info['data'] = table_data
            tables.append(table_info)
        
        sheet_data['tables'] = tables
    
    def _extract_charts(self, sheet, sheet_data: Dict[str, Any]):
        """Extract chart information."""
        charts = []
        
        for chart in sheet._charts:
            chart_info = {
                'type': type(chart).__name__,
                'title': str(chart.title) if chart.title else None,
                'x_axis_title': str(chart.x_axis.title) if chart.x_axis.title else None,
                'y_axis_title': str(chart.y_axis.title) if chart.y_axis.title else None
            }
            
            # Try to extract data sources
            try:
                if hasattr(chart, 'data') and chart.data:
                    chart_info['data_sources'] = str(chart.data)
            except:
                chart_info['data_sources'] = "Unable to extract"
            
            charts.append(chart_info)
        
        sheet_data['charts'] = charts
    
    def _extract_data_validations(self, sheet, sheet_data: Dict[str, Any]):
        """Extract data validation rules."""
        validations = []
        
        for validation in sheet.data_validations.dataValidation:
            validation_info = {
                'range': str(validation.sqref),
                'type': validation.type,
                'operator': validation.operator,
                'formula1': validation.formula1,
                'formula2': validation.formula2,
                'allow_blank': validation.allowBlank,
                'show_error_message': validation.showErrorMessage,
                'error_title': validation.errorTitle,
                'error_message': validation.errorMessage
            }
            validations.append(validation_info)
        
        sheet_data['data_validations'] = validations
    
    def _extract_merged_cells(self, sheet, sheet_data: Dict[str, Any]):
        """Extract merged cell information."""
        merged_cells = []
        
        for merged_range in sheet.merged_cells.ranges:
            merged_info = {
                'range': str(merged_range),
                'top_left': merged_range.top_left.coordinate,
                'bottom_right': merged_range.bottom_right.coordinate
            }
            merged_cells.append(merged_info)
        
        sheet_data['merged_cells'] = merged_cells
    
    def _extract_styles(self, sheet, sheet_data: Dict[str, Any]):
        """Extract basic style information."""
        styles = {
            'cells_with_styles': 0,
            'cells_with_fonts': 0,
            'cells_with_fills': 0,
            'cells_with_borders': 0
        }
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    if cell.font:
                        styles['cells_with_fonts'] += 1
                    if cell.fill:
                        styles['cells_with_fills'] += 1
                    if cell.border:
                        styles['cells_with_borders'] += 1
                    if cell.font or cell.fill or cell.border:
                        styles['cells_with_styles'] += 1
        
        sheet_data['styles'] = styles
    
    def _generate_sheet_summary(self, sheet_data: Dict[str, Any]):
        """Generate summary statistics for a sheet."""
        summary = {
            'total_cells_with_data': len(sheet_data['data']),
            'total_formulas': len(sheet_data['formulas']),
            'total_tables': len(sheet_data['tables']),
            'total_charts': len(sheet_data['charts']),
            'total_validations': len(sheet_data['data_validations']),
            'total_merged_cells': len(sheet_data['merged_cells']),
            'data_types': {},
            'formula_functions': {}
        }
        
        # Analyze data types
        for cell_info in sheet_data['data'].values():
            data_type = cell_info['data_type']
            summary['data_types'][data_type] = summary['data_types'].get(data_type, 0) + 1
        
        # Analyze formula functions
        for formula_info in sheet_data['formulas'].values():
            formula = formula_info['formula']
            # Extract function names (basic extraction)
            functions = re.findall(r'([A-Z]+)\s*\(', formula)
            for func in functions:
                summary['formula_functions'][func] = summary['formula_functions'].get(func, 0) + 1
        
        sheet_data['summary'] = summary
    
    def _extract_relationships(self):
        """Extract relationships between sheets and cells."""
        relationships = {
            'cross_sheet_references': [],
            'external_references': [],
            'circular_references': []
        }
        
        # Extract cross-sheet references from formulas
        for sheet_name, sheet_data in self.extracted_data['sheets'].items():
            for coord, formula_info in sheet_data['formulas'].items():
                formula = formula_info['formula']
                
                # Look for cross-sheet references (e.g., Sheet1!A1)
                cross_sheet_refs = re.findall(r"'?([^']+)'?!([A-Z]+\d+)", formula)
                for ref_sheet, ref_cell in cross_sheet_refs:
                    if ref_sheet in self.workbook.sheetnames:
                        relationships['cross_sheet_references'].append({
                            'source_sheet': sheet_name,
                            'source_cell': coord,
                            'target_sheet': ref_sheet,
                            'target_cell': ref_cell,
                            'formula': formula
                        })
        
        self.extracted_data['relationships'] = relationships
    
    def _generate_summary(self):
        """Generate overall workbook summary."""
        summary = {
            'total_sheets': len(self.extracted_data['sheets']),
            'total_cells_with_data': 0,
            'total_formulas': 0,
            'total_tables': 0,
            'total_charts': 0,
            'total_named_ranges': len(self.extracted_data['global_features']['named_ranges']),
            'total_external_links': len(self.extracted_data['global_features']['external_links']),
            'total_cross_sheet_references': len(self.extracted_data['relationships']['cross_sheet_references']),
            'data_types_summary': {},
            'formula_functions_summary': {},
            'complexity_score': 0
        }
        
        # Aggregate data from all sheets
        for sheet_data in self.extracted_data['sheets'].values():
            summary['total_cells_with_data'] += sheet_data['summary']['total_cells_with_data']
            summary['total_formulas'] += sheet_data['summary']['total_formulas']
            summary['total_tables'] += sheet_data['summary']['total_tables']
            summary['total_charts'] += sheet_data['summary']['total_charts']
            
            # Aggregate data types
            for data_type, count in sheet_data['summary']['data_types'].items():
                summary['data_types_summary'][data_type] = summary['data_types_summary'].get(data_type, 0) + count
            
            # Aggregate formula functions
            for func, count in sheet_data['summary']['formula_functions'].items():
                summary['formula_functions_summary'][func] = summary['formula_functions_summary'].get(func, 0) + count
        
        # Calculate complexity score
        complexity_factors = [
            summary['total_sheets'] * 10,
            summary['total_formulas'] * 5,
            summary['total_tables'] * 20,
            summary['total_charts'] * 15,
            summary['total_named_ranges'] * 10,
            summary['total_cross_sheet_references'] * 8,
            summary['total_external_links'] * 25
        ]
        summary['complexity_score'] = sum(complexity_factors)
        
        self.extracted_data['summary'] = summary
    
    def to_markdown(self) -> str:
        """Convert extracted data to comprehensive markdown format."""
        md_lines = []
        
        # Header
        md_lines.append(f"# Excel Workbook Analysis: {self.extracted_data['metadata']['filename']}")
        md_lines.append("")
        md_lines.append(f"*Generated on: {datetime.now().isoformat()}*")
        md_lines.append("")
        
        # Executive Summary
        md_lines.append("## 📊 Executive Summary")
        md_lines.append("")
        summary = self.extracted_data['summary']
        md_lines.append(f"- **File Size**: {self.extracted_data['metadata']['file_size_kb']} KB")
        md_lines.append(f"- **Sheets**: {summary['total_sheets']}")
        md_lines.append(f"- **Cells with Data**: {summary['total_cells_with_data']:,}")
        md_lines.append(f"- **Formulas**: {summary['total_formulas']:,}")
        md_lines.append(f"- **Tables**: {summary['total_tables']}")
        md_lines.append(f"- **Charts**: {summary['total_charts']}")
        md_lines.append(f"- **Named Ranges**: {summary['total_named_ranges']}")
        md_lines.append(f"- **Cross-sheet References**: {summary['total_cross_sheet_references']}")
        md_lines.append(f"- **Complexity Score**: {summary['complexity_score']}")
        md_lines.append("")
        
        # Metadata
        md_lines.append("## 📋 File Metadata")
        md_lines.append("")
        metadata = self.extracted_data['metadata']
        md_lines.append(f"- **Filename**: {metadata['filename']}")
        md_lines.append(f"- **File Size**: {metadata['file_size_kb']} KB")
        md_lines.append(f"- **Last Modified**: {metadata['last_modified']}")
        md_lines.append(f"- **File Type**: {metadata['file_extension']}")
        md_lines.append(f"- **VBA Enabled**: {metadata['has_vba']}")
        md_lines.append(f"- **Sheet Count**: {metadata['sheet_count']}")
        md_lines.append("")
        
        # Global Features
        md_lines.append("## 🌐 Global Features")
        md_lines.append("")
        
        # Named Ranges
        named_ranges = self.extracted_data['global_features']['named_ranges']
        if named_ranges:
            md_lines.append("### Named Ranges")
            md_lines.append("")
            for name, destinations in named_ranges.items():
                md_lines.append(f"- **{name}**: {', '.join(destinations)}")
            md_lines.append("")
        
        # External Links
        external_links = self.extracted_data['global_features']['external_links']
        if external_links:
            md_lines.append("### External Links")
            md_lines.append("")
            for link in external_links:
                md_lines.append(f"- {link}")
            md_lines.append("")
        
        # Properties
        properties = self.extracted_data['global_features']['properties']
        if properties:
            md_lines.append("### Document Properties")
            md_lines.append("")
            for key, value in properties.items():
                md_lines.append(f"- **{key.title()}**: {value}")
            md_lines.append("")
        
        # Sheet Analysis
        md_lines.append("## 📄 Sheet Analysis")
        md_lines.append("")
        
        for sheet_name, sheet_data in self.extracted_data['sheets'].items():
            md_lines.append(f"### Sheet: {sheet_name}")
            md_lines.append("")
            
            # Sheet summary
            sheet_summary = sheet_data['summary']
            md_lines.append(f"- **Dimensions**: {sheet_data['dimensions']['max_row']} rows × {sheet_data['dimensions']['max_column']} columns")
            md_lines.append(f"- **Cells with Data**: {sheet_summary['total_cells_with_data']:,}")
            md_lines.append(f"- **Formulas**: {sheet_summary['total_formulas']:,}")
            md_lines.append(f"- **Tables**: {sheet_summary['total_tables']}")
            md_lines.append(f"- **Charts**: {sheet_summary['total_charts']}")
            md_lines.append(f"- **Data Validations**: {sheet_summary['total_validations']}")
            md_lines.append(f"- **Merged Cells**: {sheet_summary['total_merged_cells']}")
            md_lines.append("")
            
            # Tables
            if sheet_data['tables']:
                md_lines.append("#### Formal Tables")
                md_lines.append("")
                for table in sheet_data['tables']:
                    md_lines.append(f"- **{table['name']}** (Range: {table['range']})")
                    if table['style']:
                        md_lines.append(f"  - Style: {table['style']}")
                    md_lines.append("")
            
            # Charts
            if sheet_data['charts']:
                md_lines.append("#### Charts")
                md_lines.append("")
                for chart in sheet_data['charts']:
                    md_lines.append(f"- **{chart['title'] or 'Untitled'}** ({chart['type']})")
                    if chart['x_axis_title']:
                        md_lines.append(f"  - X-Axis: {chart['x_axis_title']}")
                    if chart['y_axis_title']:
                        md_lines.append(f"  - Y-Axis: {chart['y_axis_title']}")
                    md_lines.append("")
            
            # Data Validations
            if sheet_data['data_validations']:
                md_lines.append("#### Data Validation Rules")
                md_lines.append("")
                for validation in sheet_data['data_validations']:
                    md_lines.append(f"- **Range**: {validation['range']}")
                    md_lines.append(f"  - Type: {validation['type']}")
                    if validation['formula1']:
                        md_lines.append(f"  - Formula: {validation['formula1']}")
                    md_lines.append("")
            
            # Sample Data (first 10 cells with data)
            if sheet_data['data']:
                md_lines.append("#### Sample Data")
                md_lines.append("")
                md_lines.append("| Cell | Value | Type | Formula |")
                md_lines.append("|------|-------|------|---------|")
                
                count = 0
                for coord, cell_info in sheet_data['data'].items():
                    if count >= 10:
                        break
                    
                    value = str(cell_info['value'])[:50]  # Truncate long values
                    data_type = cell_info['data_type']
                    formula = "Yes" if cell_info['is_formula'] else "No"
                    
                    md_lines.append(f"| {coord} | {value} | {data_type} | {formula} |")
                    count += 1
                
                if len(sheet_data['data']) > 10:
                    md_lines.append(f"| ... | ... | ... | ... | *(showing 10 of {len(sheet_data['data'])} cells)* |")
                md_lines.append("")
            
            # Formula Analysis
            if sheet_data['formulas']:
                md_lines.append("#### Formula Analysis")
                md_lines.append("")
                
                # Most common functions
                if sheet_summary['formula_functions']:
                    md_lines.append("**Most Common Functions:**")
                    md_lines.append("")
                    sorted_functions = sorted(sheet_summary['formula_functions'].items(), 
                                           key=lambda x: x[1], reverse=True)[:5]
                    for func, count in sorted_functions:
                        md_lines.append(f"- {func}: {count} occurrences")
                    md_lines.append("")
                
                # Sample formulas
                md_lines.append("**Sample Formulas:**")
                md_lines.append("")
                count = 0
                for coord, formula_info in sheet_data['formulas'].items():
                    if count >= 5:
                        break
                    formula = formula_info['formula'][:100]  # Truncate long formulas
                    md_lines.append(f"- **{coord}**: `{formula}`")
                    count += 1
                
                if len(sheet_data['formulas']) > 5:
                    md_lines.append(f"- ... *(showing 5 of {len(sheet_data['formulas'])} formulas)*")
                md_lines.append("")
        
        # Relationships
        md_lines.append("## 🔗 Relationships")
        md_lines.append("")
        
        relationships = self.extracted_data['relationships']
        
        if relationships['cross_sheet_references']:
            md_lines.append("### Cross-Sheet References")
            md_lines.append("")
            md_lines.append("| Source | Target | Formula |")
            md_lines.append("|--------|--------|---------|")
            
            for ref in relationships['cross_sheet_references'][:10]:  # Limit to first 10
                source = f"{ref['source_sheet']}!{ref['source_cell']}"
                target = f"{ref['target_sheet']}!{ref['target_cell']}"
                formula = ref['formula'][:50]  # Truncate
                md_lines.append(f"| {source} | {target} | `{formula}` |")
            
            if len(relationships['cross_sheet_references']) > 10:
                md_lines.append(f"| ... | ... | ... | *(showing 10 of {len(relationships['cross_sheet_references'])} references)* |")
            md_lines.append("")
        
        # Data Type Analysis
        md_lines.append("## 📈 Data Type Analysis")
        md_lines.append("")
        
        summary = self.extracted_data['summary']
        if summary['data_types_summary']:
            md_lines.append("### Data Types Distribution")
            md_lines.append("")
            sorted_types = sorted(summary['data_types_summary'].items(), 
                                key=lambda x: x[1], reverse=True)
            for data_type, count in sorted_types:
                percentage = (count / summary['total_cells_with_data']) * 100
                md_lines.append(f"- **{data_type}**: {count:,} cells ({percentage:.1f}%)")
            md_lines.append("")
        
        if summary['formula_functions_summary']:
            md_lines.append("### Formula Functions Distribution")
            md_lines.append("")
            sorted_functions = sorted(summary['formula_functions_summary'].items(), 
                                   key=lambda x: x[1], reverse=True)
            for func, count in sorted_functions:
                percentage = (count / summary['total_formulas']) * 100
                md_lines.append(f"- **{func}**: {count} uses ({percentage:.1f}%)")
            md_lines.append("")
        
        # Recommendations
        md_lines.append("## 💡 Analysis & Recommendations")
        md_lines.append("")
        
        complexity_score = summary['complexity_score']
        if complexity_score < 100:
            md_lines.append("**Complexity Level**: Simple")
            md_lines.append("- This is a straightforward Excel file suitable for basic analysis")
        elif complexity_score < 500:
            md_lines.append("**Complexity Level**: Moderate")
            md_lines.append("- This file has moderate complexity with some advanced features")
        else:
            md_lines.append("**Complexity Level**: Complex")
            md_lines.append("- This is a complex Excel file with many advanced features")
        
        md_lines.append("")
        
        # Key observations
        md_lines.append("### Key Observations")
        md_lines.append("")
        
        if summary['total_formulas'] > 0:
            md_lines.append(f"- Contains {summary['total_formulas']:,} formulas indicating active calculations")
        
        if summary['total_tables'] > 0:
            md_lines.append(f"- Uses {summary['total_tables']} formal Excel tables for structured data")
        
        if summary['total_charts'] > 0:
            md_lines.append(f"- Includes {summary['total_charts']} charts for data visualization")
        
        if summary['total_cross_sheet_references'] > 0:
            md_lines.append(f"- Has {summary['total_cross_sheet_references']} cross-sheet references showing data relationships")
        
        if summary['total_named_ranges'] > 0:
            md_lines.append(f"- Uses {summary['total_named_ranges']} named ranges for better formula readability")
        
        if summary['total_external_links'] > 0:
            md_lines.append(f"- Contains {summary['total_external_links']} external links to other files")
        
        md_lines.append("")
        
        return "\n".join(md_lines)
    
    def save_markdown(self, output_path: Path = None) -> Path:
        """Save the markdown output to a file."""
        if output_path is None:
            output_path = self.file_path.with_suffix('.md')
        
        markdown_content = self.to_markdown()
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(markdown_content)
        
        print(f"Markdown saved to: {output_path}")
        return output_path
    
    def save_json(self, output_path: Path = None) -> Path:
        """Save the raw extracted data as JSON."""
        if output_path is None:
            output_path = self.file_path.with_suffix('.json')
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.extracted_data, f, indent=2, default=str)
        
        print(f"JSON data saved to: {output_path}")
        return output_path


def extract_excel_to_markdown(file_path: Path, output_dir: Path = None) -> Tuple[Path, Path]:
    """
    Extract Excel file to markdown and JSON formats.
    
    Args:
        file_path: Path to the Excel file
        output_dir: Directory to save output files (default: same as input file)
    
    Returns:
        Tuple of (markdown_path, json_path)
    """
    if output_dir is None:
        output_dir = file_path.parent
    
    extractor = ExcelExtractor(file_path)
    extractor.extract_all()
    
    # Save outputs
    markdown_path = output_dir / f"{file_path.stem}_analysis.md"
    json_path = output_dir / f"{file_path.stem}_data.json"
    
    extractor.save_markdown(markdown_path)
    extractor.save_json(json_path)
    
    return markdown_path, json_path


if __name__ == "__main__":
    # Example usage
    import sys
    
    if len(sys.argv) > 1:
        file_path = Path(sys.argv[1])
        if file_path.exists():
            markdown_path, json_path = extract_excel_to_markdown(file_path)
            print(f"\nExtraction complete!")
            print(f"Markdown: {markdown_path}")
            print(f"JSON: {json_path}")
        else:
            print(f"File not found: {file_path}")
    else:
        print("Usage: python excel_extractor.py <excel_file_path>")
        print("\nExample: python excel_extractor.py test_files/complex_model.xlsx") 