#!/usr/bin/env python3
"""
Excel Error Sniffer - Detect and analyze common Excel errors and issues.

This module provides comprehensive error detection for Excel files including:
- Formula errors (#N/A, #VALUE!, #REF!, etc.)
- Circular references
- Broken links and references
- Data validation issues
- Performance problems
- Structural issues
- Compatibility warnings
"""

import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import warnings
from datetime import datetime
import json

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.external_reference import ExternalReference

logger = logging.getLogger(__name__)


class ExcelErrorSniffer:
    """
    Comprehensive Excel error detection and analysis tool.
    
    Detects various types of errors and issues in Excel files including:
    - Formula errors and calculation issues
    - Circular references
    - Broken links and external references
    - Data validation problems
    - Performance issues
    - Structural problems
    - Compatibility warnings
    """
    
    # Common Excel error values
    EXCEL_ERRORS = {
        '#N/A': 'Not Available - Value not available',
        '#VALUE!': 'Value Error - Wrong type of argument',
        '#REF!': 'Reference Error - Invalid cell reference',
        '#DIV/0!': 'Divide by Zero - Division by zero',
        '#NUM!': 'Number Error - Invalid number',
        '#NAME?': 'Name Error - Invalid name',
        '#NULL!': 'Null Error - Invalid intersection',
        '#SPILL!': 'Spill Error - Array formula overflow',
        '#CALC!': 'Calculation Error - Calculation failed',
        '#UNKNOWN!': 'Unknown Error - Unknown error type'
    }
    
    # Performance thresholds
    PERFORMANCE_THRESHOLDS = {
        'max_cells_per_sheet': 1000000,  # 1M cells
        'max_formulas_per_sheet': 10000,  # 10K formulas
        'max_external_links': 100,  # 100 external links
        'max_named_ranges': 1000,  # 1000 named ranges
        'max_conditional_formats': 100,  # 100 conditional formats
        'max_data_validations': 1000,  # 1000 data validations
    }
    
    def __init__(self, file_path: Path):
        """
        Initialize the Excel Error Sniffer.
        
        Args:
            file_path: Path to the Excel file to analyze
        """
        self.file_path = Path(file_path)
        self.workbook = None
        self.errors = {
            'formula_errors': [],
            'circular_references': [],
            'broken_links': [],
            'data_validation_issues': [],
            'performance_issues': [],
            'structural_issues': [],
            'compatibility_warnings': [],
            'summary': {}
        }
        
    def _load_workbook(self) -> None:
        """Load the Excel workbook safely."""
        try:
            self.workbook = openpyxl.load_workbook(
                self.file_path, 
                data_only=False,  # Keep formulas for error detection
                keep_vba=True
            )
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            raise
    
    def _detect_formula_errors(self) -> List[Dict[str, Any]]:
        """
        Detect formula errors in the workbook.
        
        Returns:
            List of formula errors found
        """
        formula_errors = []
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        # Check for Excel error values
                        cell_value = str(cell.value).strip()
                        if cell_value in self.EXCEL_ERRORS:
                            formula_errors.append({
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'error_type': cell_value,
                                'description': self.EXCEL_ERRORS[cell_value],
                                'formula': cell.formula if cell.formula else None,
                                'severity': 'high'
                            })
                        
                        # Check for potential formula issues
                        elif hasattr(cell, 'formula') and cell.formula and cell.formula.startswith('='):
                            # Check for common formula problems
                            formula = cell.formula.upper()
                            
                            # Check for hardcoded values in formulas
                            if any(op in formula for op in ['+0', '-0', '*1', '/1']):
                                                            formula_errors.append({
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'error_type': 'Inefficient Formula',
                                'description': 'Formula contains unnecessary operations',
                                'formula': cell.formula if hasattr(cell, 'formula') else None,
                                'severity': 'low'
                            })
                            
                            # Check for volatile functions
                            volatile_functions = ['NOW()', 'TODAY()', 'RAND()', 'RANDBETWEEN()', 'OFFSET()', 'INDIRECT()']
                            if any(func in formula for func in volatile_functions):
                                                            formula_errors.append({
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'error_type': 'Volatile Function',
                                'description': f'Formula uses volatile function that recalculates on every change',
                                'formula': cell.formula if hasattr(cell, 'formula') else None,
                                'severity': 'medium'
                            })
        
        return formula_errors
    
    def _detect_circular_references(self) -> List[Dict[str, Any]]:
        """
        Detect circular references in the workbook.
        
        Returns:
            List of circular references found
        """
        circular_refs = []
        
        # Check for circular references in each sheet
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            # This is a simplified check - Excel's circular reference detection
            # is more sophisticated, but we can detect obvious cases
            for row in sheet.iter_rows():
                for cell in row:
                    if hasattr(cell, 'formula') and cell.formula and cell.formula.startswith('='):
                        formula = cell.formula.upper()
                        cell_ref = cell.coordinate.upper()
                        
                        # Check if cell references itself
                        if cell_ref in formula:
                            circular_refs.append({
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'error_type': 'Circular Reference',
                                'description': f'Cell references itself: {cell.formula}',
                                'formula': cell.formula if hasattr(cell, 'formula') else None,
                                'severity': 'high'
                            })
        
        return circular_refs
    
    def _detect_broken_links(self) -> List[Dict[str, Any]]:
        """
        Detect broken external links and references.
        
        Returns:
            List of broken links found
        """
        broken_links = []
        
        # Check external references
        if hasattr(self.workbook, 'external_links') and self.workbook.external_links:
            for link in self.workbook.external_links:
                try:
                    # Try to access the external file
                    external_path = Path(link.target)
                    if not external_path.exists():
                        broken_links.append({
                            'type': 'External File',
                            'target': str(link.target),
                            'description': f'External file not found: {link.target}',
                            'severity': 'high'
                        })
                except Exception as e:
                    broken_links.append({
                        'type': 'External Link',
                        'target': str(link.target),
                        'description': f'Error accessing external link: {e}',
                        'severity': 'medium'
                    })
        
        # Check for broken references in formulas
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            for row in sheet.iter_rows():
                for cell in row:
                    if hasattr(cell, 'formula') and cell.formula and cell.formula.startswith('='):
                        formula = cell.formula
                        
                        # Check for broken sheet references
                        if '!' in formula and not any(sheet_name in formula for sheet_name in self.workbook.sheetnames):
                            broken_links.append({
                                'type': 'Sheet Reference',
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'description': f'Formula references non-existent sheet: {formula}',
                                'formula': formula,
                                'severity': 'high'
                            })
        
        return broken_links
    
    def _detect_data_validation_issues(self) -> List[Dict[str, Any]]:
        """
        Detect issues with data validation rules.
        
        Returns:
            List of data validation issues found
        """
        validation_issues = []
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            # Check data validation rules
            for row in sheet.iter_rows():
                for cell in row:
                    if hasattr(cell, 'data_validation') and cell.data_validation:
                        validation = cell.data_validation
                        
                        # Check for empty validation criteria
                        if not validation.formula1 and not validation.list:
                            validation_issues.append({
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'error_type': 'Empty Validation',
                                'description': 'Data validation rule has no criteria',
                                'severity': 'medium'
                            })
                        
                        # Check for conflicting validation types
                        if validation.formula1 and validation.list:
                            validation_issues.append({
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'error_type': 'Conflicting Validation',
                                'description': 'Data validation has both formula and list criteria',
                                'severity': 'medium'
                            })
        
        return validation_issues
    
    def _detect_performance_issues(self) -> List[Dict[str, Any]]:
        """
        Detect potential performance issues.
        
        Returns:
            List of performance issues found
        """
        performance_issues = []
        
        total_cells = 0
        total_formulas = 0
        total_external_links = len(self.workbook.external_links) if hasattr(self.workbook, 'external_links') and self.workbook.external_links else 0
        total_named_ranges = len(self.workbook.defined_names)
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            # Count cells and formulas
            sheet_cells = 0
            sheet_formulas = 0
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        sheet_cells += 1
                    if hasattr(cell, 'formula') and cell.formula and cell.formula.startswith('='):
                        sheet_formulas += 1
            
            total_cells += sheet_cells
            total_formulas += sheet_formulas
            
            # Check sheet-level performance issues
            if sheet_cells > self.PERFORMANCE_THRESHOLDS['max_cells_per_sheet']:
                performance_issues.append({
                    'sheet': sheet_name,
                    'error_type': 'Large Sheet',
                    'description': f'Sheet has {sheet_cells:,} cells (threshold: {self.PERFORMANCE_THRESHOLDS["max_cells_per_sheet"]:,})',
                    'value': sheet_cells,
                    'threshold': self.PERFORMANCE_THRESHOLDS['max_cells_per_sheet'],
                    'severity': 'medium'
                })
            
            if sheet_formulas > self.PERFORMANCE_THRESHOLDS['max_formulas_per_sheet']:
                performance_issues.append({
                    'sheet': sheet_name,
                    'error_type': 'Many Formulas',
                    'description': f'Sheet has {sheet_formulas:,} formulas (threshold: {self.PERFORMANCE_THRESHOLDS["max_formulas_per_sheet"]:,})',
                    'value': sheet_formulas,
                    'threshold': self.PERFORMANCE_THRESHOLDS['max_formulas_per_sheet'],
                    'severity': 'medium'
                })
        
        # Check workbook-level performance issues
        if total_external_links > self.PERFORMANCE_THRESHOLDS['max_external_links']:
            performance_issues.append({
                'type': 'Workbook',
                'error_type': 'Many External Links',
                'description': f'Workbook has {total_external_links} external links (threshold: {self.PERFORMANCE_THRESHOLDS["max_external_links"]})',
                'value': total_external_links,
                'threshold': self.PERFORMANCE_THRESHOLDS['max_external_links'],
                'severity': 'medium'
            })
        
        if total_named_ranges > self.PERFORMANCE_THRESHOLDS['max_named_ranges']:
            performance_issues.append({
                'type': 'Workbook',
                'error_type': 'Many Named Ranges',
                'description': f'Workbook has {total_named_ranges} named ranges (threshold: {self.PERFORMANCE_THRESHOLDS["max_named_ranges"]})',
                'value': total_named_ranges,
                'threshold': self.PERFORMANCE_THRESHOLDS['max_named_ranges'],
                'severity': 'low'
            })
        
        return performance_issues
    
    def _detect_structural_issues(self) -> List[Dict[str, Any]]:
        """
        Detect structural issues in the workbook.
        
        Returns:
            List of structural issues found
        """
        structural_issues = []
        
        # Check for hidden sheets
        hidden_sheets = [name for name in self.workbook.sheetnames 
                        if self.workbook[name].sheet_state == 'hidden']
        
        if hidden_sheets:
            structural_issues.append({
                'type': 'Workbook',
                'error_type': 'Hidden Sheets',
                'description': f'Workbook contains {len(hidden_sheets)} hidden sheets: {", ".join(hidden_sheets)}',
                'sheets': hidden_sheets,
                'severity': 'low'
            })
        
        # Check for very long sheet names
        for sheet_name in self.workbook.sheetnames:
            if len(sheet_name) > 31:  # Excel's limit
                structural_issues.append({
                    'sheet': sheet_name,
                    'error_type': 'Long Sheet Name',
                    'description': f'Sheet name exceeds 31 characters: "{sheet_name}" ({len(sheet_name)} chars)',
                    'length': len(sheet_name),
                    'severity': 'medium'
                })
        
        # Check for empty sheets
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            has_data = False
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        has_data = True
                        break
                if has_data:
                    break
            
            if not has_data:
                structural_issues.append({
                    'sheet': sheet_name,
                    'error_type': 'Empty Sheet',
                    'description': f'Sheet "{sheet_name}" contains no data',
                    'severity': 'low'
                })
        
        return structural_issues
    
    def _detect_compatibility_warnings(self) -> List[Dict[str, Any]]:
        """
        Detect compatibility warnings.
        
        Returns:
            List of compatibility warnings found
        """
        compatibility_warnings = []
        
        # Check file format
        if self.file_path.suffix.lower() == '.xls':
            compatibility_warnings.append({
                'type': 'File Format',
                'error_type': 'Legacy Format',
                'description': 'File is in legacy .xls format. Consider converting to .xlsx for better compatibility.',
                'severity': 'medium'
            })
        
        # Check for VBA code (may cause issues in some environments)
        if hasattr(self.workbook, 'vba_archive') and self.workbook.vba_archive:
            compatibility_warnings.append({
                'type': 'VBA',
                'error_type': 'VBA Code Present',
                'description': 'Workbook contains VBA code which may not work in all environments.',
                'severity': 'low'
            })
        
        # Check for very large file size
        file_size_mb = self.file_path.stat().st_size / (1024 * 1024)
        if file_size_mb > 50:  # 50MB threshold
            compatibility_warnings.append({
                'type': 'File Size',
                'error_type': 'Large File',
                'description': f'File size is {file_size_mb:.1f}MB which may cause performance issues.',
                'size_mb': file_size_mb,
                'severity': 'medium'
            })
        
        return compatibility_warnings
    
    def sniff_errors(self) -> Dict[str, Any]:
        """
        Perform comprehensive error detection on the Excel file.
        
        Returns:
            Dictionary containing all detected errors and issues
        """
        logger.info(f"Starting error detection for: {self.file_path}")
        
        try:
            self._load_workbook()
            
            # Run all error detection methods
            self.errors['formula_errors'] = self._detect_formula_errors()
            self.errors['circular_references'] = self._detect_circular_references()
            self.errors['broken_links'] = self._detect_broken_links()
            self.errors['data_validation_issues'] = self._detect_data_validation_issues()
            self.errors['performance_issues'] = self._detect_performance_issues()
            self.errors['structural_issues'] = self._detect_structural_issues()
            self.errors['compatibility_warnings'] = self._detect_compatibility_warnings()
            
            # Generate summary
            self.errors['summary'] = self._generate_summary()
            
            logger.info(f"Error detection completed. Found {self.errors['summary']['total_issues']} issues.")
            
        except Exception as e:
            logger.error(f"Error during error detection: {e}")
            raise
        finally:
            if self.workbook:
                self.workbook.close()
        
        return self.errors
    
    def _generate_summary(self) -> Dict[str, Any]:
        """
        Generate a summary of all detected issues.
        
        Returns:
            Summary dictionary
        """
        total_issues = (
            len(self.errors['formula_errors']) +
            len(self.errors['circular_references']) +
            len(self.errors['broken_links']) +
            len(self.errors['data_validation_issues']) +
            len(self.errors['performance_issues']) +
            len(self.errors['structural_issues']) +
            len(self.errors['compatibility_warnings'])
        )
        
        # Count by severity
        severity_counts = {'high': 0, 'medium': 0, 'low': 0}
        
        for error_type in self.errors.values():
            if isinstance(error_type, list):
                for error in error_type:
                    if isinstance(error, dict) and 'severity' in error:
                        severity_counts[error['severity']] += 1
        
        return {
            'total_issues': total_issues,
            'severity_breakdown': severity_counts,
            'error_types': {
                'formula_errors': len(self.errors['formula_errors']),
                'circular_references': len(self.errors['circular_references']),
                'broken_links': len(self.errors['broken_links']),
                'data_validation_issues': len(self.errors['data_validation_issues']),
                'performance_issues': len(self.errors['performance_issues']),
                'structural_issues': len(self.errors['structural_issues']),
                'compatibility_warnings': len(self.errors['compatibility_warnings'])
            },
            'timestamp': datetime.now().isoformat(),
            'file_path': str(self.file_path),
            'file_size_mb': round(self.file_path.stat().st_size / (1024 * 1024), 2)
        }
    
    def to_markdown(self) -> str:
        """
        Convert error analysis to markdown format.
        
        Returns:
            Markdown formatted error report
        """
        if not self.errors['summary']:
            return "# Excel Error Analysis: No Data Available\n\n*No error analysis has been performed yet.*"
        
        md_lines = []
        
        # Header
        md_lines.append(f"# Excel Error Analysis: {self.file_path.name}")
        md_lines.append("")
        md_lines.append(f"**Analysis Date:** {self.errors['summary']['timestamp']}")
        md_lines.append(f"**File Size:** {self.errors['summary']['file_size_mb']} MB")
        md_lines.append("")
        
        # Summary
        summary = self.errors['summary']
        md_lines.append("## Summary")
        md_lines.append("")
        md_lines.append(f"- **Total Issues:** {summary['total_issues']}")
        md_lines.append(f"- **High Severity:** {summary['severity_breakdown']['high']}")
        md_lines.append(f"- **Medium Severity:** {summary['severity_breakdown']['medium']}")
        md_lines.append(f"- **Low Severity:** {summary['severity_breakdown']['low']}")
        md_lines.append("")
        
        # Error type breakdown
        md_lines.append("### Issue Breakdown")
        md_lines.append("")
        for error_type, count in summary['error_types'].items():
            if count > 0:
                md_lines.append(f"- **{error_type.replace('_', ' ').title()}:** {count}")
        md_lines.append("")
        
        # Detailed sections
        sections = [
            ('Formula Errors', 'formula_errors', 'high'),
            ('Circular References', 'circular_references', 'high'),
            ('Broken Links', 'broken_links', 'high'),
            ('Data Validation Issues', 'data_validation_issues', 'medium'),
            ('Performance Issues', 'performance_issues', 'medium'),
            ('Structural Issues', 'structural_issues', 'low'),
            ('Compatibility Warnings', 'compatibility_warnings', 'low')
        ]
        
        for title, key, default_severity in sections:
            errors = self.errors[key]
            if errors:
                md_lines.append(f"## {title}")
                md_lines.append("")
                
                for error in errors:
                    severity = error.get('severity', default_severity)
                    severity_emoji = {'high': '🔴', 'medium': '🟡', 'low': '🟢'}[severity]
                    
                    md_lines.append(f"### {severity_emoji} {error.get('error_type', 'Issue')}")
                    
                    if 'sheet' in error:
                        md_lines.append(f"**Sheet:** {error['sheet']}")
                    if 'cell' in error:
                        md_lines.append(f"**Cell:** {error['cell']}")
                    
                    md_lines.append(f"**Description:** {error['description']}")
                    
                    if 'formula' in error and error['formula']:
                        md_lines.append(f"**Formula:** `{error['formula']}`")
                    
                    if 'value' in error and 'threshold' in error:
                        md_lines.append(f"**Value:** {error['value']:,} (Threshold: {error['threshold']:,})")
                    
                    md_lines.append("")
        
        return "\n".join(md_lines)
    
    def save_markdown(self, output_path: Path) -> None:
        """
        Save error analysis as markdown file.
        
        Args:
            output_path: Path to save the markdown file
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        markdown_content = self.to_markdown()
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(markdown_content)
        
        logger.info(f"Error analysis saved to: {output_path}")
    
    def save_json(self, output_path: Path) -> None:
        """
        Save error analysis as JSON file.
        
        Args:
            output_path: Path to save the JSON file
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.errors, f, indent=2, default=str)
        
        logger.info(f"Error analysis saved to: {output_path}")


def sniff_excel_errors(file_path: Path, output_dir: Optional[Path] = None) -> Dict[str, Any]:
    """
    Convenience function to sniff Excel errors and optionally save reports.
    
    Args:
        file_path: Path to the Excel file
        output_dir: Optional directory to save reports
        
    Returns:
        Dictionary containing error analysis results
    """
    sniffer = ExcelErrorSniffer(file_path)
    errors = sniffer.sniff_errors()
    
    if output_dir:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Save reports
        markdown_path = output_dir / f"{file_path.stem}_error_analysis.md"
        json_path = output_dir / f"{file_path.stem}_error_analysis.json"
        
        sniffer.save_markdown(markdown_path)
        sniffer.save_json(json_path)
    
    return errors


if __name__ == '__main__':
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python excel_error_sniffer.py <excel_file> [output_dir]")
        sys.exit(1)
    
    file_path = Path(sys.argv[1])
    output_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    
    try:
        errors = sniff_excel_errors(file_path, output_dir)
        print(f"Error analysis completed. Found {errors['summary']['total_issues']} issues.")
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1) 