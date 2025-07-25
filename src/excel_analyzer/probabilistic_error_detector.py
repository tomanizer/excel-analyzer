#!/usr/bin/env python3
"""
Probabilistic Excel Error Detector - Advanced error detection using pattern analysis and probability models.

This module provides a framework for detecting complex Excel errors using probabilistic models
and pattern analysis. Each error type has its own detection algorithm that returns a probability
score, allowing users to set thresholds to control false positives.
"""

import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple, Callable
import warnings
from datetime import datetime
import json
import re
from dataclasses import dataclass
from enum import Enum
from collections import Counter

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.external_reference import ExternalReference

logger = logging.getLogger(__name__)


class ErrorSeverity(Enum):
    """Error severity levels."""
    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"


@dataclass
class ErrorDetectionResult:
    """Result of an error detection algorithm."""
    error_type: str
    description: str
    probability: float  # 0.0 to 1.0
    severity: ErrorSeverity
    location: Optional[str] = None  # e.g., "Sheet1!A1", "NamedRange:Revenue"
    details: Optional[Dict[str, Any]] = None
    suggested_fix: Optional[str] = None


class ErrorDetector:
    """Base class for error detection algorithms."""
    
    def __init__(self, name: str, description: str, severity: ErrorSeverity):
        self.name = name
        self.description = description
        self.severity = severity
    
    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        """
        Detect errors of this type in the workbook.
        
        Args:
            workbook: The Excel workbook to analyze
            **kwargs: Additional parameters for detection
            
        Returns:
            List of detected errors with probability scores
        """
        raise NotImplementedError("Subclasses must implement detect()")


class ProbabilisticErrorSniffer:
    """
    Advanced Excel error detection using probabilistic models and pattern analysis.
    
    This framework allows for:
    - Probabilistic error detection with confidence scores
    - Configurable thresholds to control false positives
    - Extensible architecture for adding new error types
    - Pattern-based analysis for complex error detection
    """
    
    def __init__(self, file_path: Path, error_threshold: float = 0.7):
        """
        Initialize the Probabilistic Error Sniffer.
        
        Args:
            file_path: Path to the Excel file to analyze
            error_threshold: Minimum probability threshold for reporting errors (0.0 to 1.0)
        """
        self.file_path = Path(file_path)
        self.workbook = None
        self.error_threshold = error_threshold
        self.detectors: List[ErrorDetector] = []
        self.detection_results: Dict[str, List[ErrorDetectionResult]] = {}
        
        # Register built-in error detectors
        self._register_builtin_detectors()
    
    def _load_workbook(self) -> None:
        """Load the Excel workbook safely."""
        try:
            self.workbook = openpyxl.load_workbook(
                self.file_path, 
                data_only=False,  # Keep formulas for error detection
                keep_vba=True
            )
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            raise
    
    def register_detector(self, detector: ErrorDetector) -> None:
        """Register a new error detector."""
        self.detectors.append(detector)
        logger.info(f"Registered error detector: {detector.name}")
    
    def _register_builtin_detectors(self) -> None:
        """Register all built-in error detectors."""
        # We'll add these one by one as we implement them
        pass
    
    def detect_all_errors(self) -> Dict[str, Any]:
        """
        Run all registered error detectors and return results above threshold.
        
        Returns:
            Dictionary containing all detected errors organized by type
        """
        logger.info(f"Starting probabilistic error detection for: {self.file_path}")
        
        try:
            self._load_workbook()
            
            # Run all detectors
            for detector in self.detectors:
                try:
                    results = detector.detect(self.workbook)
                    # Filter results by threshold
                    filtered_results = [
                        result for result in results 
                        if result.probability >= self.error_threshold
                    ]
                    self.detection_results[detector.name] = filtered_results
                    
                    logger.info(f"Detector '{detector.name}' found {len(filtered_results)} errors above threshold")
                    
                except Exception as e:
                    logger.error(f"Error in detector '{detector.name}': {e}")
                    self.detection_results[detector.name] = []
            
            # Generate summary
            summary = self._generate_summary()
            self.detection_results['summary'] = summary
            
            logger.info(f"Error detection completed. Found {summary['total_errors']} errors above threshold.")
            
        except Exception as e:
            logger.error(f"Error during detection: {e}")
            raise
        finally:
            if self.workbook:
                self.workbook.close()
        
        return self.detection_results
    
    def _generate_summary(self) -> Dict[str, Any]:
        """Generate summary of detection results."""
        total_errors = sum(len(results) for key, results in self.detection_results.items() if key != 'summary')
        
        severity_counts = {severity.value: 0 for severity in ErrorSeverity}
        for results in self.detection_results.values():
            if isinstance(results, list):
                for result in results:
                    severity_counts[result.severity.value] += 1
        
        return {
            'total_errors': total_errors,
            'severity_breakdown': severity_counts,
            'error_types': {
                detector.name: len(self.detection_results.get(detector.name, []))
                for detector in self.detectors
            },
            'threshold_used': self.error_threshold,
            'timestamp': datetime.now().isoformat(),
            'file_path': str(self.file_path),
            'file_size_mb': round(self.file_path.stat().st_size / (1024 * 1024), 2)
        }


# ============================================================================
# ERROR DETECTION ALGORITHMS
# ============================================================================

class HiddenDataInRangesDetector(ErrorDetector):
    """
    Detector for hidden rows/columns in data ranges.
    
    Algorithm:
    1. Identify data ranges (continuous cells with data)
    2. Check if ranges include hidden rows/columns
    3. Analyze data consistency within ranges
    4. Calculate probability based on hidden data patterns
    """
    
    def __init__(self):
        super().__init__(
            name="hidden_data_in_ranges",
            description="Hidden rows/columns in data ranges that may contain incorrect data",
            severity=ErrorSeverity.HIGH
        )
    
    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        """Detect hidden data in ranges."""
        results = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Find data ranges
            data_ranges = self._find_data_ranges(sheet)
            
            for range_info in data_ranges:
                # Check for hidden rows/columns in this range
                hidden_analysis = self._analyze_hidden_data(sheet, range_info)
                
                if hidden_analysis['has_hidden_data']:
                    probability = self._calculate_hidden_data_probability(hidden_analysis)
                    
                    if probability > 0:
                        results.append(ErrorDetectionResult(
                            error_type=self.name,
                            description=f"Data range {range_info['range']} contains hidden {hidden_analysis['hidden_type']} with potentially inconsistent data",
                            probability=probability,
                            severity=self.severity,
                            location=f"{sheet_name}!{range_info['range']}",
                            details=hidden_analysis,
                            suggested_fix="Review hidden data in range and ensure consistency with visible data"
                        ))
        
        return results
    
    def _find_data_ranges(self, sheet) -> List[Dict[str, Any]]:
        """Find continuous data ranges in the sheet."""
        ranges = []
        
        # Get the used range
        min_row = sheet.min_row
        max_row = sheet.max_row
        min_col = sheet.min_column
        max_col = sheet.max_column
        
        if min_row is None or max_row is None:
            return ranges
        
        # Find continuous data blocks
        current_range_start = None
        current_range_end = None
        
        for row in range(min_row, max_row + 1):
            row_has_data = any(
                sheet.cell(row=row, column=col).value is not None
                for col in range(min_col, max_col + 1)
            )
            
            if row_has_data and current_range_start is None:
                current_range_start = row
            elif not row_has_data and current_range_start is not None:
                current_range_end = row - 1
                ranges.append({
                    'start_row': current_range_start,
                    'end_row': current_range_end,
                    'range': f"{get_column_letter(min_col)}{current_range_start}:{get_column_letter(max_col)}{current_range_end}"
                })
                current_range_start = None
        
        # Handle range that extends to the end
        if current_range_start is not None:
            ranges.append({
                'start_row': current_range_start,
                'end_row': max_row,
                'range': f"{get_column_letter(min_col)}{current_range_start}:{get_column_letter(max_col)}{max_row}"
            })
        
        return ranges
    
    def _analyze_hidden_data(self, sheet, range_info: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze hidden data within a range."""
        hidden_rows = []
        hidden_cols = []
        
        # Check for hidden rows
        for row in range(range_info['start_row'], range_info['end_row'] + 1):
            if sheet.row_dimensions[row].hidden:
                hidden_rows.append(row)
        
        # Check for hidden columns
        for col in range(sheet.min_column, sheet.max_column + 1):
            if sheet.column_dimensions[get_column_letter(col)].hidden:
                hidden_cols.append(col)
        
        # Analyze data consistency
        visible_data = []
        hidden_data = []
        
        for row in range(range_info['start_row'], range_info['end_row'] + 1):
            for col in range(sheet.min_column, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None:
                    if row in hidden_rows or col in hidden_cols:
                        hidden_data.append(cell.value)
                    else:
                        visible_data.append(cell.value)
        
        return {
            'has_hidden_data': len(hidden_rows) > 0 or len(hidden_cols) > 0,
            'hidden_rows': hidden_rows,
            'hidden_cols': hidden_cols,
            'hidden_type': 'rows' if len(hidden_rows) > 0 else 'columns' if len(hidden_cols) > 0 else 'none',
            'visible_data_count': len(visible_data),
            'hidden_data_count': len(hidden_data),
            'data_consistency_score': self._calculate_data_consistency(visible_data, hidden_data)
        }
    
    def _calculate_data_consistency(self, visible_data: List, hidden_data: List) -> float:
        """Calculate consistency score between visible and hidden data."""
        if not visible_data or not hidden_data:
            return 1.0  # No inconsistency if one set is empty
        
        # Simple consistency check - can be enhanced with more sophisticated analysis
        visible_types = set(type(val) for val in visible_data)
        hidden_types = set(type(val) for val in hidden_data)
        
        type_overlap = len(visible_types.intersection(hidden_types))
        total_types = len(visible_types.union(hidden_types))
        
        return type_overlap / total_types if total_types > 0 else 1.0
    
    def _calculate_hidden_data_probability(self, analysis: Dict[str, Any]) -> float:
        """Calculate probability of hidden data causing issues."""
        if not analysis['has_hidden_data']:
            return 0.0
        
        # Base probability from having hidden data
        base_prob = 0.3
        
        # Adjust based on data consistency
        consistency_factor = 1.0 - analysis['data_consistency_score']
        base_prob += consistency_factor * 0.4
        
        # Adjust based on amount of hidden data
        total_data = analysis['visible_data_count'] + analysis['hidden_data_count']
        if total_data > 0:
            hidden_ratio = analysis['hidden_data_count'] / total_data
            base_prob += hidden_ratio * 0.3
        
        return min(base_prob, 1.0)


class CircularNamedRangesDetector(ErrorDetector):
    """
    Detector for circular references in named ranges.
    
    Algorithm:
    1. Extract all named ranges from the workbook
    2. Parse each named range's formula to identify dependencies
    3. Build a dependency graph
    4. Detect cycles using graph algorithms
    5. Calculate probability based on cycle characteristics
    """
    
    def __init__(self):
        super().__init__(
            name="circular_named_ranges",
            description="Circular references in named ranges that can cause infinite calculation loops",
            severity=ErrorSeverity.HIGH
        )
    
    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        """Detect circular references in named ranges."""
        results = []
        
        # Extract all named ranges
        named_ranges = self._extract_named_ranges(workbook)
        
        if not named_ranges:
            return results
        
        # Build dependency graph
        dependency_graph = self._build_dependency_graph(named_ranges)
        
        # Detect cycles
        cycles = self._detect_cycles(dependency_graph)
        
        # Generate results for each cycle
        for cycle in cycles:
            probability = self._calculate_circular_probability(cycle, named_ranges, dependency_graph)
            
            if probability > 0:
                # Get formulas for the cycle
                cycle_formulas = {name: named_ranges[name]['formula'] for name in cycle}
                
                results.append(ErrorDetectionResult(
                    error_type=self.name,
                    description=f"Circular reference detected: {' → '.join(cycle)} → {cycle[0]}",
                    probability=probability,
                    severity=self.severity,
                    location=f"NamedRanges: {', '.join(cycle[:-1])}",  # Exclude the duplicate at the end
                    details={
                        'cycle': cycle,
                        'cycle_length': len(cycle),
                        'formulas': cycle_formulas,
                        'dependency_graph': dependency_graph
                    },
                    suggested_fix="Break the circular dependency by introducing intermediate calculations or using iterative calculation settings"
                ))
        
        return results
    
    def _extract_named_ranges(self, workbook: openpyxl.Workbook) -> Dict[str, Dict[str, Any]]:
        """Extract all named ranges and their formulas."""
        named_ranges = {}
        
        for name in workbook.defined_names:
            try:
                # Get the DefinedName object
                defined_name = workbook.defined_names[name]
                
                # Get the formula/reference
                formula = defined_name.attr_text if hasattr(defined_name, 'attr_text') else str(defined_name)
                
                named_ranges[name] = {
                    'formula': formula,
                    'scope': defined_name.localSheetId if hasattr(defined_name, 'localSheetId') else None,
                    'comment': defined_name.comment if hasattr(defined_name, 'comment') else None
                }
            except Exception as e:
                logger.warning(f"Could not extract named range {name}: {e}")
                continue
        
        return named_ranges
    
    def _parse_named_range_formula(self, formula: str) -> List[str]:
        """Parse a named range formula to extract dependencies."""
        dependencies = []
        
        if not formula or not isinstance(formula, str):
            return dependencies
        
        # Remove leading '=' if present
        if formula.startswith('='):
            formula = formula[1:]
        
        # Extract named range references using regex
        import re
        
        # Pattern for named range references
        # Matches: standalone names, names in functions, names in operations
        patterns = [
            r'\b([A-Za-z_][A-Za-z0-9_]*)\b',  # Basic named range pattern
            r'([A-Za-z_][A-Za-z0-9_]*)',      # More permissive pattern
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, formula)
            for match in matches:
                # Filter out common Excel functions and keywords (case-sensitive)
                excel_functions = {
                    'SUM', 'AVERAGE', 'COUNT', 'MAX', 'MIN', 'IF', 'AND', 'OR',
                    'TRUE', 'FALSE', 'PI', 'TODAY', 'NOW', 'ROW', 'COLUMN',
                    'ABS', 'ROUND', 'INT', 'MOD', 'POWER', 'SQRT', 'LOG', 'LN',
                    'SIN', 'COS', 'TAN', 'ASIN', 'ACOS', 'ATAN', 'RAND', 'RANDBETWEEN'
                }
                
                # Filter out cell references (like A1, B2, etc.)
                is_cell_reference = (
                    len(match) >= 2 and 
                    match[0].isalpha() and 
                    match[1:].isdigit()
                )
                
                if (match not in excel_functions and 
                    len(match) >= 1 and 
                    not is_cell_reference):
                    dependencies.append(match)
        
        return list(set(dependencies))  # Remove duplicates
    
    def _build_dependency_graph(self, named_ranges: Dict[str, Dict[str, Any]]) -> Dict[str, List[str]]:
        """Build a dependency graph from named ranges."""
        graph = {}
        
        for name, info in named_ranges.items():
            dependencies = self._parse_named_range_formula(info['formula'])
            # Only include dependencies that are actual named ranges
            valid_dependencies = [dep for dep in dependencies if dep in named_ranges]
            graph[name] = valid_dependencies
        
        return graph
    
    def _detect_cycles(self, graph: Dict[str, List[str]]) -> List[List[str]]:
        """Detect cycles in the dependency graph using DFS."""
        cycles = []
        
        def dfs(node: str, path: List[str], visited: set, rec_stack: set):
            """Depth-first search to detect cycles."""
            if node in rec_stack:
                # Found a cycle
                cycle_start = path.index(node)
                cycle = path[cycle_start:] + [node]
                # Add cycles with at least 2 nodes (including self-references)
                if len(cycle) >= 2:  # Allow self-references and multi-node cycles
                    cycles.append(cycle)
                return
            
            if node in visited:
                return
            
            visited.add(node)
            rec_stack.add(node)
            path.append(node)
            
            for neighbor in graph.get(node, []):
                dfs(neighbor, path.copy(), visited.copy(), rec_stack.copy())
            
            rec_stack.remove(node)
        
        # Run DFS from each node
        for node in graph:
            visited = set()
            rec_stack = set()
            dfs(node, [], visited, rec_stack)
        
        # Remove duplicate cycles (same cycle starting from different nodes)
        unique_cycles = []
        for cycle in cycles:
            # Normalize cycle by starting from the lexicographically smallest node
            cycle_nodes = cycle[:-1]  # Exclude the last node (duplicate of first)
            min_node = min(cycle_nodes)
            start_idx = cycle_nodes.index(min_node)
            normalized_cycle = cycle_nodes[start_idx:] + cycle_nodes[:start_idx] + [min_node]
            
            # Check if this normalized cycle is already in unique_cycles
            is_duplicate = False
            for existing_cycle in unique_cycles:
                existing_nodes = existing_cycle[:-1]
                if set(existing_nodes) == set(cycle_nodes):
                    is_duplicate = True
                    break
            
            if not is_duplicate:
                unique_cycles.append(normalized_cycle)
        
        return unique_cycles
    
    def _calculate_circular_probability(self, cycle: List[str], named_ranges: Dict[str, Dict[str, Any]], graph: Dict[str, List[str]]) -> float:
        """Calculate probability of circular reference being problematic."""
        if not cycle:
            return 0.0
        
        # Base probability based on cycle length
        cycle_length = len(cycle)
        if cycle_length == 2:
            base_prob = 0.9  # Very high probability for 2-range cycles
        elif cycle_length == 3:
            base_prob = 0.8  # High probability for 3-range cycles
        elif cycle_length == 4:
            base_prob = 0.7  # Medium-high probability for 4-range cycles
        else:
            base_prob = 0.6  # Medium probability for longer cycles
        
        # Adjust based on formula complexity
        complexity_factor = 0.0
        for name in cycle:
            formula = named_ranges[name]['formula']
            # Count functions, operators, and references
            function_count = formula.count('(') + formula.count(')')
            operator_count = formula.count('+') + formula.count('-') + formula.count('*') + formula.count('/')
            reference_count = len(self._parse_named_range_formula(formula))
            
            complexity = (function_count + operator_count + reference_count) / 10.0
            complexity_factor = max(complexity_factor, complexity)
        
        base_prob += complexity_factor * 0.2
        
        # Adjust based on usage frequency (how many other named ranges depend on these)
        usage_factor = 0.0
        for name in cycle:
            # Count how many other named ranges depend on this one
            dependents = sum(1 for deps in graph.values() if name in deps)
            usage_factor = max(usage_factor, dependents / 10.0)  # Normalize
        
        base_prob += usage_factor * 0.1
        
        # Check for aggregation functions in cycle (more dangerous)
        aggregation_functions = ['SUM', 'AVERAGE', 'COUNT', 'MAX', 'MIN', 'SUMPRODUCT']
        for name in cycle:
            formula = named_ranges[name]['formula'].upper()
            if any(func in formula for func in aggregation_functions):
                base_prob += 0.1
                break
        
        return min(base_prob, 1.0)


class InconsistentDateFormatsDetector(ErrorDetector):
    """
    Detector for inconsistent date formats in date calculations.
    
    Algorithm:
    1. Scan for formulas that perform date arithmetic or use date functions
    2. Identify ranges/columns involved in date calculations
    3. Analyze data types in those ranges (Excel date, text that looks like date, other)
    4. Flag if a range contains a mix of true dates and text-formatted dates
    5. Calculate probability based on proportion of inconsistencies
    """
    def __init__(self):
        super().__init__(
            name="inconsistent_date_formats",
            description="Mixed date formats (text vs. actual dates) in date-based calculations",
            severity=ErrorSeverity.HIGH
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        results = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            checked_ranges = set()
            # 1. Scan all columns for mixed date types or all text dates
            for col in range(1, sheet.max_column + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                if col_letter in checked_ranges:
                    continue
                checked_ranges.add(col_letter)
                analysis = self._analyze_range(sheet, col_letter)
                # Flag if mixed types OR all text dates (and at least 1 text date)
                is_all_text_dates = analysis['text_date_count'] > 0 and analysis['date_count'] == 0
                if analysis['mixed_types'] or is_all_text_dates:
                    probability = self._calculate_probability(analysis, is_all_text_dates=is_all_text_dates)
                    if probability > 0:
                        results.append(ErrorDetectionResult(
                            error_type=self.name,
                            description=f"{'All text dates' if is_all_text_dates else 'Mixed date formats'} detected in column {col_letter} on sheet {sheet_name}",
                            probability=probability,
                            severity=self.severity,
                            location=f"{sheet_name}!{col_letter}",
                            details=analysis,
                            suggested_fix="Convert all dates in the column to Excel date format for consistency"
                        ))
            # 2. Also scan columns referenced by date formulas (legacy logic)
            date_formula_cells = self._find_date_formula_cells(sheet)
            for cell in date_formula_cells:
                referenced_ranges = self._extract_referenced_ranges(cell)
                for rng in referenced_ranges:
                    if rng in checked_ranges:
                        continue
                    checked_ranges.add(rng)
                    analysis = self._analyze_range(sheet, rng)
                    is_all_text_dates = analysis['text_date_count'] > 0 and analysis['date_count'] == 0
                    if analysis['mixed_types'] or is_all_text_dates:
                        probability = self._calculate_probability(analysis, is_all_text_dates=is_all_text_dates)
                        if probability > 0:
                            results.append(ErrorDetectionResult(
                                error_type=self.name,
                                description=f"{'All text dates' if is_all_text_dates else 'Mixed date formats'} detected in range {rng} on sheet {sheet_name}",
                                probability=probability,
                                severity=self.severity,
                                location=f"{sheet_name}!{rng}",
                                details=analysis,
                                suggested_fix="Convert all dates in the range to Excel date format for consistency"
                            ))
        return results

    def _find_date_formula_cells(self, sheet) -> List[openpyxl.cell.cell.Cell]:
        """Find cells with formulas that perform date arithmetic or use date functions."""
        date_functions = {'DATEDIF', 'DATE', 'YEAR', 'MONTH', 'DAY', 'EDATE', 'EOMONTH', 'TODAY', 'NOW', 'NETWORKDAYS', 'WORKDAY'}
        date_formula_cells = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:
                    formula = str(cell.value).upper()
                    # Check for date functions or date arithmetic (e.g., +, -, with cell refs)
                    if any(func in formula for func in date_functions) or self._is_date_arithmetic(formula):
                        date_formula_cells.append(cell)
        return date_formula_cells

    def _is_date_arithmetic(self, formula: str) -> bool:
        # Simple heuristic: look for + or - between cell references
        import re
        # e.g., =A1-B1 or =B2+30
        return bool(re.search(r"[A-Z]+[0-9]+\s*[-+]\s*[A-Z0-9]+", formula))

    def _extract_referenced_ranges(self, cell) -> List[str]:
        # For simplicity, just extract all cell references in the formula
        import re
        formula = str(cell.value)
        # Matches A1, B2, C10, etc.
        refs = re.findall(r"[A-Z]+[0-9]+", formula)
        # Group by column (e.g., all A1, A2, A3 -> A)
        columns = set(ref[0] for ref in refs if len(ref) >= 2)
        # For now, treat each column as a range (e.g., A)
        return [f"{col}" for col in columns]

    def _analyze_range(self, sheet, col: str) -> dict:
        # Analyze all cells in the given column
        from openpyxl.utils import column_index_from_string
        col_idx = column_index_from_string(col)
        date_count = 0
        text_date_count = 0
        other_count = 0
        total = 0
        text_date_examples = []
        for row in sheet.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is None:
                    continue
                total += 1
                if self._is_excel_date(cell):
                    date_count += 1
                elif self._looks_like_date(cell.value):
                    text_date_count += 1
                    if len(text_date_examples) < 3:
                        text_date_examples.append(cell.value)
                else:
                    other_count += 1
        mixed_types = date_count > 0 and text_date_count > 0
        return {
            'date_count': date_count,
            'text_date_count': text_date_count,
            'other_count': other_count,
            'total': total,
            'mixed_types': mixed_types,
            'text_date_examples': text_date_examples
        }

    def _is_excel_date(self, cell) -> bool:
        # openpyxl stores Excel dates as numbers with a date format
        if cell.is_date:
            return True
        # Sometimes dates are stored as numbers with a date format
        if cell.data_type == 'n' and cell.number_format and 'yy' in cell.number_format.lower():
            return True
        return False

    def _looks_like_date(self, value) -> bool:
        import re
        # Match common date patterns: YYYY-MM-DD, DD/MM/YYYY, MM/DD/YYYY, etc.
        if not isinstance(value, str):
            return False
        patterns = [
            r'\b\d{4}-\d{2}-\d{2}\b',  # 2023-01-01
            r'\b\d{2}/\d{2}/\d{4}\b',  # 01/01/2023
            r'\b\d{1,2} [A-Za-z]{3,9} \d{4}\b',  # 1 Jan 2023
            r'\b\d{2}\.\d{2}\.\d{4}\b',  # 01.01.2023
        ]
        return any(re.search(pat, value) for pat in patterns)

    def _calculate_probability(self, analysis: dict, is_all_text_dates: bool = False) -> float:
        if analysis['total'] == 0:
            return 0.0
        if is_all_text_dates:
            # All text dates, no Excel dates
            ratio = analysis['text_date_count'] / analysis['total']
            if ratio > 0.9:
                return 0.5  # Lower probability, but still a warning
            elif ratio > 0.5:
                return 0.2
            else:
                return 0.1
        ratio = analysis['text_date_count'] / analysis['total']
        if ratio > 0.1:
            return 0.9
        elif ratio > 0.01:
            return 0.5
        elif ratio > 0:
            return 0.2
        return 0.0


class ArrayFormulaSpillErrorsDetector(ErrorDetector):
    """
    Detector for array formula spill errors.
    
    Algorithm:
    1. Scan for array formulas (curly braces or dynamic array functions)
    2. Look for #SPILL! errors in cells
    3. For each array formula, check the intended spill range for conflicts
    4. Calculate probability based on evidence of spill errors or conflicts
    """
    def __init__(self):
        super().__init__(
            name="array_formula_spill_errors",
            description="Array formulas that can't spill properly due to insufficient space or conflicts",
            severity=ErrorSeverity.HIGH
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        results = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    # 1. Check for #SPILL! error
                    if isinstance(cell.value, str) and cell.value.strip().upper() == '#SPILL!':
                        results.append(ErrorDetectionResult(
                            error_type=self.name,
                            description=f"#SPILL! error detected in cell {cell.coordinate} on sheet {sheet_name}",
                            probability=0.95,
                            severity=self.severity,
                            location=f"{sheet_name}!{cell.coordinate}",
                            details={'cell': cell.coordinate, 'error': '#SPILL!'},
                            suggested_fix="Check for blocked cells, merged cells, or other conflicts in the intended spill range."
                        ))
                    # 2. Check for array formulas (legacy or dynamic)
                    if self._is_array_formula(cell):
                        spill_range, conflict_cells = self._analyze_spill_range(sheet, cell)
                        if conflict_cells:
                            # High probability if most/all spill cells are blocked
                            spill_ratio = len(conflict_cells) / (len(spill_range) - 1)  # -1 to exclude original cell
                            probability = 0.9 if spill_ratio >= 0.8 else 0.7
                            results.append(ErrorDetectionResult(
                                error_type=self.name,
                                description=f"Array formula in {cell.coordinate} cannot spill due to conflicts in range {spill_range[0]}:{spill_range[-1]}",
                                probability=probability,
                                severity=self.severity,
                                location=f"{sheet_name}!{cell.coordinate}",
                                details={
                                    'cell': cell.coordinate,
                                    'spill_range': [c.coordinate for c in spill_range],
                                    'conflict_cells': [c.coordinate for c in conflict_cells],
                                },
                                suggested_fix="Clear or move conflicting cells in the spill range."
                            ))
        return results

    def _is_array_formula(self, cell) -> bool:
        # Legacy array formulas: openpyxl marks with cell.data_type == 'f' and formula starts with {= or ends with }
        if cell.data_type == 'f' and cell.value:
            formula = str(cell.value)
            if formula.startswith('{=') or formula.endswith('}'):  # legacy CSE
                return True
            # Dynamic array functions
            dynamic_funcs = ['SEQUENCE', 'UNIQUE', 'SORT', 'FILTER', 'RANDARRAY', 'TRANSPOSE', 'XMATCH', 'XLOOKUP']
            if any(func in formula.upper() for func in dynamic_funcs):
                return True
        return False

    def _analyze_spill_range(self, sheet, cell):
        # Heuristic: For legacy arrays, assume 1xN or Nx1 block; for dynamic, try to estimate
        # For this implementation, just check the right and down cells for conflicts (up to 10x10 block)
        spill_range = [cell]
        conflict_cells = []
        max_rows, max_cols = 10, 10
        # Check rightwards
        for dc in range(1, max_cols):
            col = cell.column + dc
            if col > sheet.max_column:
                break
            c = sheet.cell(row=cell.row, column=col)
            if c.value is not None:
                conflict_cells.append(c)
            spill_range.append(c)
        # Check downwards
        for dr in range(1, max_rows):
            row = cell.row + dr
            if row > sheet.max_row:
                break
            c = sheet.cell(row=row, column=cell.column)
            if c.value is not None:
                conflict_cells.append(c)
            spill_range.append(c)
        return spill_range, conflict_cells


class VolatileFunctionsDetector(ErrorDetector):
    """
    Detector for volatile functions that can cause performance issues in large models.
    
    Algorithm:
    1. Scan for volatile functions (NOW, TODAY, RAND, OFFSET, INDIRECT, etc.)
    2. Analyze usage patterns and dependency impact
    3. Calculate performance impact based on frequency and context
    4. Flag if volatile functions are used inappropriately or excessively
    """
    def __init__(self):
        super().__init__(
            name="volatile_functions",
            description="Volatile functions causing excessive recalculations and performance issues",
            severity=ErrorSeverity.MEDIUM
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        results = []
        
        # Collect all volatile functions across the workbook
        volatile_cells = []
        total_formulas = 0
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:
                        total_formulas += 1
                        volatile_funcs = self._find_volatile_functions(str(cell.value))
                        if volatile_funcs:
                            volatile_cells.append({
                                'cell': cell,
                                'sheet': sheet_name,
                                'functions': volatile_funcs,
                                'dependencies': self._count_dependencies(sheet, cell)
                            })
        
        if not volatile_cells:
            return results
        
        # Analyze overall impact
        total_volatile_funcs = sum(len(cell['functions']) for cell in volatile_cells)
        model_size_factor = min(total_formulas / 100, 1.0)  # Normalize by model size
        
        # Calculate probability based on usage patterns
        probability = self._calculate_volatile_probability(
            total_volatile_funcs, 
            len(volatile_cells), 
            total_formulas,
            volatile_cells
        )
        
        if probability > 0:
            # Group by severity level
            high_impact_cells = [c for c in volatile_cells if c['dependencies'] > 5]
            medium_impact_cells = [c for c in volatile_cells if 2 <= c['dependencies'] <= 5]
            
            results.append(ErrorDetectionResult(
                error_type=self.name,
                description=f"Found {total_volatile_funcs} volatile functions across {len(volatile_cells)} cells",
                probability=probability,
                severity=self.severity,
                location=f"Workbook-wide: {len(volatile_cells)} cells affected",
                details={
                    'total_volatile_functions': total_volatile_funcs,
                    'affected_cells': len(volatile_cells),
                    'total_formulas': total_formulas,
                    'high_impact_cells': len(high_impact_cells),
                    'medium_impact_cells': len(medium_impact_cells),
                    'volatile_cells': [
                        {
                            'cell': f"{cell['sheet']}!{cell['cell'].coordinate}",
                            'functions': cell['functions'],
                            'dependencies': cell['dependencies']
                        }
                        for cell in volatile_cells
                    ]
                },
                suggested_fix="Replace volatile functions with static alternatives where possible, or use iterative calculation settings."
            ))
        
        return results

    def _find_volatile_functions(self, formula: str) -> List[str]:
        """Find volatile functions in a formula."""
        volatile_functions = {
            # Time-based
            'NOW', 'TODAY', 'RAND', 'RANDBETWEEN',
            # Reference-based
            'OFFSET', 'INDIRECT', 'ADDRESS', 'COLUMN', 'ROW',
            # Information functions
            'CELL', 'INFO',
            # Database functions
            'DSUM', 'DCOUNT', 'DAVERAGE', 'DMAX', 'DMIN', 'DSTDEV', 'DVAR',
            # Other volatile functions
            'AREAS', 'COLUMNS', 'ROWS', 'HYPERLINK'
        }
        
        found_funcs = []
        formula_upper = formula.upper()
        
        for func in volatile_functions:
            if func in formula_upper:
                # Check if it's actually a function call (not part of another word)
                pattern = r'\b' + func + r'\s*\('
                import re
                if re.search(pattern, formula_upper):
                    found_funcs.append(func)
        
        return found_funcs

    def _count_dependencies(self, sheet, cell) -> int:
        """Count how many other cells reference this cell."""
        dependencies = 0
        cell_ref = cell.coordinate
        
        for row in sheet.iter_rows():
            for other_cell in row:
                if other_cell.data_type == 'f' and other_cell.value:
                    formula = str(other_cell.value)
                    if cell_ref in formula:
                        dependencies += 1
        
        return dependencies

    def _calculate_volatile_probability(self, total_volatile_funcs: int, affected_cells: int, total_formulas: int, volatile_cells: List[dict]) -> float:
        """Calculate probability of performance issues from volatile functions."""
        if total_formulas == 0:
            return 0.0
        
        # For small models, use absolute counts
        if total_formulas < 10:
            if total_volatile_funcs == 1:
                base_prob = 0.2
            elif total_volatile_funcs <= 3:
                base_prob = 0.4
            else:
                base_prob = 0.6
        else:
            # Base probability from frequency (more nuanced)
            volatile_ratio = total_volatile_funcs / max(total_formulas, 1)
            if volatile_ratio > 0.3:  # More than 30% of formulas are volatile
                base_prob = 0.8
            elif volatile_ratio > 0.15:  # 15-30% volatile
                base_prob = 0.6
            elif volatile_ratio > 0.05:  # 5-15% volatile
                base_prob = 0.4
            else:  # Less than 5% volatile
                base_prob = 0.2
        
        # Adjust based on impact (high dependency cells)
        high_impact_count = sum(1 for cell in volatile_cells if cell['dependencies'] > 5)
        if high_impact_count > 0:
            impact_factor = min(high_impact_count / max(affected_cells, 1), 1.0)
            base_prob += impact_factor * 0.7  # High impact can add up to 70%
        
        # Adjust based on model size (larger models are more sensitive)
        if total_formulas > 100:  # Large model
            base_prob += 0.4
        elif total_formulas > 50:  # Medium model
            base_prob += 0.2
        
        # Ensure high probability for high-impact or large model
        if high_impact_count > 0 or total_formulas > 100:
            base_prob = max(base_prob, 0.8)
        
        return min(base_prob, 1.0)


class CrossSheetReferenceErrorsDetector(ErrorDetector):
    """
    Detector for cross-sheet reference errors (broken, missing, or invalid references).
    
    Algorithm:
    1. Scan all formulas for cross-sheet references
    2. Check if referenced sheet exists
    3. Check if referenced cell/range exists in the target sheet
    4. Look for #REF! errors in formulas or cell values
    5. Calculate probability based on severity
    """
    def __init__(self):
        super().__init__(
            name="cross_sheet_reference_errors",
            description="References to cells in other sheets that have been moved or deleted",
            severity=ErrorSeverity.HIGH
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        results = []
        sheet_names = set(workbook.sheetnames)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:
                        formula = str(cell.value)
                        cross_refs = self._extract_cross_sheet_references(formula)
                        for ref in cross_refs:
                            ref_sheet, ref_cell = ref
                            if ref_sheet not in sheet_names:
                                # Missing sheet
                                results.append(ErrorDetectionResult(
                                    error_type=self.name,
                                    description=f"Reference to missing sheet '{ref_sheet}' in formula {cell.coordinate} on sheet {sheet_name}",
                                    probability=0.95,
                                    severity=self.severity,
                                    location=f"{sheet_name}!{cell.coordinate}",
                                    details={'formula': formula, 'missing_sheet': ref_sheet},
                                    suggested_fix="Restore the missing sheet or update the formula to reference an existing sheet."
                                ))
                            else:
                                # Check if cell exists in target sheet
                                target_sheet = workbook[ref_sheet]
                                if not self._cell_exists(target_sheet, ref_cell):
                                    # Check for #REF! in formula
                                    if '#REF!' in formula.upper():
                                        prob = 0.95
                                    else:
                                        prob = 0.7
                                    results.append(ErrorDetectionResult(
                                        error_type=self.name,
                                        description=f"Reference to missing cell/range '{ref_cell}' in sheet '{ref_sheet}' from formula {cell.coordinate} on sheet {sheet_name}",
                                        probability=prob,
                                        severity=self.severity,
                                        location=f"{sheet_name}!{cell.coordinate}",
                                        details={'formula': formula, 'target_sheet': ref_sheet, 'missing_cell': ref_cell},
                                        suggested_fix="Update the formula to reference a valid cell/range in the target sheet."
                                    ))
                                else:
                                    # Check if target cell is empty
                                    tgt_cell = target_sheet[ref_cell]  # openpyxl always returns a cell object
                                    if tgt_cell.value is None or tgt_cell.value == '':
                                        results.append(ErrorDetectionResult(
                                            error_type=self.name,
                                            description=f"Reference to empty cell '{ref_cell}' in sheet '{ref_sheet}' from formula {cell.coordinate} on sheet {sheet_name}",
                                            probability=0.3,
                                            severity=ErrorSeverity.LOW,
                                            location=f"{sheet_name}!{cell.coordinate}",
                                            details={'formula': formula, 'target_sheet': ref_sheet, 'empty_cell': ref_cell},
                                            suggested_fix="Check if the referenced cell should contain data."
                                        ))
                    # Also check for #REF! in cell value (not just formula)
                    if isinstance(cell.value, str) and '#REF!' in cell.value.upper():
                        results.append(ErrorDetectionResult(
                            error_type=self.name,
                            description=f"#REF! error in cell {cell.coordinate} on sheet {sheet_name}",
                            probability=0.95,
                            severity=self.severity,
                            location=f"{sheet_name}!{cell.coordinate}",
                            details={'cell': cell.coordinate, 'value': cell.value},
                            suggested_fix="Update the formula to reference a valid cell or sheet."
                        ))
        return results

    def _extract_cross_sheet_references(self, formula: str) -> List[tuple]:
        # Extract references like 'Sheet2'!A1 or Sheet3!B2
        import re
        pattern = r"(?:'([^']+)'|([A-Za-z0-9_]+))!([A-Za-z]+[0-9]+)"
        matches = re.findall(pattern, formula)
        refs = []
        for match in matches:
            sheet = match[0] if match[0] else match[1]
            cell = match[2]
            refs.append((sheet, cell))
        return refs

    def _cell_exists(self, sheet, cell_ref: str) -> bool:
        from openpyxl.utils import coordinate_to_tuple
        try:
            row, col = coordinate_to_tuple(cell_ref)
            max_row = sheet.max_row
            max_col = sheet.max_column
            return 1 <= row <= max_row and 1 <= col <= max_col
        except Exception:
            return False


class DataTypeInconsistenciesInLookupTablesDetector(ErrorDetector):
    """
    Detector for data type inconsistencies in lookup tables (e.g., numbers stored as text).
    
    Algorithm:
    1. Scan for lookup functions (VLOOKUP, HLOOKUP, XLOOKUP, MATCH, INDEX)
    2. Extract lookup ranges
    3. Analyze data types in lookup key columns/rows
    4. Flag if mixed types (number/text/date)
    5. Calculate probability based on proportion of inconsistencies
    """
    def __init__(self):
        super().__init__(
            name="data_type_inconsistencies_in_lookup_tables",
            description="Mixed data types in lookup tables (numbers stored as text, etc.)",
            severity=ErrorSeverity.HIGH
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        results = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:
                        formula = str(cell.value)
                        lookup_ranges = self._extract_lookup_ranges(formula)
                        for rng in lookup_ranges:
                            analysis = self._analyze_lookup_range(sheet, rng)
                            if analysis['mixed_types']:
                                probability = self._calculate_probability(analysis)
                                if probability > 0:
                                    results.append(ErrorDetectionResult(
                                        error_type=self.name,
                                        description=f"Mixed data types detected in lookup range {rng} on sheet {sheet_name}",
                                        probability=probability,
                                        severity=self.severity,
                                        location=f"{sheet_name}!{rng}",
                                        details=analysis,
                                        suggested_fix="Convert all lookup keys to a consistent data type (all numbers or all text)."
                                    ))
        return results

    def _extract_lookup_ranges(self, formula: str) -> List[str]:
        # Extract ranges from lookup functions (VLOOKUP, HLOOKUP, XLOOKUP, MATCH, INDEX)
        import re
        # Simple pattern for ranges like A1:B10, Sheet2!A1:B10, etc.
        pattern = r"([A-Za-z0-9_']+!|)([A-Za-z]+[0-9]+:[A-Za-z]+[0-9]+)"
        matches = re.findall(pattern, formula)
        ranges = []
        for match in matches:
            prefix = match[0]
            rng = match[1]
            ranges.append(prefix + rng)
        return ranges

    def _analyze_lookup_range(self, sheet, rng: str) -> dict:
        from openpyxl.utils import range_boundaries
        # Only analyze ranges on the current sheet for now
        if '!' in rng:
            sheet_name, rng = rng.split('!')
            if sheet.title != sheet_name.replace("'", ""):
                return {'mixed_types': False}
        try:
            min_col, min_row, max_col, max_row = range_boundaries(rng)
        except Exception:
            return {'mixed_types': False}
        type_counts = {'number': 0, 'text': 0, 'date': 0, 'other': 0, 'numeric_text': 0}
        total = 0
        all_text = True
        all_numeric_text = True
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if cell.value is None:
                    continue
                total += 1
                if isinstance(cell.value, (int, float)):
                    type_counts['number'] += 1
                    all_text = False
                elif self._is_date(cell):
                    type_counts['date'] += 1
                    all_text = False
                elif isinstance(cell.value, str):
                    if self._is_numeric_string(cell.value):
                        type_counts['numeric_text'] += 1
                    else:
                        type_counts['text'] += 1
                        all_numeric_text = False
                else:
                    type_counts['other'] += 1
                    all_text = False
                    all_numeric_text = False
        # If all non-empty cells are strings and all are numeric strings, treat as numbers
        if total > 0 and all_text and all_numeric_text and type_counts['numeric_text'] > 0:
            type_counts['number'] = type_counts['numeric_text']
            type_counts['numeric_text'] = 0
        else:
            # Otherwise, count numeric_text as text
            type_counts['text'] += type_counts['numeric_text']
            type_counts['numeric_text'] = 0
        # Mixed types if more than one type is present (excluding 'other')
        nonzero_types = [k for k, v in type_counts.items() if v > 0 and k not in ('other', 'numeric_text')]
        mixed_types = len(nonzero_types) > 1
        return {
            'type_counts': type_counts,
            'total': total,
            'mixed_types': mixed_types
        }

    def _is_number(self, value) -> bool:
        return isinstance(value, (int, float))

    def _is_numeric_string(self, value) -> bool:
        if not isinstance(value, str):
            return False
        try:
            float(value.replace(',', ''))
            return True
        except Exception:
            return False

    def _is_date(self, cell) -> bool:
        return hasattr(cell, 'is_date') and cell.is_date

    def _calculate_probability(self, analysis: dict) -> float:
        if analysis['total'] == 0:
            return 0.0
        type_counts = analysis['type_counts']
        majority_type = max((k for k in type_counts if k != 'other'), key=lambda k: type_counts[k], default=None)
        inconsistent = sum(v for k, v in type_counts.items() if k != majority_type and k != 'other')
        ratio = inconsistent / analysis['total'] if analysis['total'] else 0
        if ratio > 0.1:
            return 0.9
        elif ratio > 0.01:
            return 0.5
        elif ratio > 0:
            return 0.2
        return 0.0


class ConditionalFormattingOverlapConflictsDetector(ErrorDetector):
    """
    Detector for overlapping/conflicting conditional formatting rules.
    
    Algorithm:
    1. Extract all conditional formatting rules and their ranges
    2. Detect overlapping ranges
    3. Analyze rule types and formats for conflicts
    4. Calculate probability based on severity of overlap/conflict
    """
    def __init__(self):
        super().__init__(
            name="conditional_formatting_overlap_conflicts",
            description="Multiple conditional formatting rules that conflict or overlap",
            severity=ErrorSeverity.MEDIUM
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        results = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            # 1. Extract all conditional formatting rules and their ranges
            cf_rules = self._extract_conditional_formatting_rules(sheet)
            # Compare all pairs, including rules within the same range
            n = len(cf_rules)
            for i in range(n):
                for j in range(i + 1, n):
                    rule1 = cf_rules[i]
                    rule2 = cf_rules[j]
                    overlap_cells = rule1['cells'] & rule2['cells']
                    if not overlap_cells:
                        continue
                    # 3. Analyze for conflicts
                    conflict_type, probability = self._analyze_conflict(rule1, rule2)
                    if probability > 0:
                        results.append(ErrorDetectionResult(
                            error_type=self.name,
                            description=f"Conditional formatting overlap/conflict between rules on {sheet_name}: {rule1['range']} and {rule2['range']}",
                            probability=probability,
                            severity=self.severity,
                            location=f"{sheet_name}!{rule1['range']} & {rule2['range']}",
                            details={
                                'rule1': rule1,
                                'rule2': rule2,
                                'overlap_cells': list(overlap_cells),
                                'conflict_type': conflict_type
                            },
                            suggested_fix="Review overlapping conditional formatting rules and resolve conflicts."
                        ))
        return results

    def _extract_conditional_formatting_rules(self, sheet) -> List[dict]:
        # openpyxl stores conditional formatting in sheet.conditional_formatting
        rules = []
        cf = getattr(sheet, 'conditional_formatting', None)
        if cf is None:
            return rules
        # Iterate over cf._cf_rules.items(), use cf_range.sqref for the range string
        for cf_range, rule_list in cf._cf_rules.items():
            range_str = str(getattr(cf_range, 'sqref', '')) if hasattr(cf_range, 'sqref') else None
            if not range_str:
                continue
            from openpyxl.utils import range_boundaries, get_column_letter
            try:
                min_col, min_row, max_col, max_row = range_boundaries(range_str)
            except Exception:
                continue
            cells = set()
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cell = f"{get_column_letter(col)}{row}"
                    cells.add(cell)
            for rule in rule_list:
                rules.append({
                    'range': range_str,
                    'cells': cells,
                    'type': getattr(rule, 'type', None),
                    'formula': getattr(rule, 'formula', None),
                    'dxf': getattr(rule, 'dxf', None),
                    'priority': getattr(rule, 'priority', None),
                    'rule_obj': rule
                })
        return rules

    def _analyze_conflict(self, rule1: dict, rule2: dict) -> tuple:
        # Check for type and format conflicts
        type1 = rule1['type']
        type2 = rule2['type']
        dxf1 = rule1['dxf']
        dxf2 = rule2['dxf']
        # High probability: both set fill/font and are different
        if dxf1 and dxf2:
            fill1 = getattr(dxf1, 'fill', None)
            fill2 = getattr(dxf2, 'fill', None)
            font1 = getattr(dxf1, 'font', None)
            font2 = getattr(dxf2, 'font', None)
            if (fill1 and fill2 and fill1 != fill2) or (font1 and font2 and font1 != font2):
                return ('conflicting_format', 0.9)
        # Medium probability: different types (e.g., color scale vs formula)
        if type1 != type2:
            return ('different_types', 0.6)
        # Low probability: overlap but compatible
        return ('overlap', 0.3)


class ExternalDataConnectionFailuresDetector(ErrorDetector):
    """
    Detector for external data connection failures (broken or outdated links).
    
    Algorithm:
    1. Scan for external links and data connections
    2. Check if target file/database exists (if accessible)
    3. Check for broken/unavailable links in workbook metadata
    4. Check for error values in cells that depend on external data
    5. Check last refresh date/time if available
    6. Calculate probability based on severity
    """
    def __init__(self):
        super().__init__(
            name="external_data_connection_failures",
            description="Links to external databases or files that are broken or outdated",
            severity=ErrorSeverity.HIGH
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        import os
        results = []
        # 1. Scan for external links (to other workbooks/files)
        external_links = getattr(workbook, 'external_links', [])
        for link in external_links:
            target = getattr(link, 'target', None)
            if not target:
                continue
            # 2. Check if the target file exists (if it's a file path)
            if os.path.isfile(target):
                probability = 0.2  # Low probability if file exists
            else:
                probability = 0.95  # High probability if file is missing
            results.append(ErrorDetectionResult(
                error_type=self.name,
                description=f"External link to file '{target}' is {'missing' if probability > 0.5 else 'present'}",
                probability=probability,
                severity=self.severity if probability > 0.5 else ErrorSeverity.LOW,
                location=f"external_link:{target}",
                details={'target': target, 'link': str(link)},
                suggested_fix="Update or remove the broken external link."
            ))
        # 2b. Scan for data connections (databases, web queries, etc.)
        connections = getattr(workbook, 'connections', None)
        if connections:
            for conn in connections:
                # Check last refresh date/time if available
                last_refresh = getattr(conn, 'last_refresh', None)
                if last_refresh:
                    import datetime
                    if isinstance(last_refresh, str):
                        try:
                            last_refresh = datetime.datetime.fromisoformat(last_refresh)
                        except Exception:
                            last_refresh = None
                    if last_refresh:
                        days_since = (datetime.datetime.now() - last_refresh).days
                        if days_since > 30:
                            probability = 0.6
                        else:
                            probability = 0.2
                    else:
                        probability = 0.3
                else:
                    probability = 0.3
                results.append(ErrorDetectionResult(
                    error_type=self.name,
                    description=f"External data connection '{getattr(conn, 'name', 'unknown')}' may be outdated or unverifiable",
                    probability=probability,
                    severity=ErrorSeverity.MEDIUM if probability > 0.3 else ErrorSeverity.LOW,
                    location=f"connection:{getattr(conn, 'name', 'unknown')}",
                    details={'connection': str(conn)},
                    suggested_fix="Check the data connection and refresh if needed."
                ))
        # 3. Check for error values in cells that depend on external data
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.upper() in {'#REF!', '#VALUE!', '#N/A'}:
                        results.append(ErrorDetectionResult(
                            error_type=self.name,
                            description=f"Error value '{cell.value}' in cell {cell.coordinate} on sheet {sheet_name} may be due to external data connection failure",
                            probability=0.8,
                            severity=self.severity,
                            location=f"{sheet_name}!{cell.coordinate}",
                            details={'cell': cell.coordinate, 'value': cell.value},
                            suggested_fix="Check external data sources and update or fix broken links."
                        ))
        return results


class PrecisionErrorsInFinancialCalculationsDetector(ErrorDetector):
    """
    Detector for floating-point precision errors in financial calculations.
    
    Algorithm:
    1. Scan for formulas with decimal arithmetic or financial functions
    2. Detect absence of rounding in such formulas
    3. Analyze for known precision issues (chained arithmetic, subtraction of nearly equal numbers)
    4. Calculate probability based on severity
    """
    def __init__(self):
        super().__init__(
            name="precision_errors_in_financial_calculations",
            description="Floating-point precision errors in financial calculations",
            severity=ErrorSeverity.MEDIUM
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        import re
        results = []
        financial_funcs = {'PMT', 'NPV', 'IRR', 'FV', 'PV', 'RATE', 'XNPV', 'XIRR', 'MIRR', 'DURATION', 'YIELD', 'COUPON', 'PRICE', 'DISC', 'TBILL', 'SLN', 'SYD', 'DB', 'DDB', 'VDB', 'AMORDEGRC', 'AMORLINC'}
        rounding_funcs = {'ROUND', 'ROUNDUP', 'ROUNDDOWN', 'MROUND', 'TRUNC', 'INT', 'CEILING', 'FLOOR'}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:
                        formula = str(cell.value).upper()
                        # 1. Check for financial functions
                        uses_financial_func = any(func in formula for func in financial_funcs)
                        uses_rounding = any(func in formula for func in rounding_funcs)
                        # 2. Check for decimal arithmetic (division, multiplication, subtraction, decimal point)
                        has_decimal_point = '.' in formula
                        cell_refs = re.findall(r"[A-Z][0-9]+", formula)
                        any_float = False
                        for ref in cell_refs:
                            try:
                                ref_cell = sheet[ref]
                                if isinstance(ref_cell.value, float):
                                    any_float = True
                                    break
                            except Exception:
                                continue
                        has_decimal_arithmetic = has_decimal_point or any_float
                        # 3. Check for subtraction of nearly equal numbers (e.g., A1-A2)
                        subtraction_matches = re.findall(r"([A-Z][0-9]+)\s*-\s*([A-Z][0-9]+)", formula)
                        # 4. Check for chained arithmetic (multiple operators)
                        operator_count = formula.count('+') + formula.count('-') + formula.count('*') + formula.count('/')
                        # 5. Ignore integer-only calculations
                        cell_refs = re.findall(r"[A-Z][0-9]+", formula)
                        all_integer = True
                        for ref in cell_refs:
                            try:
                                ref_cell = sheet[ref]
                                if not isinstance(ref_cell.value, int):
                                    all_integer = False
                                    break
                            except Exception:
                                all_integer = False
                                break
                        is_integer_only = all_integer and not has_decimal_arithmetic and not uses_financial_func
                        if is_integer_only:
                            continue
                        # 6. Probability calculation
                        probability = 0.0
                        if uses_financial_func and not uses_rounding:
                            probability = 0.9
                        elif operator_count >= 3 and not uses_rounding:
                            probability = 0.8
                        elif subtraction_matches and not uses_rounding:
                            probability = 0.8
                        elif has_decimal_arithmetic and not uses_rounding:
                            probability = 0.6
                        elif uses_rounding and operator_count >= 2:
                            probability = 0.3
                        if probability > 0:
                            results.append(ErrorDetectionResult(
                                error_type=self.name,
                                description=f"Potential precision error in formula {cell.coordinate} on sheet {sheet_name}",
                                probability=probability,
                                severity=self.severity if probability >= 0.6 else ErrorSeverity.LOW,
                                location=f"{sheet_name}!{cell.coordinate}",
                                details={
                                    'formula': formula,
                                    'uses_financial_func': uses_financial_func,
                                    'uses_rounding': uses_rounding,
                                    'has_decimal_arithmetic': has_decimal_arithmetic,
                                    'subtraction_matches': subtraction_matches,
                                    'operator_count': operator_count
                                },
                                suggested_fix="Use explicit rounding (e.g., ROUND) in all financial and decimal calculations."
                            ))
        return results


class IncompleteDragFormulaDetector(ErrorDetector):
    """
    Detector for incomplete drag formulas (formula cutoff / range incomplete formula errors).
    
    Algorithm:
    1. For each column, scan for contiguous blocks of formulas
    2. Identify expected data range (based on adjacent columns or max data row)
    3. Detect cutoffs (formulas stop before end of data) or gaps (missing formulas in middle)
    4. Calculate probability based on severity
    """
    def __init__(self):
        super().__init__(
            name="incomplete_drag_formula",
            description="Formula drag/copy stopped short, leaving cells without formulas (formula cutoff)",
            severity=ErrorSeverity.HIGH
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        results = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            # For each column, scan for formula blocks
            for col in range(1, max_col + 1):
                formula_rows = []
                data_rows = []
                for row in range(1, max_row + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value is not None:
                        data_rows.append(row)
                    if cell.data_type == 'f' and cell.value:
                        formula_rows.append(row)
                if not formula_rows or not data_rows:
                    continue
                # Expected range: from min(data_rows) to max(data_rows)
                expected_rows = set(range(min(data_rows), max(data_rows) + 1))
                formula_rows_set = set(formula_rows)
                missing_formula_rows = expected_rows - formula_rows_set
                if not missing_formula_rows:
                    continue  # No cutoff/gap
                # Probability calculation
                cutoff_at_end = max(formula_rows) < max(data_rows)
                gap_in_middle = any(row > min(formula_rows) and row < max(formula_rows) for row in missing_formula_rows)
                if gap_in_middle:
                    probability = 0.9
                elif cutoff_at_end and len(missing_formula_rows) > 1:
                    probability = 0.8
                elif cutoff_at_end:
                    probability = 0.6
                else:
                    probability = 0.3
                if probability > 0:
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(col)
                    results.append(ErrorDetectionResult(
                        error_type=self.name,
                        description=f"Incomplete drag/copy of formula in column {col_letter} on sheet {sheet_name}; missing formulas at rows {sorted(missing_formula_rows)}",
                        probability=probability,
                        severity=self.severity if probability >= 0.6 else ErrorSeverity.LOW,
                        location=f"{sheet_name}!{col_letter}{min(data_rows)}:{col_letter}{max(data_rows)}",
                        details={
                            'column': col_letter,
                            'missing_formula_rows': sorted(missing_formula_rows),
                            'formula_rows': formula_rows,
                            'data_rows': data_rows
                        },
                        suggested_fix="Drag/copy the formula to all data rows; check for gaps or cutoffs."
                    ))
        return results


class FalseRangeEndDetectionDetector(ErrorDetector):
    """
    Detector for false range end detection (empty cell trap).
    
    Algorithm:
    1. For each column, find contiguous blocks of non-empty cells
    2. Identify empty cells in the middle of a data range
    3. Check if formulas only cover up to the first empty cell
    4. Flag if data exists after the gap but formulas stop at the gap
    5. Calculate probability based on severity
    """
    def __init__(self):
        super().__init__(
            name="false_range_end_detection",
            description="Empty cell in the middle of a data range causes formulas to stop early (false range end)",
            severity=ErrorSeverity.HIGH
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        results = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            for col in range(1, max_col + 1):
                data_rows = []
                formula_rows = []
                for row in range(1, max_row + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value is not None:
                        data_rows.append(row)
                    if cell.data_type == 'f' and cell.value:
                        formula_rows.append(row)
                if not data_rows or not formula_rows:
                    continue
                # Find the first empty cell in the data range
                min_data_row = min(data_rows)
                max_data_row = max(data_rows)
                empty_in_middle = None
                for row in range(min_data_row, max_data_row + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value is None:
                        empty_in_middle = row
                        break
                if empty_in_middle is None:
                    continue  # No gap in the middle
                # Check if there is data after the gap
                data_after_gap = [row for row in range(empty_in_middle + 1, max_data_row + 1) if sheet.cell(row=row, column=col).value is not None]
                if not data_after_gap:
                    continue  # No data after gap
                # Find missing formulas after the gap
                formula_rows_set = set(formula_rows)
                missing_formula_rows = set(row for row in range(empty_in_middle + 1, max_data_row + 1) if row not in formula_rows_set and sheet.cell(row=row, column=col).value is not None)
                if not missing_formula_rows:
                    continue
                # Probability calculation
                gap_size = len(missing_formula_rows)
                if gap_size > 2:
                    probability = 0.9
                elif gap_size > 0:
                    probability = 0.6
                else:
                    probability = 0.3
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(col)
                results.append(ErrorDetectionResult(
                    error_type=self.name,
                    description=f"False range end detected in column {col_letter} on sheet {sheet_name}; missing formulas after empty cell at row {empty_in_middle}: {sorted(missing_formula_rows)}",
                    probability=probability,
                    severity=self.severity if probability >= 0.6 else ErrorSeverity.LOW,
                    location=f"{sheet_name}!{col_letter}{min_data_row}:{col_letter}{max_data_row}",
                    details={
                        'column': col_letter,
                        'empty_in_middle': empty_in_middle,
                        'data_after_gap': data_after_gap,
                        'missing_formula_rows': sorted(missing_formula_rows)
                    },
                    suggested_fix="Check for empty cells in the middle of data ranges and ensure formulas cover all data rows."
                ))
        return results


class PartialFormulaPropagationDetector(ErrorDetector):
    def __init__(self):
        super().__init__(
            name="partial_formula_propagation",
            description="Cells in a data range are missing formulas that are present in most other cells (partial formula propagation)",
            severity=ErrorSeverity.HIGH
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        results = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            for col in range(1, max_col + 1):
                formula_rows = []
                non_formula_rows = []
                formulas = {}
                for row in range(1, max_row + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.data_type == 'f' and cell.value:
                        formula_rows.append(row)
                        formulas[row] = cell.value
                    elif cell.value is not None:
                        non_formula_rows.append(row)
                total_data_rows = len(formula_rows) + len(non_formula_rows)
                if total_data_rows < 5:
                    continue  # skip small ranges
                if len(formula_rows) / total_data_rows < 0.7:
                    continue  # skip if not mostly formulas
                # Find non-formula cells surrounded by formulas or at the edge
                missing_candidates = []
                for row in non_formula_rows:
                    prev_formula = any(r < row for r in formula_rows)
                    next_formula = any(r > row for r in formula_rows)
                    if (prev_formula and next_formula) or row == 1 or row == max_row:
                        missing_candidates.append(row)
                if not missing_candidates:
                    continue
                # Use the most common formula as the expected one
                from collections import Counter
                formula_counter = Counter(formulas.values())
                if not formula_counter:
                    continue
                expected_formula, _ = formula_counter.most_common(1)[0]
                # Probability: more missing = lower, more surrounded = higher, edge = 0.5
                for row in missing_candidates:
                    if row == 1 or row == max_row:
                        probability = 0.5
                        severity = ErrorSeverity.MEDIUM
                    else:
                        probability = 0.8 if len(missing_candidates) <= 2 else 0.6
                        severity = self.severity if probability >= 0.7 else ErrorSeverity.MEDIUM
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(col)
                    results.append(ErrorDetectionResult(
                        error_type=self.name,
                        description=f"Cell {col_letter}{row} on sheet {sheet_name} is missing a formula, but most cells in this column have '{expected_formula}'",
                        probability=probability,
                        severity=severity,
                        location=f"{sheet_name}!{col_letter}{row}",
                        details={
                            'column': col_letter,
                            'row': row,
                            'expected_formula': expected_formula,
                            'missing_candidates': missing_candidates
                        },
                        suggested_fix=f"Copy the formula '{expected_formula}' to cell {col_letter}{row} if appropriate."
                    ))
        return results


class FormulaBoundaryMismatchDetector(ErrorDetector):
    def __init__(self):
        super().__init__(
            name="formula_boundary_mismatch",
            description="Aggregation formula references a range that does not cover all data (range misalignment)",
            severity=ErrorSeverity.HIGH
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        results = []
        agg_funcs = ["SUM", "AVERAGE", "COUNT", "COUNTA", "MAX", "MIN"]
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.data_type == 'f' and cell.value:
                        formula = str(cell.value).upper()
                        for func in agg_funcs:
                            if formula.startswith(f"={func}"):
                                # Extract range, e.g., =SUM(A1:A50)
                                import re
                                match = re.search(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", formula)
                                if not match:
                                    continue
                                start_col, start_row, end_col, end_row = match.groups()
                                if start_col != end_col:
                                    continue  # Only handle single-column ranges for now
                                col_idx = openpyxl.utils.column_index_from_string(start_col)
                                start_row = int(start_row)
                                end_row = int(end_row)
                                # Find actual data extent in this column
                                data_rows = [r for r in range(1, max_row + 1) if sheet.cell(row=r, column=col_idx).value is not None]
                                if not data_rows:
                                    continue
                                max_data_row = max(data_rows)
                                if max_data_row > end_row:
                                    # Data exists beyond the referenced range
                                    extra_data_count = max_data_row - end_row
                                    total_data_count = max_data_row - start_row + 1
                                    probability = min(0.9, 0.5 + 0.4 * (extra_data_count / total_data_count))
                                    from openpyxl.utils import get_column_letter
                                    col_letter = get_column_letter(col_idx)
                                    results.append(ErrorDetectionResult(
                                        error_type=self.name,
                                        description=f"Formula in {sheet_name}!{get_column_letter(col)}{row} references {col_letter}{start_row}:{col_letter}{end_row}, but data extends to row {max_data_row}.",
                                        probability=probability,
                                        severity=self.severity if probability >= 0.7 else ErrorSeverity.MEDIUM,
                                        location=f"{sheet_name}!{get_column_letter(col)}{row}",
                                        details={
                                            'formula': formula,
                                            'referenced_range': f"{col_letter}{start_row}:{col_letter}{end_row}",
                                            'max_data_row': max_data_row,
                                            'extra_data_count': extra_data_count
                                        },
                                        suggested_fix=f"Check if the aggregation formula should include data up to row {max_data_row}."
                                    ))
        return results


class CopyPasteFormulaGapsDetector(ErrorDetector):
    def __init__(self):
        super().__init__(
            name="copy_paste_formula_gaps",
            description="Gaps in formula sequences due to copy-paste errors (missing formulas in expected locations)",
            severity=ErrorSeverity.HIGH
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        results = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            for col in range(1, max_col + 1):
                formula_rows = []
                non_formula_rows = []
                formulas = {}
                for row in range(1, max_row + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.data_type == 'f' and cell.value:
                        formula_rows.append(row)
                        formulas[row] = cell.value
                    elif cell.value is not None:
                        non_formula_rows.append(row)
                if len(formula_rows) < 3:
                    continue  # Need at least 3 formulas to detect gaps
                # Find gaps in formula sequences
                formula_rows.sort()
                gaps = []
                for i in range(len(formula_rows) - 1):
                    current_row = formula_rows[i]
                    next_row = formula_rows[i + 1]
                    if next_row - current_row > 1:
                        # Check if there are non-formula cells in the gap
                        gap_cells = [r for r in range(current_row + 1, next_row) if r in non_formula_rows]
                        if gap_cells:
                            gaps.append((current_row, next_row, gap_cells))
                for start_row, end_row, gap_cells in gaps:
                    # Check if surrounding formulas are similar
                    start_formula = formulas[start_row]
                    end_formula = formulas[end_row]
                    if self._are_formulas_similar(start_formula, end_formula):
                        probability = 0.8 if len(gap_cells) <= 2 else 0.6
                        from openpyxl.utils import get_column_letter
                        col_letter = get_column_letter(col)
                        results.append(ErrorDetectionResult(
                            error_type=self.name,
                            description=f"Formula gap detected in column {col_letter} on sheet {sheet_name} between rows {start_row} and {end_row}; missing formulas in rows {gap_cells}.",
                            probability=probability,
                            severity=self.severity if probability >= 0.7 else ErrorSeverity.MEDIUM,
                            location=f"{sheet_name}!{col_letter}{start_row}:{col_letter}{end_row}",
                            details={
                                'column': col_letter,
                                'start_row': start_row,
                                'end_row': end_row,
                                'gap_cells': gap_cells,
                                'start_formula': start_formula,
                                'end_formula': end_formula
                            },
                            suggested_fix=f"Check for missing formulas in rows {gap_cells}; consider copying the formula pattern from adjacent cells."
                        ))
        return results

    def _are_formulas_similar(self, formula1: str, formula2: str) -> bool:
        """Check if two formulas follow a similar pattern (e.g., incrementing row references)."""
        # Simple heuristic: check if formulas have similar structure
        # This could be enhanced with more sophisticated pattern matching
        if not formula1 or not formula2:
            return False
        # Remove cell references and compare structure
        import re
        clean1 = re.sub(r'[A-Z]+\d+', 'CELL', formula1)
        clean2 = re.sub(r'[A-Z]+\d+', 'CELL', formula2)
        return clean1 == clean2


class FormulaRangeVsDataRangeDiscrepancyDetector(ErrorDetector):
    def __init__(self):
        super().__init__(
            name="formula_range_vs_data_range_discrepancy",
            description="Lookup formula references a range that does not cover all data (range vs data mismatch)",
            severity=ErrorSeverity.HIGH
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        results = []
        lookup_funcs = ["VLOOKUP", "HLOOKUP", "XLOOKUP", "INDEX", "MATCH"]
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.data_type == 'f' and cell.value:
                        formula = str(cell.value).upper()
                        for func in lookup_funcs:
                            if func in formula:
                                # Extract range, e.g., VLOOKUP(A1,A1:B50,2)
                                import re
                                match = re.search(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", formula)
                                if not match:
                                    continue
                                start_col, start_row, end_col, end_row = match.groups()
                                col_start_idx = openpyxl.utils.column_index_from_string(start_col)
                                col_end_idx = openpyxl.utils.column_index_from_string(end_col)
                                start_row = int(start_row)
                                end_row = int(end_row)
                                # Check if data extends beyond the referenced range
                                data_beyond = False
                                max_data_row = end_row
                                max_data_col = col_end_idx
                                for r in range(end_row + 1, max_row + 1):
                                    for c in range(col_start_idx, col_end_idx + 1):
                                        if sheet.cell(row=r, column=c).value is not None:
                                            data_beyond = True
                                            max_data_row = max(max_data_row, r)
                                for c in range(col_end_idx + 1, max_col + 1):
                                    for r in range(start_row, end_row + 1):
                                        if sheet.cell(row=r, column=c).value is not None:
                                            data_beyond = True
                                            max_data_col = max(max_data_col, c)
                                if data_beyond:
                                    extra_rows = max_data_row - end_row
                                    extra_cols = max_data_col - col_end_idx
                                    probability = min(0.9, 0.5 + 0.4 * ((extra_rows + extra_cols) / (end_row - start_row + 1 + col_end_idx - col_start_idx + 1)))
                                    from openpyxl.utils import get_column_letter
                                    start_col_letter = get_column_letter(col_start_idx)
                                    end_col_letter = get_column_letter(col_end_idx)
                                    max_data_col_letter = get_column_letter(max_data_col)
                                    results.append(ErrorDetectionResult(
                                        error_type=self.name,
                                        description=f"Lookup formula in {sheet_name}!{get_column_letter(col)}{row} references {start_col_letter}{start_row}:{end_col_letter}{end_row}, but data extends to {max_data_col_letter}{max_data_row}.",
                                        probability=probability,
                                        severity=self.severity if probability >= 0.7 else ErrorSeverity.MEDIUM,
                                        location=f"{sheet_name}!{get_column_letter(col)}{row}",
                                        details={
                                            'formula': formula,
                                            'referenced_range': f"{start_col_letter}{start_row}:{end_col_letter}{end_row}",
                                            'max_data_row': max_data_row,
                                            'max_data_col': max_data_col,
                                            'extra_rows': extra_rows,
                                            'extra_cols': extra_cols
                                        },
                                        suggested_fix=f"Check if the lookup range should include data up to {max_data_col_letter}{max_data_row}."
                                    ))
        return results


class InconsistentFormulaApplicationDetector(ErrorDetector):
    def __init__(self):
        super().__init__(
            name="inconsistent_formula_application",
            description="Mixed formulas and hardcoded values in the same range (inconsistent calculation methods)",
            severity=ErrorSeverity.HIGH
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        results = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            for col in range(1, max_col + 1):
                formula_rows = []
                hardcoded_rows = []
                formulas = {}
                hardcoded_values = {}
                for row in range(1, max_row + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.data_type == 'f' and cell.value:
                        formula_rows.append(row)
                        formulas[row] = cell.value
                    elif cell.value is not None:
                        hardcoded_rows.append(row)
                        hardcoded_values[row] = cell.value
                total_data_rows = len(formula_rows) + len(hardcoded_rows)
                if total_data_rows < 3:
                    continue  # Skip small ranges
                if len(formula_rows) == 0 or len(hardcoded_rows) == 0:
                    continue  # Skip if all one type
                # Calculate proportions
                formula_ratio = len(formula_rows) / total_data_rows
                hardcoded_ratio = len(hardcoded_rows) / total_data_rows
                # Flag if both types are significant (at least 20% each)
                if formula_ratio >= 0.2 and hardcoded_ratio >= 0.2:
                    # Probability: more balanced = higher probability
                    if 0.4 <= formula_ratio <= 0.6:
                        probability = 0.9  # Very balanced mix
                    elif 0.3 <= formula_ratio <= 0.7:
                        probability = 0.7  # Moderately balanced
                    else:
                        probability = 0.5  # Less balanced but still mixed
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(col)
                    min_row = min(min(formula_rows), min(hardcoded_rows))
                    max_row = max(max(formula_rows), max(hardcoded_rows))
                    example_formula = next(iter(formulas.values()))
                    example_hardcoded = next(iter(hardcoded_values.values()))
                    results.append(ErrorDetectionResult(
                        error_type=self.name,
                        description=f"Mixed formulas and hardcoded values in column {col_letter} on sheet {sheet_name}; {len(formula_rows)} formulas and {len(hardcoded_rows)} hardcoded values.",
                        probability=probability,
                        severity=self.severity if probability >= 0.7 else ErrorSeverity.MEDIUM,
                        location=f"{sheet_name}!{col_letter}{min_row}:{col_letter}{max_row}",
                        details={
                            'column': col_letter,
                            'formula_rows': sorted(formula_rows),
                            'hardcoded_rows': sorted(hardcoded_rows),
                            'formula_ratio': formula_ratio,
                            'hardcoded_ratio': hardcoded_ratio,
                            'example_formula': example_formula,
                            'example_hardcoded': example_hardcoded
                        },
                        suggested_fix="Check for mixed formulas and hardcoded values in this range; consider standardizing calculation logic."
                    ))
        return results


class MissingDollarSignAnchorsDetector(ErrorDetector):
    def __init__(self):
        super().__init__(
            name="missing_dollar_sign_anchors",
            description="Formulas missing dollar sign anchors for references that should remain fixed",
            severity=ErrorSeverity.HIGH
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        results = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.data_type == 'f' and cell.value:
                        formula = str(cell.value)
                        # Look for cell references without dollar signs
                        import re
                        cell_refs = re.findall(r'[A-Z]+\d+', formula)
                        for ref in cell_refs:
                            # Check if this reference should be anchored
                            if self._should_be_anchored(sheet, ref, row, col):
                                # Check if it's already anchored
                                if not self._is_anchored(formula, ref):
                                    probability = self._calculate_anchor_probability(sheet, ref, row, col)
                                    if probability > 0.5:
                                        from openpyxl.utils import get_column_letter
                                        col_letter = get_column_letter(col)
                                        expected_formula = self._suggest_anchored_formula(formula, ref)
                                        results.append(ErrorDetectionResult(
                                            error_type=self.name,
                                            description=f"Missing dollar sign anchor for reference {ref} in formula at {sheet_name}!{col_letter}{row}",
                                            probability=probability,
                                            severity=self.severity if probability >= 0.7 else ErrorSeverity.MEDIUM,
                                            location=f"{sheet_name}!{col_letter}{row}",
                                            details={
                                                'formula': formula,
                                                'reference': ref,
                                                'expected_formula': expected_formula,
                                                'current_row': row,
                                                'current_col': col
                                            },
                                            suggested_fix=f"Consider anchoring reference {ref} with dollar signs: {expected_formula}"
                                        ))
        return results

    def _should_be_anchored(self, sheet, ref: str, current_row: int, current_col: int) -> bool:
        """Check if a cell reference should be anchored based on its usage pattern."""
        # Extract row and column from reference
        import re
        match = re.match(r'([A-Z]+)(\d+)', ref)
        if not match:
            return False
        col_str, row_str = match.groups()
        ref_row = int(row_str)
        ref_col = openpyxl.utils.column_index_from_string(col_str)
        
        # Check if reference is to a header row (row 1)
        if ref_row == 1:
            return True
        
        # Check if reference is to a constant value (same value in multiple cells)
        ref_cell = sheet.cell(row=ref_row, column=ref_col)
        if ref_cell.value is None:
            return False
        
        # Check if this value appears in multiple cells in the same column
        constant_count = 0
        total_cells = 0
        for r in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=r, column=ref_col)
            if cell.value is not None:
                total_cells += 1
                if cell.value == ref_cell.value:
                    constant_count += 1
        
        # If the same value appears in more than 2 cells and represents more than 50% of non-empty cells
        if constant_count > 2 and total_cells > 0 and (constant_count / total_cells) > 0.5:
            return True
        
        return False

    def _is_anchored(self, formula: str, ref: str) -> bool:
        """Check if a reference is already anchored with dollar signs."""
        # Look for anchored version of the reference
        import re
        anchored_pattern = re.escape(ref).replace('\\', '\\\\')
        anchored_pattern = anchored_pattern.replace('A', '\\$?A').replace('B', '\\$?B').replace('C', '\\$?C').replace('D', '\\$?D').replace('E', '\\$?E').replace('F', '\\$?F').replace('G', '\\$?G').replace('H', '\\$?H').replace('I', '\\$?I').replace('J', '\\$?J').replace('K', '\\$?K').replace('L', '\\$?L').replace('M', '\\$?M').replace('N', '\\$?N').replace('O', '\\$?O').replace('P', '\\$?P').replace('Q', '\\$?Q').replace('R', '\\$?R').replace('S', '\\$?S').replace('T', '\\$?T').replace('U', '\\$?U').replace('V', '\\$?V').replace('W', '\\$?W').replace('X', '\\$?X').replace('Y', '\\$?Y').replace('Z', '\\$?Z')
        anchored_pattern = anchored_pattern.replace('1', '\\$?1').replace('2', '\\$?2').replace('3', '\\$?3').replace('4', '\\$?4').replace('5', '\\$?5').replace('6', '\\$?6').replace('7', '\\$?7').replace('8', '\\$?8').replace('9', '\\$?9').replace('0', '\\$?0')
        
        # Check if any anchored version exists
        matches = re.findall(anchored_pattern, formula)
        for match in matches:
            if '$' in match:
                return True
        return False

    def _calculate_anchor_probability(self, sheet, ref: str, current_row: int, current_col: int) -> float:
        """Calculate probability that a reference should be anchored."""
        import re
        match = re.match(r'([A-Z]+)(\d+)', ref)
        if not match:
            return 0.0
        col_str, row_str = match.groups()
        ref_row = int(row_str)
        
        # Higher probability for header references
        if ref_row == 1:
            return 0.9
        
        # Check if it's a constant value
        ref_col = openpyxl.utils.column_index_from_string(col_str)
        ref_cell = sheet.cell(row=ref_row, column=ref_col)
        if ref_cell.value is None:
            return 0.0
        
        constant_count = 0
        total_cells = 0
        for r in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=r, column=ref_col)
            if cell.value is not None:
                total_cells += 1
                if cell.value == ref_cell.value:
                    constant_count += 1
        
        if total_cells == 0:
            return 0.0
        
        constant_ratio = constant_count / total_cells
        if constant_ratio > 0.8:
            return 0.8
        elif constant_ratio > 0.5:
            return 0.6
        else:
            return 0.3

    def _suggest_anchored_formula(self, formula: str, ref: str) -> str:
        """Suggest a formula with proper anchoring for the given reference."""
        # Simple approach: add $ to both row and column
        import re
        anchored_ref = re.sub(r'([A-Z]+)(\d+)', r'$\1$\2', ref)
        return formula.replace(ref, anchored_ref)


class WrongRowColumnAnchoringDetector(ErrorDetector):
    def __init__(self):
        super().__init__(
            name="wrong_row_column_anchoring",
            description="Dollar signs on wrong part of cell reference (partial lock errors)",
            severity=ErrorSeverity.HIGH
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        results = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.data_type == 'f' and cell.value:
                        formula = str(cell.value)
                        # Look for cell references with dollar signs
                        import re
                        cell_refs = re.findall(r'\$?[A-Z]+\$?\d+', formula)
                        for ref in cell_refs:
                            if self._has_wrong_anchoring(sheet, ref, row, col):
                                probability = self._calculate_wrong_anchoring_probability(sheet, ref, row, col)
                                if probability > 0.5:
                                    from openpyxl.utils import get_column_letter
                                    col_letter = get_column_letter(col)
                                    expected_ref = self._suggest_correct_anchoring(sheet, ref, row, col)
                                    expected_formula = self._suggest_correct_formula(formula, ref, expected_ref)
                                    results.append(ErrorDetectionResult(
                                        error_type=self.name,
                                        description=f"Wrong anchoring for reference {ref} in formula at {sheet_name}!{col_letter}{row}; should be {expected_ref}",
                                        probability=probability,
                                        severity=self.severity if probability >= 0.7 else ErrorSeverity.MEDIUM,
                                        location=f"{sheet_name}!{col_letter}{row}",
                                        details={
                                            'formula': formula,
                                            'current_reference': ref,
                                            'expected_reference': expected_ref,
                                            'expected_formula': expected_formula,
                                            'current_row': row,
                                            'current_col': col
                                        },
                                        suggested_fix=f"Adjust anchoring for {ref} to {expected_ref}: {expected_formula}"
                                    ))
        return results

    def _has_wrong_anchoring(self, sheet, ref: str, current_row: int, current_col: int) -> bool:
        """Check if a reference has wrong anchoring based on usage pattern."""
        # Extract row and column from reference
        import re
        match = re.match(r'(\$?)([A-Z]+)(\$?)(\d+)', ref)
        if not match:
            return False
        col_dollar, col_str, row_dollar, row_str = match.groups()
        ref_row = int(row_str)
        ref_col = openpyxl.utils.column_index_from_string(col_str)
        
        # Determine expected anchoring based on usage pattern
        expected_anchoring = self._determine_expected_anchoring(sheet, ref_col, ref_row, current_col, current_row)
        
        # Compare actual vs expected anchoring
        actual_anchoring = self._get_anchoring_type(col_dollar, row_dollar)
        
        # Only flag if expected is not relative and actual doesn't match expected
        if expected_anchoring != "relative" and actual_anchoring != expected_anchoring:
            return True
        
        return False

    def _determine_expected_anchoring(self, sheet, ref_col: int, ref_row: int, current_col: int, current_row: int) -> str:
        """Determine the expected anchoring type for a reference."""
        # Check if this is a constant value (should be fully locked)
        ref_cell = sheet.cell(row=ref_row, column=ref_col)
        if ref_cell.value is None:
            return "relative"
        
        # Check if it's a header (row 1) - should be row-locked
        if ref_row == 1:
            return "row_locked"
        
        # Check if it's a constant value
        constant_count = 0
        total_cells = 0
        for r in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=r, column=ref_col)
            if cell.value is not None:
                total_cells += 1
                if cell.value == ref_cell.value:
                    constant_count += 1
        
        if constant_count > 2 and total_cells > 0 and (constant_count / total_cells) > 0.5:
            return "fully_locked"
        
        # For varying values, assume relative (don't flag as wrong)
        return "relative"

    def _get_anchoring_type(self, col_dollar: str, row_dollar: str) -> str:
        """Get the anchoring type from dollar signs."""
        if col_dollar and row_dollar:
            return "fully_locked"
        elif col_dollar:
            return "column_locked"
        elif row_dollar:
            return "row_locked"
        else:
            return "relative"

    def _calculate_wrong_anchoring_probability(self, sheet, ref: str, current_row: int, current_col: int) -> float:
        """Calculate probability that anchoring is wrong."""
        import re
        match = re.match(r'(\$?)([A-Z]+)(\$?)(\d+)', ref)
        if not match:
            return 0.0
        col_dollar, col_str, row_dollar, row_str = match.groups()
        ref_row = int(row_str)
        
        # Higher probability for header references
        if ref_row == 1:
            return 0.9
        
        # Check if it's a constant value
        ref_col = openpyxl.utils.column_index_from_string(col_str)
        ref_cell = sheet.cell(row=ref_row, column=ref_col)
        if ref_cell.value is None:
            return 0.0
        
        constant_count = 0
        total_cells = 0
        for r in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=r, column=ref_col)
            if cell.value is not None:
                total_cells += 1
                if cell.value == ref_cell.value:
                    constant_count += 1
        
        if total_cells == 0:
            return 0.0
        
        constant_ratio = constant_count / total_cells
        if constant_ratio > 0.8:
            return 0.8
        elif constant_ratio > 0.5:
            return 0.6
        else:
            return 0.4

    def _suggest_correct_anchoring(self, sheet, ref: str, current_row: int, current_col: int) -> str:
        """Suggest the correct anchoring for a reference."""
        import re
        match = re.match(r'(\$?)([A-Z]+)(\$?)(\d+)', ref)
        if not match:
            return ref
        col_dollar, col_str, row_dollar, row_str = match.groups()
        ref_row = int(row_str)
        ref_col = openpyxl.utils.column_index_from_string(col_str)
        
        expected_anchoring = self._determine_expected_anchoring(sheet, ref_col, ref_row, current_col, current_row)
        
        if expected_anchoring == "fully_locked":
            return f"${col_str}${row_str}"
        elif expected_anchoring == "column_locked":
            return f"${col_str}{row_str}"
        elif expected_anchoring == "row_locked":
            return f"{col_str}${row_str}"
        else:
            return f"{col_str}{row_str}"

    def _suggest_correct_formula(self, formula: str, current_ref: str, expected_ref: str) -> str:
        """Suggest the correct formula with proper anchoring."""
        return formula.replace(current_ref, expected_ref)


class OverAnchoredReferencesDetector(ErrorDetector):
    def __init__(self):
        super().__init__(
            name="over_anchored_references",
            description="Dollar signs on references that should be relative (unnecessary absolute)",
            severity=ErrorSeverity.MEDIUM
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        results = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            # Find copied formula patterns
            copied_patterns = self._find_copied_formula_patterns(sheet)
            
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.data_type == 'f' and cell.value:
                        formula = str(cell.value)
                        # Look for over-anchored references
                        over_anchored_refs = self._find_over_anchored_references(sheet, formula, row, col, copied_patterns)
                        for ref in over_anchored_refs:
                            probability = self._calculate_over_anchoring_probability(sheet, ref, row, col, copied_patterns)
                            if probability > 0.5:
                                from openpyxl.utils import get_column_letter
                                col_letter = get_column_letter(col)
                                expected_ref = self._suggest_relative_reference(ref)
                                expected_formula = self._suggest_relative_formula(formula, ref, expected_ref)
                                results.append(ErrorDetectionResult(
                                    error_type=self.name,
                                    description=f"Over-anchored reference {ref} in formula at {sheet_name}!{col_letter}{row}; should be {expected_ref}",
                                    probability=probability,
                                    severity=self.severity if probability >= 0.7 else ErrorSeverity.LOW,
                                    location=f"{sheet_name}!{col_letter}{row}",
                                    details={
                                        'formula': formula,
                                        'over_anchored_reference': ref,
                                        'expected_reference': expected_ref,
                                        'expected_formula': expected_formula,
                                        'current_row': row,
                                        'current_col': col
                                    },
                                    suggested_fix=f"Remove unnecessary anchoring for {ref}: {expected_formula}"
                                ))
        return results

    def _find_copied_formula_patterns(self, sheet) -> dict:
        """Find patterns of copied formulas in the sheet."""
        patterns = {}
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        for col in range(1, max_col + 1):
            formula_patterns = {}
            for row in range(1, max_row + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.data_type == 'f' and cell.value:
                    formula = str(cell.value)
                    # Normalize formula by removing cell references
                    normalized = self._normalize_formula(formula)
                    if normalized not in formula_patterns:
                        formula_patterns[normalized] = []
                    formula_patterns[normalized].append(row)
            
            # Keep only patterns with multiple occurrences (copied formulas)
            for pattern, rows in formula_patterns.items():
                if len(rows) > 1:
                    patterns[pattern] = rows
        
        return patterns

    def _normalize_formula(self, formula: str) -> str:
        """Normalize formula by replacing cell references with placeholders."""
        import re
        # Replace cell references with placeholders
        normalized = re.sub(r'\$?[A-Z]+\$?\d+', 'CELL', formula)
        return normalized

    def _find_over_anchored_references(self, sheet, formula: str, current_row: int, current_col: int, copied_patterns: dict) -> List[str]:
        """Find over-anchored references in a formula."""
        import re
        cell_refs = re.findall(r'\$?[A-Z]+\$?\d+', formula)
        over_anchored = []
        
        for ref in cell_refs:
            if self._is_over_anchored(sheet, ref, current_row, current_col, copied_patterns):
                over_anchored.append(ref)
        
        return over_anchored

    def _is_over_anchored(self, sheet, ref: str, current_row: int, current_col: int, copied_patterns: dict) -> bool:
        """Check if a reference is over-anchored."""
        import re
        match = re.match(r'(\$?)([A-Z]+)(\$?)(\d+)', ref)
        if not match:
            return False
        
        col_dollar, col_str, row_dollar, row_str = match.groups()
        ref_row = int(row_str)
        ref_col = openpyxl.utils.column_index_from_string(col_str)
        
        # Check if this is part of a copied pattern
        is_copied = self._is_in_copied_pattern(sheet, current_row, current_col, copied_patterns)
        
        # Check if the referenced cell contains varying values (not a constant)
        is_varying = self._is_varying_value(sheet, ref_col, ref_row)
        
        # Check if it's a header (row 1) - headers should be anchored
        is_header = ref_row == 1
        
        # Over-anchored if: fully locked AND (copied pattern OR varying value) AND not header
        if col_dollar and row_dollar and is_copied and not is_header:
            return True
        
        # Over-anchored if: partially locked AND varying value AND not header
        if (col_dollar or row_dollar) and is_varying and not is_header:
            return True
        
        return False

    def _is_in_copied_pattern(self, sheet, row: int, col: int, copied_patterns: dict) -> bool:
        """Check if a cell is part of a copied formula pattern."""
        cell = sheet.cell(row=row, column=col)
        if cell.data_type != 'f' or not cell.value:
            return False
        
        formula = str(cell.value)
        normalized = self._normalize_formula(formula)
        
        return normalized in copied_patterns and row in copied_patterns[normalized]

    def _is_varying_value(self, sheet, col: int, row: int) -> bool:
        """Check if a cell contains a varying value (not constant)."""
        ref_cell = sheet.cell(row=row, column=col)
        if ref_cell.value is None:
            return True  # Empty cells are considered varying
        
        # Check if this value appears in multiple cells in the same column
        constant_count = 0
        total_cells = 0
        for r in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=r, column=col)
            if cell.value is not None:
                total_cells += 1
                if cell.value == ref_cell.value:
                    constant_count += 1
        
        # If the same value appears in more than 50% of non-empty cells, it's a constant
        if total_cells > 0 and (constant_count / total_cells) > 0.5:
            return False
        
        return True

    def _calculate_over_anchoring_probability(self, sheet, ref: str, current_row: int, current_col: int, copied_patterns: dict) -> float:
        """Calculate probability that a reference is over-anchored."""
        import re
        match = re.match(r'(\$?)([A-Z]+)(\$?)(\d+)', ref)
        if not match:
            return 0.0
        
        col_dollar, col_str, row_dollar, row_str = match.groups()
        ref_row = int(row_str)
        ref_col = openpyxl.utils.column_index_from_string(col_str)
        
        # Base probability
        base_prob = 0.5
        
        # Increase probability for copied patterns
        if self._is_in_copied_pattern(sheet, current_row, current_col, copied_patterns):
            base_prob += 0.3
        
        # Increase probability for varying values
        if self._is_varying_value(sheet, ref_col, ref_row):
            base_prob += 0.2
        
        # Decrease probability for headers
        if ref_row == 1:
            base_prob -= 0.3
        
        # Increase probability for fully locked references
        if col_dollar and row_dollar:
            base_prob += 0.2
        
        return min(0.9, max(0.1, base_prob))

    def _suggest_relative_reference(self, ref: str) -> str:
        """Suggest a relative reference by removing dollar signs."""
        import re
        return re.sub(r'\$', '', ref)

    def _suggest_relative_formula(self, formula: str, current_ref: str, expected_ref: str) -> str:
        """Suggest a formula with relative references."""
        return formula.replace(current_ref, expected_ref)


class InconsistentAnchoringInRangesDetector(ErrorDetector):
    def __init__(self):
        super().__init__(
            name="inconsistent_anchoring_in_ranges",
            description="Mixed anchoring within the same range reference (inconsistent anchoring patterns)",
            severity=ErrorSeverity.MEDIUM
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        results = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.data_type == 'f' and cell.value:
                        formula = str(cell.value)
                        # Look for inconsistent anchoring in ranges
                        inconsistent_ranges = self._find_inconsistent_ranges(formula)
                        for range_ref in inconsistent_ranges:
                            probability = self._calculate_inconsistency_probability(sheet, range_ref, row, col)
                            if probability > 0.5:
                                from openpyxl.utils import get_column_letter
                                col_letter = get_column_letter(col)
                                expected_range = self._suggest_consistent_range(range_ref)
                                expected_formula = self._suggest_consistent_formula(formula, range_ref, expected_range)
                                results.append(ErrorDetectionResult(
                                    error_type=self.name,
                                    description=f"Inconsistent anchoring in range {range_ref} in formula at {sheet_name}!{col_letter}{row}; should be {expected_range}",
                                    probability=probability,
                                    severity=self.severity if probability >= 0.7 else ErrorSeverity.LOW,
                                    location=f"{sheet_name}!{col_letter}{row}",
                                    details={
                                        'formula': formula,
                                        'inconsistent_range': range_ref,
                                        'expected_range': expected_range,
                                        'expected_formula': expected_formula,
                                        'current_row': row,
                                        'current_col': col
                                    },
                                    suggested_fix=f"Make anchoring consistent in range {range_ref}: {expected_formula}"
                                ))
        return results

    def _find_inconsistent_ranges(self, formula: str) -> List[str]:
        """Find ranges with inconsistent anchoring in a formula."""
        import re
        # Find range patterns like A1:B10, $A$1:A10, A1:$B$10, etc.
        range_patterns = re.findall(r'\$?[A-Z]+\$?\d+:\$?[A-Z]+\$?\d+', formula)
        inconsistent_ranges = []
        
        for range_ref in range_patterns:
            if self._has_inconsistent_anchoring(range_ref):
                inconsistent_ranges.append(range_ref)
        
        return inconsistent_ranges

    def _has_inconsistent_anchoring(self, range_ref: str) -> bool:
        """Check if a range has inconsistent anchoring."""
        import re
        # Split range into start and end parts
        parts = range_ref.split(':')
        if len(parts) != 2:
            return False
        
        start_ref, end_ref = parts
        
        # Parse anchoring for start reference
        start_match = re.match(r'(\$?)([A-Z]+)(\$?)(\d+)', start_ref)
        if not start_match:
            return False
        
        start_col_dollar, start_col, start_row_dollar, start_row = start_match.groups()
        start_anchoring = self._get_anchoring_type(start_col_dollar, start_row_dollar)
        
        # Parse anchoring for end reference
        end_match = re.match(r'(\$?)([A-Z]+)(\$?)(\d+)', end_ref)
        if not end_match:
            return False
        
        end_col_dollar, end_col, end_row_dollar, end_row = end_match.groups()
        end_anchoring = self._get_anchoring_type(end_col_dollar, end_row_dollar)
        
        # Check for inconsistency
        return start_anchoring != end_anchoring

    def _get_anchoring_type(self, col_dollar: str, row_dollar: str) -> str:
        """Get the anchoring type from dollar signs."""
        if col_dollar and row_dollar:
            return "fully_locked"
        elif col_dollar:
            return "column_locked"
        elif row_dollar:
            return "row_locked"
        else:
            return "relative"

    def _calculate_inconsistency_probability(self, sheet, range_ref: str, current_row: int, current_col: int) -> float:
        """Calculate probability that inconsistent anchoring will cause problems."""
        # Base probability
        base_prob = 0.6
        
        # Check if this is in a calculation function
        cell = sheet.cell(row=current_row, column=current_col)
        if cell.data_type == 'f' and cell.value:
            formula = str(cell.value).upper()
            calc_functions = ["SUM", "AVERAGE", "COUNT", "COUNTA", "MAX", "MIN", "VLOOKUP", "HLOOKUP", "INDEX"]
            if any(func in formula for func in calc_functions):
                base_prob += 0.2
        
        # Check if this is likely to be copied (part of a pattern)
        if self._is_likely_copied(sheet, current_row, current_col):
            base_prob += 0.2
        
        # Check severity of inconsistency
        severity = self._get_inconsistency_severity(range_ref)
        if severity == "high":
            base_prob += 0.1
        elif severity == "medium":
            base_prob += 0.05
        
        return min(0.9, base_prob)

    def _is_likely_copied(self, sheet, row: int, col: int) -> bool:
        """Check if a cell is likely part of a copied pattern."""
        # Simple heuristic: check if adjacent cells have similar formulas
        adjacent_formulas = []
        for r in range(max(1, row-1), min(sheet.max_row + 1, row+2)):
            for c in range(max(1, col-1), min(sheet.max_column + 1, col+2)):
                if r != row or c != col:
                    cell = sheet.cell(row=r, column=c)
                    if cell.data_type == 'f' and cell.value:
                        adjacent_formulas.append(str(cell.value))
        
        # If there are similar formulas nearby, it's likely copied
        if len(adjacent_formulas) >= 2:
            return True
        
        return False

    def _get_inconsistency_severity(self, range_ref: str) -> str:
        """Get the severity of anchoring inconsistency."""
        import re
        parts = range_ref.split(':')
        if len(parts) != 2:
            return "low"
        
        start_ref, end_ref = parts
        
        # Parse anchoring
        start_match = re.match(r'(\$?)([A-Z]+)(\$?)(\d+)', start_ref)
        end_match = re.match(r'(\$?)([A-Z]+)(\$?)(\d+)', end_ref)
        
        if not start_match or not end_match:
            return "low"
        
        start_col_dollar, _, start_row_dollar, _ = start_match.groups()
        end_col_dollar, _, end_row_dollar, _ = end_match.groups()
        
        start_anchoring = self._get_anchoring_type(start_col_dollar, start_row_dollar)
        end_anchoring = self._get_anchoring_type(end_col_dollar, end_row_dollar)
        
        # High severity: fully locked vs relative
        if (start_anchoring == "fully_locked" and end_anchoring == "relative") or \
           (start_anchoring == "relative" and end_anchoring == "fully_locked"):
            return "high"
        
        # Medium severity: partial vs relative or partial vs fully locked
        if start_anchoring != end_anchoring:
            return "medium"
        
        return "low"

    def _suggest_consistent_range(self, range_ref: str) -> str:
        """Suggest a consistent anchoring for a range."""
        import re
        parts = range_ref.split(':')
        if len(parts) != 2:
            return range_ref
        
        start_ref, end_ref = parts
        
        # Determine the most appropriate anchoring based on the range
        # For now, suggest relative anchoring (most common case)
        start_relative = re.sub(r'\$', '', start_ref)
        end_relative = re.sub(r'\$', '', end_ref)
        
        return f"{start_relative}:{end_relative}"

    def _suggest_consistent_formula(self, formula: str, current_range: str, expected_range: str) -> str:
        """Suggest a formula with consistent anchoring."""
        return formula.replace(current_range, expected_range)


class LookupFunctionAnchoringDetector(ErrorDetector):
    def __init__(self):
        super().__init__(
            name="lookup_function_anchoring_errors",
            description="Wrong anchoring in VLOOKUP, HLOOKUP, INDEX/MATCH functions",
            severity=ErrorSeverity.HIGH
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        results = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.data_type == 'f' and cell.value:
                        formula = str(cell.value)
                        # Look for anchoring errors in lookup functions
                        lookup_errors = self._find_lookup_anchoring_errors(sheet, formula, row, col)
                        for error in lookup_errors:
                            probability = self._calculate_lookup_error_probability(sheet, error, row, col)
                            if probability > 0.5:
                                from openpyxl.utils import get_column_letter
                                col_letter = get_column_letter(col)
                                results.append(ErrorDetectionResult(
                                    error_type=self.name,
                                    description=error['description'],
                                    probability=probability,
                                    severity=self.severity if probability >= 0.7 else ErrorSeverity.MEDIUM,
                                    location=f"{sheet_name}!{col_letter}{row}",
                                    details={
                                        'formula': formula,
                                        'function_type': error['function_type'],
                                        'parameter': error['parameter'],
                                        'current_anchoring': error['current_anchoring'],
                                        'expected_anchoring': error['expected_anchoring'],
                                        'copy_direction': error['copy_direction'],
                                        'current_row': row,
                                        'current_col': col
                                    },
                                    suggested_fix=error['suggested_fix']
                                ))
        return results

    def _find_lookup_anchoring_errors(self, sheet, formula: str, current_row: int, current_col: int) -> List[dict]:
        """Find anchoring errors in lookup functions."""
        errors = []
        
        # Check for VLOOKUP errors
        vlookup_errors = self._check_vlookup_anchoring(sheet, formula, current_row, current_col)
        errors.extend(vlookup_errors)
        
        # Check for HLOOKUP errors
        hlookup_errors = self._check_hlookup_anchoring(sheet, formula, current_row, current_col)
        errors.extend(hlookup_errors)
        
        # Check for INDEX/MATCH errors
        index_match_errors = self._check_index_match_anchoring(sheet, formula, current_row, current_col)
        errors.extend(index_match_errors)
        
        return errors

    def _check_vlookup_anchoring(self, sheet, formula: str, current_row: int, current_col: int) -> List[dict]:
        """Check anchoring in VLOOKUP functions."""
        import re
        errors = []
        
        # Find VLOOKUP functions
        vlookup_pattern = r'VLOOKUP\s*\(\s*([^,]+)\s*,\s*([^,]+)\s*,\s*([^,]+)\s*(?:,\s*([^)]+))?\s*\)'
        matches = re.finditer(vlookup_pattern, formula, re.IGNORECASE)
        
        for match in matches:
            lookup_value = match.group(1).strip()
            table_array = match.group(2).strip()
            col_index = match.group(3).strip()
            
            # Determine copy direction
            copy_direction = self._determine_copy_direction(sheet, current_row, current_col)
            
            # Check lookup value anchoring (flag if not column-locked, regardless of copy direction)
            if not self._is_column_locked(lookup_value):
                errors.append({
                    'function_type': 'VLOOKUP',
                    'parameter': 'lookup_value',
                    'current_anchoring': self._get_anchoring_type_from_ref(lookup_value),
                    'expected_anchoring': 'column_locked',
                    'copy_direction': copy_direction,
                    'description': f"VLOOKUP lookup value {lookup_value} should be column-locked when copying across",
                    'suggested_fix': f"Change {lookup_value} to {self._make_column_locked(lookup_value)}"
                })
            
            # Check table array anchoring (always flag if not fully locked)
            if not self._is_fully_locked(table_array):
                errors.append({
                    'function_type': 'VLOOKUP',
                    'parameter': 'table_array',
                    'current_anchoring': self._get_anchoring_type_from_ref(table_array),
                    'expected_anchoring': 'fully_locked',
                    'copy_direction': copy_direction,
                    'description': f"VLOOKUP table array {table_array} should be fully locked",
                    'suggested_fix': f"Change {table_array} to {self._make_fully_locked(table_array)}"
                })
        
        return errors

    def _check_hlookup_anchoring(self, sheet, formula: str, current_row: int, current_col: int) -> List[dict]:
        """Check anchoring in HLOOKUP functions."""
        import re
        errors = []
        
        # Find HLOOKUP functions
        hlookup_pattern = r'HLOOKUP\s*\(\s*([^,]+)\s*,\s*([^,]+)\s*,\s*([^,]+)\s*(?:,\s*([^)]+))?\s*\)'
        matches = re.finditer(hlookup_pattern, formula, re.IGNORECASE)
        
        for match in matches:
            lookup_value = match.group(1).strip()
            table_array = match.group(2).strip()
            row_index = match.group(3).strip()
            
            # Determine copy direction
            copy_direction = self._determine_copy_direction(sheet, current_row, current_col)
            
            # Check lookup value anchoring (flag if not row-locked, regardless of copy direction)
            if not self._is_row_locked(lookup_value):
                errors.append({
                    'function_type': 'HLOOKUP',
                    'parameter': 'lookup_value',
                    'current_anchoring': self._get_anchoring_type_from_ref(lookup_value),
                    'expected_anchoring': 'row_locked',
                    'copy_direction': copy_direction,
                    'description': f"HLOOKUP lookup value {lookup_value} should be row-locked when copying down",
                    'suggested_fix': f"Change {lookup_value} to {self._make_row_locked(lookup_value)}"
                })
            
            # Check table array anchoring (always flag if not fully locked)
            if not self._is_fully_locked(table_array):
                errors.append({
                    'function_type': 'HLOOKUP',
                    'parameter': 'table_array',
                    'current_anchoring': self._get_anchoring_type_from_ref(table_array),
                    'expected_anchoring': 'fully_locked',
                    'copy_direction': copy_direction,
                    'description': f"HLOOKUP table array {table_array} should be fully locked",
                    'suggested_fix': f"Change {table_array} to {self._make_fully_locked(table_array)}"
                })
        
        return errors

    def _check_index_match_anchoring(self, sheet, formula: str, current_row: int, current_col: int) -> List[dict]:
        """Check anchoring in INDEX/MATCH functions."""
        import re
        errors = []
        
        # Find INDEX functions with MATCH
        index_match_pattern = r'INDEX\s*\(\s*([^,]+)\s*,\s*MATCH\s*\(\s*([^,]+)\s*,\s*([^,]+)\s*(?:,\s*([^)]+))?\s*\)\s*\)'
        matches = re.finditer(index_match_pattern, formula, re.IGNORECASE)
        
        for match in matches:
            array = match.group(1).strip()
            lookup_value = match.group(2).strip()
            lookup_array = match.group(3).strip()
            
            # Determine copy direction
            copy_direction = self._determine_copy_direction(sheet, current_row, current_col)
            
            # Check lookup value anchoring (flag if not column-locked, regardless of copy direction)
            if not self._is_column_locked(lookup_value):
                errors.append({
                    'function_type': 'INDEX/MATCH',
                    'parameter': 'lookup_value',
                    'current_anchoring': self._get_anchoring_type_from_ref(lookup_value),
                    'expected_anchoring': 'column_locked',
                    'copy_direction': copy_direction,
                    'description': f"INDEX/MATCH lookup value {lookup_value} should be column-locked when copying across",
                    'suggested_fix': f"Change {lookup_value} to {self._make_column_locked(lookup_value)}"
                })
            
            # Check array anchoring (always flag if not fully locked)
            if not self._is_fully_locked(array):
                errors.append({
                    'function_type': 'INDEX/MATCH',
                    'parameter': 'array',
                    'current_anchoring': self._get_anchoring_type_from_ref(array),
                    'expected_anchoring': 'fully_locked',
                    'copy_direction': copy_direction,
                    'description': f"INDEX/MATCH array {array} should be fully locked",
                    'suggested_fix': f"Change {array} to {self._make_fully_locked(array)}"
                })
            
            # Check lookup array anchoring (always flag if not fully locked)
            if not self._is_fully_locked(lookup_array):
                errors.append({
                    'function_type': 'INDEX/MATCH',
                    'parameter': 'lookup_array',
                    'current_anchoring': self._get_anchoring_type_from_ref(lookup_array),
                    'expected_anchoring': 'fully_locked',
                    'copy_direction': copy_direction,
                    'description': f"INDEX/MATCH lookup array {lookup_array} should be fully locked",
                    'suggested_fix': f"Change {lookup_array} to {self._make_fully_locked(lookup_array)}"
                })
        
        return errors

    def _determine_copy_direction(self, sheet, row: int, col: int) -> str:
        """Determine the likely copy direction for a cell."""
        # Simple heuristic: check if adjacent cells have similar formulas
        horizontal_similar = 0
        vertical_similar = 0
        
        # Check horizontal (across)
        for c in range(max(1, col-1), min(sheet.max_column + 1, col+2)):
            if c != col:
                cell = sheet.cell(row=row, column=c)
                if cell.data_type == 'f' and cell.value:
                    horizontal_similar += 1
        
        # Check vertical (down)
        for r in range(max(1, row-1), min(sheet.max_row + 1, row+2)):
            if r != row:
                cell = sheet.cell(row=r, column=col)
                if cell.data_type == 'f' and cell.value:
                    vertical_similar += 1
        
        if horizontal_similar > vertical_similar:
            return "across"
        elif vertical_similar > horizontal_similar:
            return "down"
        else:
            return "unknown"

    def _is_column_locked(self, ref: str) -> bool:
        """Check if a reference is column-locked."""
        import re
        match = re.match(r'(\$?)([A-Z]+)(\$?)(\d+)', ref)
        if not match:
            return False
        col_dollar, _, row_dollar, _ = match.groups()
        return bool(col_dollar) and not bool(row_dollar)

    def _is_row_locked(self, ref: str) -> bool:
        """Check if a reference is row-locked."""
        import re
        match = re.match(r'(\$?)([A-Z]+)(\$?)(\d+)', ref)
        if not match:
            return False
        col_dollar, _, row_dollar, _ = match.groups()
        return not bool(col_dollar) and bool(row_dollar)

    def _is_fully_locked(self, ref: str) -> bool:
        """Check if a reference is fully locked."""
        import re
        match = re.match(r'(\$?)([A-Z]+)(\$?)(\d+)', ref)
        if not match:
            return False
        col_dollar, _, row_dollar, _ = match.groups()
        return bool(col_dollar) and bool(row_dollar)

    def _get_anchoring_type_from_ref(self, ref: str) -> str:
        """Get anchoring type from a reference."""
        if self._is_fully_locked(ref):
            return "fully_locked"
        elif self._is_column_locked(ref):
            return "column_locked"
        elif self._is_row_locked(ref):
            return "row_locked"
        else:
            return "relative"

    def _make_column_locked(self, ref: str) -> str:
        """Make a reference column-locked."""
        import re
        return re.sub(r'([A-Z]+)(\d+)', r'$\1\2', ref)

    def _make_row_locked(self, ref: str) -> str:
        """Make a reference row-locked."""
        import re
        return re.sub(r'([A-Z]+)(\d+)', r'\1$\2', ref)

    def _make_fully_locked(self, ref: str) -> str:
        """Make a reference fully locked."""
        import re
        return re.sub(r'([A-Z]+)(\d+)', r'$\1$\2', ref)

    def _calculate_lookup_error_probability(self, sheet, error: dict, current_row: int, current_col: int) -> float:
        """Calculate probability that a lookup anchoring error will cause problems."""
        base_prob = 0.7  # High base probability for lookup functions
        
        # Increase probability for critical parameters
        if error['parameter'] in ['table_array', 'array', 'lookup_array']:
            base_prob += 0.2  # Table arrays are critical
        
        # Increase probability for known copy direction
        if error['copy_direction'] != "unknown":
            base_prob += 0.1
        
        # Increase probability for high severity functions
        if error['function_type'] in ['VLOOKUP', 'INDEX/MATCH']:
            base_prob += 0.1
        
        return min(0.9, base_prob)


class ArrayFormulaAnchoringDetector(ErrorDetector):
    def __init__(self):
        super().__init__(
            name="array_formula_anchoring_errors",
            description="Incorrect anchoring in array formulas",
            severity=ErrorSeverity.MEDIUM
        )

    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        results = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.data_type == 'f' and cell.value:
                        formula = str(cell.value)
                        # Look for anchoring errors in array formulas
                        array_errors = self._find_array_anchoring_errors(sheet, formula, row, col)
                        for error in array_errors:
                            probability = self._calculate_array_error_probability(sheet, error, row, col)
                            if probability > 0.5:
                                from openpyxl.utils import get_column_letter
                                col_letter = get_column_letter(col)
                                results.append(ErrorDetectionResult(
                                    error_type=self.name,
                                    description=error['description'],
                                    probability=probability,
                                    severity=self.severity if probability >= 0.7 else ErrorSeverity.LOW,
                                    location=f"{sheet_name}!{col_letter}{row}",
                                    details={
                                        'formula': formula,
                                        'array_function': error['array_function'],
                                        'range_reference': error['range_reference'],
                                        'current_anchoring': error['current_anchoring'],
                                        'expected_anchoring': error['expected_anchoring'],
                                        'current_row': row,
                                        'current_col': col
                                    },
                                    suggested_fix=error['suggested_fix']
                                ))
        return results

    def _find_array_anchoring_errors(self, sheet, formula: str, current_row: int, current_col: int) -> List[dict]:
        """Find anchoring errors in array formulas."""
        errors = []
        
        # Check for SUM(IF()) array formulas
        sum_if_errors = self._check_sum_if_anchoring(sheet, formula, current_row, current_col)
        errors.extend(sum_if_errors)
        
        # Check for modern array functions
        modern_array_errors = self._check_modern_array_functions(sheet, formula, current_row, current_col)
        errors.extend(modern_array_errors)
        
        return errors

    def _check_sum_if_anchoring(self, sheet, formula: str, current_row: int, current_col: int) -> List[dict]:
        """Check anchoring in SUM(IF()) array formulas."""
        import re
        errors = []
        
        # Find SUM(IF()) patterns
        sum_if_pattern = r'SUM\s*\(\s*IF\s*\(\s*([^,]+)\s*,\s*([^,]+)\s*(?:,\s*([^)]+))?\s*\)\s*\)'
        matches = re.finditer(sum_if_pattern, formula, re.IGNORECASE)
        
        for match in matches:
            condition_range = match.group(1).strip()
            true_range = match.group(2).strip()
            false_value = match.group(3).strip() if match.group(3) else "0"
            
            # Check if ranges are over-anchored (should be relative for flexibility)
            if self._is_over_anchored_for_array(condition_range):
                errors.append({
                    'array_function': 'SUM(IF)',
                    'range_reference': condition_range,
                    'current_anchoring': self._get_anchoring_type_from_ref(condition_range),
                    'expected_anchoring': 'relative',
                    'description': f"SUM(IF) condition range {condition_range} should be relative for flexibility",
                    'suggested_fix': f"Change {condition_range} to {self._make_relative(condition_range)}"
                })
            
            if self._is_over_anchored_for_array(true_range):
                errors.append({
                    'array_function': 'SUM(IF)',
                    'range_reference': true_range,
                    'current_anchoring': self._get_anchoring_type_from_ref(true_range),
                    'expected_anchoring': 'relative',
                    'description': f"SUM(IF) true range {true_range} should be relative for flexibility",
                    'suggested_fix': f"Change {true_range} to {self._make_relative(true_range)}"
                })
        
        return errors

    def _check_modern_array_functions(self, sheet, formula: str, current_row: int, current_col: int) -> List[dict]:
        """Check anchoring in modern array functions."""
        import re
        errors = []
        
        # Array functions that should typically have relative ranges
        array_functions = [
            ('UNIQUE', r'UNIQUE\s*\(\s*([^)]+)\s*\)'),
            ('FILTER', r'FILTER\s*\(\s*([^,]+)\s*,\s*([^)]+)\s*\)'),
            ('SORT', r'SORT\s*\(\s*([^)]+)\s*\)'),
            ('SORTBY', r'SORTBY\s*\(\s*([^,]+)\s*,\s*([^)]+)\s*\)'),
            ('SEQUENCE', r'SEQUENCE\s*\(\s*([^)]+)\s*\)'),
            ('RANDARRAY', r'RANDARRAY\s*\(\s*([^)]+)\s*\)')
        ]
        
        for func_name, pattern in array_functions:
            matches = re.finditer(pattern, formula, re.IGNORECASE)
            for match in matches:
                # Check each range parameter
                for i in range(1, len(match.groups()) + 1):
                    range_ref = match.group(i).strip()
                    if self._is_over_anchored_for_array(range_ref):
                        errors.append({
                            'array_function': func_name,
                            'range_reference': range_ref,
                            'current_anchoring': self._get_anchoring_type_from_ref(range_ref),
                            'expected_anchoring': 'relative',
                            'description': f"{func_name} range {range_ref} should be relative for flexibility",
                            'suggested_fix': f"Change {range_ref} to {self._make_relative(range_ref)}"
                        })
        
        return errors

    def _is_over_anchored_for_array(self, range_ref: str) -> bool:
        """Check if a range is over-anchored for array formula flexibility."""
        # Check if it's a range reference
        if ':' not in range_ref:
            return False
        
        # Split range into start and end parts
        parts = range_ref.split(':')
        if len(parts) != 2:
            return False
        
        start_ref, end_ref = parts
        
        # Check if both parts are fully locked
        if self._is_fully_locked(start_ref) and self._is_fully_locked(end_ref):
            return True
        
        # Check if it's a very large range (likely should be relative)
        if self._is_large_range(start_ref, end_ref):
            return True
        
        return False

    def _is_large_range(self, start_ref: str, end_ref: str) -> bool:
        """Check if a range is large enough to warrant relative anchoring."""
        import re
        
        # Extract row numbers
        start_match = re.match(r'[A-Z]+(\d+)', start_ref)
        end_match = re.match(r'[A-Z]+(\d+)', end_ref)
        
        if not start_match or not end_match:
            return False
        
        start_row = int(start_match.group(1))
        end_row = int(end_match.group(1))
        
        # Consider ranges with more than 50 rows as large
        return (end_row - start_row) > 50

    def _is_fully_locked(self, ref: str) -> bool:
        """Check if a reference is fully locked."""
        import re
        match = re.match(r'(\$?)([A-Z]+)(\$?)(\d+)', ref)
        if not match:
            return False
        col_dollar, _, row_dollar, _ = match.groups()
        return bool(col_dollar) and bool(row_dollar)

    def _get_anchoring_type_from_ref(self, ref: str) -> str:
        """Get anchoring type from a reference."""
        if self._is_fully_locked(ref):
            return "fully_locked"
        elif self._is_column_locked(ref):
            return "column_locked"
        elif self._is_row_locked(ref):
            return "row_locked"
        else:
            return "relative"

    def _is_column_locked(self, ref: str) -> bool:
        """Check if a reference is column-locked."""
        import re
        match = re.match(r'(\$?)([A-Z]+)(\$?)(\d+)', ref)
        if not match:
            return False
        col_dollar, _, row_dollar, _ = match.groups()
        return bool(col_dollar) and not bool(row_dollar)

    def _is_row_locked(self, ref: str) -> bool:
        """Check if a reference is row-locked."""
        import re
        match = re.match(r'(\$?)([A-Z]+)(\$?)(\d+)', ref)
        if not match:
            return False
        col_dollar, _, row_dollar, _ = match.groups()
        return not bool(col_dollar) and bool(row_dollar)

    def _make_relative(self, ref: str) -> str:
        """Make a reference relative by removing dollar signs."""
        import re
        return re.sub(r'\$', '', ref)

    def _calculate_array_error_probability(self, sheet, error: dict, current_row: int, current_col: int) -> float:
        """Calculate probability that an array anchoring error will cause problems."""
        base_prob = 0.6  # Medium base probability for array formulas
        
        # Increase probability for large ranges
        if ':' in error['range_reference']:
            parts = error['range_reference'].split(':')
            if len(parts) == 2:
                start_ref, end_ref = parts
                if self._is_large_range(start_ref, end_ref):
                    base_prob += 0.2
        
        # Increase probability for critical array functions
        if error['array_function'] in ['SUM(IF)', 'FILTER', 'UNIQUE']:
            base_prob += 0.1
        
        # Increase probability for fully locked ranges
        if error['current_anchoring'] == 'fully_locked':
            base_prob += 0.1
        
        return min(0.9, base_prob)


class CrossSheetAnchoringDetector(ErrorDetector):
    """
    Detects wrong anchoring when referencing other sheets.
    
    Cross-sheet references need special consideration for anchoring:
    - When copying across: References should be column-locked (Sheet1!$A1)
    - When copying down: References should be row-locked (Sheet1!A$1)  
    - When copying both directions: References should be relative (Sheet1!A1)
    - Fixed references: Should be fully locked (Sheet1!$A$1)
    """
    
    def __init__(self):
        super().__init__(
            name="cross_sheet_anchoring_errors",
            description="Wrong anchoring when referencing other sheets",
            severity=ErrorSeverity.MEDIUM
        )
    
    def detect(self, workbook: openpyxl.Workbook, **kwargs) -> List[ErrorDetectionResult]:
        """Detect cross-sheet anchoring errors."""
        results = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.data_type == 'f' and cell.value:
                        formula = str(cell.value)
                        cross_sheet_errors = self._find_cross_sheet_anchoring_errors(
                            workbook, sheet, formula, row, col
                        )
                        
                        for error in cross_sheet_errors:
                            probability = self._calculate_cross_sheet_error_probability(
                                workbook, sheet, error, row, col
                            )
                            if probability > 0.5:
                                from openpyxl.utils import get_column_letter
                                col_letter = get_column_letter(col)
                                results.append(ErrorDetectionResult(
                                    error_type=self.name,
                                    description=error['description'],
                                    probability=probability,
                                    severity=self.severity if probability >= 0.7 else ErrorSeverity.LOW,
                                    location=f"{sheet_name}!{col_letter}{row}",
                                    details=error['details'],
                                    suggested_fix=error['suggested_fix']
                                ))
        
        return results
    
    def _find_cross_sheet_anchoring_errors(self, workbook, sheet, formula: str, current_row: int, current_col: int) -> List[dict]:
        """Find cross-sheet anchoring errors in the formula."""
        errors = []
        
        # Extract cross-sheet references
        cross_sheet_refs = self._extract_cross_sheet_references(formula)
        
        for ref_info in cross_sheet_refs:
            sheet_name, cell_ref = ref_info['sheet_name'], ref_info['cell_ref']
            
            # Check if the referenced sheet exists
            if sheet_name not in workbook.sheetnames:
                continue
            
            ref_sheet = workbook[sheet_name]
            expected_anchoring = self._determine_expected_cross_sheet_anchoring(
                workbook, sheet, ref_sheet, cell_ref, current_row, current_col
            )
            actual_anchoring = self._get_anchoring_type_from_ref(cell_ref)
            
            if expected_anchoring != actual_anchoring:
                expected_ref = self._suggest_correct_cross_sheet_reference(
                    cell_ref, expected_anchoring
                )
                expected_formula = self._suggest_correct_cross_sheet_formula(
                    formula, ref_info['full_ref'], expected_ref, sheet_name
                )
                
                errors.append({
                    'description': f"Wrong cross-sheet anchoring: {sheet_name}!{cell_ref} should be {sheet_name}!{expected_ref}",
                    'details': {
                        'formula': formula,
                        'cross_sheet_reference': f"{sheet_name}!{cell_ref}",
                        'expected_reference': f"{sheet_name}!{expected_ref}",
                        'expected_formula': expected_formula,
                        'current_anchoring': actual_anchoring,
                        'expected_anchoring': expected_anchoring,
                        'current_row': current_row,
                        'current_col': current_col,
                        'referenced_sheet': sheet_name
                    },
                    'suggested_fix': f"Update cross-sheet reference: {expected_formula}"
                })
        
        return errors
    
    def _extract_cross_sheet_references(self, formula: str) -> List[dict]:
        """Extract cross-sheet references from formula."""
        refs = []
        
        # Pattern for quoted sheet names: 'Sheet Name'!A1 or 'Sheet Name'!A1:B3
        quoted_pattern = r"'([^']+)'!(\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?)"
        quoted_matches = re.findall(quoted_pattern, formula)
        for sheet_name, cell_ref in quoted_matches:
            refs.append({
                'sheet_name': sheet_name,
                'cell_ref': cell_ref,
                'full_ref': f"'{sheet_name}'!{cell_ref}"
            })
        
        # Pattern for unquoted sheet names: SheetName!A1 or SheetName!A1:B3
        unquoted_pattern = r'([A-Za-z][A-Za-z0-9_]*?)!(\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?)'
        unquoted_matches = re.findall(unquoted_pattern, formula)
        for sheet_name, cell_ref in unquoted_matches:
            # Skip if this is already captured as a quoted reference
            if not any(ref['sheet_name'] == sheet_name and ref['cell_ref'] == cell_ref for ref in refs):
                refs.append({
                    'sheet_name': sheet_name,
                    'cell_ref': cell_ref,
                    'full_ref': f"{sheet_name}!{cell_ref}"
                })
        
        return refs
    
    def _determine_expected_cross_sheet_anchoring(self, workbook, current_sheet, ref_sheet, cell_ref: str, current_row: int, current_col: int) -> str:
        """Determine the expected anchoring for a cross-sheet reference."""
        # Parse the cell reference
        ref_col, ref_row = self._parse_cell_reference(cell_ref)
        
        # Check if this is a fixed reference (like a constant or header)
        if self._is_fixed_cross_sheet_reference(ref_sheet, ref_col, ref_row):
            return "fully_locked"
        
        # Check if this is part of a range in a function (like VLOOKUP, SUM, etc.)
        if self._is_part_of_function_range(current_sheet, current_row, current_col):
            return "relative"  # Ranges in functions should typically be relative
        
        # Determine copy direction based on context
        copy_direction = self._determine_cross_sheet_copy_direction(
            workbook, current_sheet, current_row, current_col
        )
        
        if copy_direction == "across":
            return "column_locked"
        elif copy_direction == "down":
            return "row_locked"
        elif copy_direction == "both":
            return "relative"
        else:
            return "relative"  # Default to relative
    
    def _is_fixed_cross_sheet_reference(self, ref_sheet, ref_col: int, ref_row: int) -> bool:
        """Check if a cross-sheet reference should be fixed."""
        try:
            cell = ref_sheet.cell(row=ref_row, column=ref_col)
            
            # Check if it's a header row (row 1)
            if ref_row == 1:
                return True
            
            # Check if it's a constant value (same value in multiple cells)
            if cell.value is not None:
                constant_count = 0
                total_cells = 0
                
                # Check the column for consistency
                for row in range(1, min(ref_row + 10, ref_sheet.max_row + 1)):
                    check_cell = ref_sheet.cell(row=row, column=ref_col)
                    if check_cell.value is not None:
                        total_cells += 1
                        if check_cell.value == cell.value:
                            constant_count += 1
                
                # If more than 50% of cells have the same value, consider it fixed
                if total_cells > 0 and (constant_count / total_cells) > 0.5:
                    return True
            
            return False
        except:
            return False
    
    def _determine_cross_sheet_copy_direction(self, workbook, sheet, row: int, col: int) -> str:
        """Determine the likely copy direction for cross-sheet formulas."""
        # Check if this cell is part of a copied pattern
        copied_pattern = self._find_cross_sheet_copied_pattern(sheet, row, col)
        
        if copied_pattern:
            return copied_pattern['direction']
        
        # Default based on position and surrounding formulas
        return self._guess_cross_sheet_copy_direction(sheet, row, col)
    
    def _find_cross_sheet_copied_pattern(self, sheet, row: int, col: int) -> Optional[dict]:
        """Find if the cell is part of a copied pattern with cross-sheet references."""
        # Look for similar formulas in adjacent cells
        similar_formulas = []
        
        # Check horizontal pattern (copying across)
        for c in range(max(1, col - 3), min(col + 4, sheet.max_column + 1)):
            if c != col:
                cell = sheet.cell(row=row, column=c)
                if cell.data_type == 'f' and cell.value:
                    if self._has_cross_sheet_reference(str(cell.value)):
                        similar_formulas.append({
                            'col': c,
                            'row': row,
                            'formula': str(cell.value)
                        })
        
        # Check vertical pattern (copying down)
        for r in range(max(1, row - 3), min(row + 4, sheet.max_row + 1)):
            if r != row:
                cell = sheet.cell(row=r, column=col)
                if cell.data_type == 'f' and cell.value:
                    if self._has_cross_sheet_reference(str(cell.value)):
                        similar_formulas.append({
                            'col': col,
                            'row': r,
                            'formula': str(cell.value)
                        })
        
        if len(similar_formulas) >= 2:
            # Determine direction based on pattern
            horizontal_count = len([f for f in similar_formulas if f['row'] == row])
            vertical_count = len([f for f in similar_formulas if f['col'] == col])
            
            if horizontal_count > vertical_count:
                return {'direction': 'across', 'count': horizontal_count}
            elif vertical_count > horizontal_count:
                return {'direction': 'down', 'count': vertical_count}
            else:
                return {'direction': 'both', 'count': len(similar_formulas)}
        
        return None
    
    def _has_cross_sheet_reference(self, formula: str) -> bool:
        """Check if formula contains cross-sheet references."""
        return bool(re.search(r"'?[^']+'?!", formula))
    
    def _guess_cross_sheet_copy_direction(self, sheet, row: int, col: int) -> str:
        """Guess copy direction based on cell position and context."""
        # If near the top, likely copying down
        if row <= 3:
            return "down"
        # If near the left, likely copying across
        elif col <= 3:
            return "across"
        # Default to both directions
        else:
            return "both"
    
    def _parse_cell_reference(self, cell_ref: str) -> Tuple[int, int]:
        """Parse cell reference to get column and row numbers."""
        # Remove any anchoring symbols
        clean_ref = cell_ref.replace('$', '')
        
        # Extract column letters and row number
        col_match = re.match(r'([A-Z]+)', clean_ref)
        row_match = re.search(r'(\d+)$', clean_ref)
        
        if col_match and row_match:
            col_letters = col_match.group(1)
            row_num = int(row_match.group(1))
            col_num = openpyxl.utils.column_index_from_string(col_letters)
            return col_num, row_num
        
        return 1, 1  # Default fallback
    
    def _get_anchoring_type_from_ref(self, cell_ref: str) -> str:
        """Get anchoring type from cell reference."""
        if cell_ref.startswith('$') and '$' in cell_ref[1:]:
            return "fully_locked"
        elif cell_ref.startswith('$'):
            return "column_locked"
        elif '$' in cell_ref:
            return "row_locked"
        else:
            return "relative"
    
    def _suggest_correct_cross_sheet_reference(self, cell_ref: str, expected_anchoring: str) -> str:
        """Suggest the correct cross-sheet reference with proper anchoring."""
        # Remove existing anchoring
        clean_ref = cell_ref.replace('$', '')
        
        # Extract column letters and row number
        col_match = re.match(r'([A-Z]+)', clean_ref)
        row_match = re.search(r'(\d+)$', clean_ref)
        
        if not col_match or not row_match:
            return cell_ref  # Return original if parsing fails
        
        col_letters = col_match.group(1)
        row_num = row_match.group(1)
        
        # Add appropriate anchoring
        if expected_anchoring == "fully_locked":
            return f"${col_letters}${row_num}"
        elif expected_anchoring == "column_locked":
            return f"${col_letters}{row_num}"
        elif expected_anchoring == "row_locked":
            return f"{col_letters}${row_num}"
        else:  # relative
            return f"{col_letters}{row_num}"
    
    def _suggest_correct_cross_sheet_formula(self, formula: str, current_ref: str, expected_ref: str, sheet_name: str) -> str:
        """Suggest the correct formula with proper cross-sheet anchoring."""
        # Extract the cell reference part from current_ref
        cell_ref_match = re.search(r'!(\$?[A-Z]+\$?\d+)$', current_ref)
        if not cell_ref_match:
            return formula
        
        current_cell_ref = cell_ref_match.group(1)
        
        # Handle both quoted and unquoted sheet names
        if "'" in current_ref:
            # Quoted sheet name
            old_pattern = f"'{sheet_name}'!{current_cell_ref}"
            new_pattern = f"'{sheet_name}'!{expected_ref}"
        else:
            # Unquoted sheet name
            old_pattern = f"{sheet_name}!{current_cell_ref}"
            new_pattern = f"{sheet_name}!{expected_ref}"
        
        return formula.replace(old_pattern, new_pattern)
    
    def _calculate_cross_sheet_error_probability(self, workbook, sheet, error: dict, current_row: int, current_col: int) -> float:
        """Calculate probability for cross-sheet anchoring error."""
        base_prob = 0.6
        
        # Higher probability for critical calculations
        if self._is_critical_cross_sheet_calculation(sheet, current_row, current_col):
            base_prob += 0.2
        
        # Higher probability for complex cross-sheet references
        if self._is_complex_cross_sheet_reference(error['details']['formula']):
            base_prob += 0.1
        
        # Higher probability for frequently copied patterns
        copied_pattern = self._find_cross_sheet_copied_pattern(sheet, current_row, current_col)
        if copied_pattern and copied_pattern['count'] > 3:
            base_prob += 0.1
        
        # Higher probability for wrong anchoring type
        current_anchoring = error['details']['current_anchoring']
        expected_anchoring = error['details']['expected_anchoring']
        
        if current_anchoring == "fully_locked" and expected_anchoring != "fully_locked":
            base_prob += 0.1
        elif current_anchoring == "relative" and expected_anchoring in ["column_locked", "row_locked"]:
            base_prob += 0.1
        
        return min(0.9, base_prob)
    
    def _is_critical_cross_sheet_calculation(self, sheet, row: int, col: int) -> bool:
        """Check if this is a critical cross-sheet calculation."""
        cell = sheet.cell(row=row, column=col)
        if not cell.data_type == 'f':
            return False
        
        formula = str(cell.value).upper()
        
        # Critical functions that often use cross-sheet references
        critical_functions = ['VLOOKUP', 'HLOOKUP', 'INDEX', 'MATCH', 'SUMIF', 'COUNTIF', 'AVERAGEIF']
        
        return any(func in formula for func in critical_functions)
    
    def _is_part_of_function_range(self, sheet, row: int, col: int) -> bool:
        """Check if the cell is part of a function range."""
        cell = sheet.cell(row=row, column=col)
        if not cell.data_type == 'f':
            return False
        
        formula = str(cell.value).upper()
        
        # Functions that typically use ranges
        range_functions = ['VLOOKUP', 'HLOOKUP', 'SUM', 'AVERAGE', 'COUNT', 'MAX', 'MIN', 'INDEX', 'MATCH']
        
        return any(func in formula for func in range_functions)
    
    def _is_complex_cross_sheet_reference(self, formula: str) -> bool:
        """Check if formula has complex cross-sheet references."""
        cross_sheet_refs = self._extract_cross_sheet_references(formula)
        return len(cross_sheet_refs) > 1


# ============================================================================
# MAIN FUNCTION
# ============================================================================

def detect_excel_errors_probabilistic(
    file_path: Path, 
    error_threshold: float = 0.7,
    output_dir: Optional[Path] = None
) -> Dict[str, Any]:
    """
    Detect Excel errors using probabilistic models.
    
    Args:
        file_path: Path to the Excel file
        error_threshold: Minimum probability threshold (0.0 to 1.0)
        output_dir: Optional directory to save reports
        
    Returns:
        Dictionary containing detection results
    """
    sniffer = ProbabilisticErrorSniffer(file_path, error_threshold)
    
    # Register detectors
    sniffer.register_detector(HiddenDataInRangesDetector())
    sniffer.register_detector(CircularNamedRangesDetector())
    sniffer.register_detector(InconsistentDateFormatsDetector())
    sniffer.register_detector(ArrayFormulaSpillErrorsDetector())
    sniffer.register_detector(VolatileFunctionsDetector())
    sniffer.register_detector(CrossSheetReferenceErrorsDetector())
    sniffer.register_detector(DataTypeInconsistenciesInLookupTablesDetector())
    sniffer.register_detector(ConditionalFormattingOverlapConflictsDetector())
    sniffer.register_detector(ExternalDataConnectionFailuresDetector())
    sniffer.register_detector(PrecisionErrorsInFinancialCalculationsDetector())
    sniffer.register_detector(IncompleteDragFormulaDetector())
    sniffer.register_detector(FalseRangeEndDetectionDetector())
    sniffer.register_detector(PartialFormulaPropagationDetector())
    sniffer.register_detector(FormulaBoundaryMismatchDetector())
    sniffer.register_detector(CopyPasteFormulaGapsDetector())
    sniffer.register_detector(FormulaRangeVsDataRangeDiscrepancyDetector())
    sniffer.register_detector(InconsistentFormulaApplicationDetector())
    sniffer.register_detector(MissingDollarSignAnchorsDetector())
    sniffer.register_detector(WrongRowColumnAnchoringDetector())
    sniffer.register_detector(OverAnchoredReferencesDetector())
    sniffer.register_detector(InconsistentAnchoringInRangesDetector())
    sniffer.register_detector(LookupFunctionAnchoringDetector())
    sniffer.register_detector(ArrayFormulaAnchoringDetector())
    sniffer.register_detector(CrossSheetAnchoringDetector())
    
    # Run detection
    results = sniffer.detect_all_errors()
    
    # Save reports if output directory specified
    if output_dir:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Ensure file_path is a Path object
        file_path = Path(file_path)
        
        # Save JSON report
        json_path = output_dir / f"{file_path.stem}_probabilistic_errors.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            # Convert results to serializable format
            serializable_results = {}
            for key, value in results.items():
                if key == 'summary':
                    serializable_results[key] = value
                else:
                    serializable_results[key] = [
                        {
                            'error_type': r.error_type,
                            'description': r.description,
                            'probability': r.probability,
                            'severity': r.severity.value,
                            'location': r.location,
                            'details': r.details,
                            'suggested_fix': r.suggested_fix
                        }
                        for r in value
                    ]
            json.dump(serializable_results, f, indent=2, default=str)
        
        logger.info(f"Probabilistic error analysis saved to: {json_path}")
    
    return results


if __name__ == '__main__':
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python probabilistic_error_detector.py <excel_file> [threshold] [output_dir]")
        sys.exit(1)
    
    file_path = Path(sys.argv[1])
    threshold = float(sys.argv[2]) if len(sys.argv) > 2 else 0.7
    output_dir = Path(sys.argv[3]) if len(sys.argv) > 3 else None
    
    try:
        results = detect_excel_errors_probabilistic(file_path, threshold, output_dir)
        print(f"Probabilistic error detection completed. Found {results['summary']['total_errors']} errors above threshold {threshold}.")
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1) 