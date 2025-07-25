import pytest
import openpyxl
from src.excel_analyzer.probabilistic_error_detector import FormulaRangeVsDataRangeDiscrepancyDetector, ErrorSeverity

def create_sheet_with_lookup_formula(data_rows, data_cols, formula_end_row, formula_end_col):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Fill data
    for row in range(1, data_rows + 1):
        for col in range(1, data_cols + 1):
            ws.cell(row=row, column=col).value = f"{row}-{col}"
    # Place VLOOKUP formula in row 1, column 3
    ws.cell(row=1, column=3).value = f"=VLOOKUP(A1,A1:{openpyxl.utils.get_column_letter(formula_end_col)}{formula_end_row},2)"
    ws.cell(row=1, column=3).data_type = 'f'
    return wb

def test_lookup_range_covers_all_data():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Fill data only in columns A and B, up to row 50
    for row in range(1, 51):
        ws.cell(row=row, column=1).value = f"{row}-1"
        ws.cell(row=row, column=2).value = f"{row}-2"
    # Place VLOOKUP formula in row 51, column D (not in data range)
    ws.cell(row=51, column=4).value = "=VLOOKUP(A1,A1:B50,2)"
    ws.cell(row=51, column=4).data_type = 'f'
    detector = FormulaRangeVsDataRangeDiscrepancyDetector()
    results = detector.detect(wb)
    assert not results

def test_lookup_range_misses_some_data():
    wb = create_sheet_with_lookup_formula(100, 3, 50, 2)
    detector = FormulaRangeVsDataRangeDiscrepancyDetector()
    results = detector.detect(wb)
    assert results
    r = results[0]
    assert r.details['referenced_range'] == 'A1:B50'
    assert r.details['max_data_row'] == 100
    assert r.details['max_data_col'] == 3
    assert r.probability > 0.5
    assert r.severity in (ErrorSeverity.HIGH, ErrorSeverity.MEDIUM)

def test_non_lookup_formula():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in range(1, 20):
        ws.cell(row=row, column=1).value = row
    ws.cell(row=1, column=2).value = "=SUM(A1:A10)"
    ws.cell(row=1, column=2).data_type = 'f'
    detector = FormulaRangeVsDataRangeDiscrepancyDetector()
    results = detector.detect(wb)
    assert not results

def test_no_range_in_formula():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1).value = "=VLOOKUP(A1,B1,1)"
    ws.cell(row=1, column=1).data_type = 'f'
    detector = FormulaRangeVsDataRangeDiscrepancyDetector()
    results = detector.detect(wb)
    assert not results 