import pytest
import openpyxl
from src.excel_analyzer.probabilistic_error_detector import InconsistentAnchoringInRangesDetector, ErrorSeverity

def create_sheet_with_inconsistent_ranges():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Set up data
    for row in range(1, 11):
        ws.cell(row=row, column=1).value = row
        ws.cell(row=row, column=2).value = row * 10
    
    # Add formulas with inconsistent anchoring in ranges
    ws.cell(row=1, column=3).value = "=SUM($A$1:A10)"  # Inconsistent: start anchored, end not
    ws.cell(row=1, column=3).data_type = 'f'
    
    ws.cell(row=2, column=3).value = "=AVERAGE(A1:$B$10)"  # Inconsistent: start not anchored, end anchored
    ws.cell(row=2, column=3).data_type = 'f'
    
    ws.cell(row=3, column=3).value = "=VLOOKUP(A1,$B$1:C$10,2)"  # Inconsistent in lookup range
    ws.cell(row=3, column=3).data_type = 'f'
    
    ws.cell(row=4, column=3).value = "=SUM(A1:B10)"  # Consistent: both relative
    ws.cell(row=4, column=3).data_type = 'f'
    
    ws.cell(row=5, column=3).value = "=SUM($A$1:$B$10)"  # Consistent: both anchored
    ws.cell(row=5, column=3).data_type = 'f'
    
    return wb

def test_consistent_anchoring():
    wb = create_sheet_with_inconsistent_ranges()
    detector = InconsistentAnchoringInRangesDetector()
    results = detector.detect(wb)
    
    # Should not flag consistently anchored ranges
    consistent_results = [r for r in results if "A1:B10" in r.details['inconsistent_range'] or "$A$1:$B$10" in r.details['inconsistent_range']]
    assert len(consistent_results) == 0

def test_inconsistent_anchoring_in_sum():
    wb = create_sheet_with_inconsistent_ranges()
    detector = InconsistentAnchoringInRangesDetector()
    results = detector.detect(wb)
    
    # Should flag inconsistent anchoring in SUM function
    sum_results = [r for r in results if "SUM" in r.details['formula'] and "$A$1:A10" in r.details['inconsistent_range']]
    assert len(sum_results) == 1
    assert sum_results[0].probability > 0.7  # High probability for calculation functions
    assert sum_results[0].severity == ErrorSeverity.MEDIUM

def test_inconsistent_anchoring_in_vlookup():
    wb = create_sheet_with_inconsistent_ranges()
    detector = InconsistentAnchoringInRangesDetector()
    results = detector.detect(wb)
    
    # Should flag inconsistent anchoring in VLOOKUP range
    vlookup_results = [r for r in results if "VLOOKUP" in r.details['formula'] and "$B$1:C$10" in r.details['inconsistent_range']]
    assert len(vlookup_results) == 1
    assert vlookup_results[0].probability > 0.7  # High probability for lookup functions
    assert vlookup_results[0].severity == ErrorSeverity.MEDIUM

def test_mixed_partial_anchoring():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Set up data
    for row in range(1, 6):
        ws.cell(row=row, column=1).value = row
        ws.cell(row=row, column=2).value = row * 10
    
    # Add formula with mixed partial anchoring
    ws.cell(row=1, column=3).value = "=SUM($A1:B$5)"  # Mixed partial anchoring
    ws.cell(row=1, column=3).data_type = 'f'
    
    detector = InconsistentAnchoringInRangesDetector()
    results = detector.detect(wb)
    
    # Should flag mixed partial anchoring
    mixed_results = [r for r in results if "$A1:B$5" in r.details['inconsistent_range']]
    assert len(mixed_results) == 1
    assert mixed_results[0].probability > 0.5
    assert mixed_results[0].severity in (ErrorSeverity.MEDIUM, ErrorSeverity.LOW)

def test_no_ranges():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Add formulas without ranges
    ws.cell(row=1, column=2).value = "=A1+B1"
    ws.cell(row=1, column=2).data_type = 'f'
    
    ws.cell(row=2, column=2).value = "=SUM(A1)"
    ws.cell(row=2, column=2).data_type = 'f'
    
    detector = InconsistentAnchoringInRangesDetector()
    results = detector.detect(wb)
    assert len(results) == 0

def test_calculation_functions():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Set up data
    for row in range(1, 6):
        ws.cell(row=row, column=1).value = row
    
    # Add calculation functions with inconsistent ranges
    ws.cell(row=1, column=2).value = "=COUNT($A$1:A5)"  # Inconsistent
    ws.cell(row=1, column=2).data_type = 'f'
    
    ws.cell(row=2, column=2).value = "=MAX(A1:$A$5)"  # Inconsistent
    ws.cell(row=2, column=2).data_type = 'f'
    
    detector = InconsistentAnchoringInRangesDetector()
    results = detector.detect(wb)
    
    # Should flag both calculation functions
    calc_results = [r for r in results if any(func in r.details['formula'] for func in ["COUNT", "MAX"])]
    assert len(calc_results) == 2
    
    for r in calc_results:
        assert r.probability > 0.7  # High probability for calculation functions
        assert r.severity == ErrorSeverity.MEDIUM

def test_high_severity_inconsistency():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Set up data
    for row in range(1, 6):
        ws.cell(row=row, column=1).value = row
    
    # Add formula with high severity inconsistency (fully locked vs relative)
    ws.cell(row=1, column=2).value = "=SUM($A$1:A5)"  # Fully locked vs relative
    ws.cell(row=1, column=2).data_type = 'f'
    
    detector = InconsistentAnchoringInRangesDetector()
    results = detector.detect(wb)
    
    # Should flag high severity inconsistency
    high_severity_results = [r for r in results if "$A$1:A5" in r.details['inconsistent_range']]
    assert len(high_severity_results) == 1
    assert high_severity_results[0].probability > 0.7  # Higher probability for high severity 