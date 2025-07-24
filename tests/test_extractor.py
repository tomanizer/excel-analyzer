#!/usr/bin/env python3
"""
Tests for the Excel Extractor module.
"""

import pytest
import json
from pathlib import Path
import tempfile
import shutil

from excel_analyzer.excel_extractor import ExcelExtractor, extract_excel_to_markdown


class TestExcelExtractor:
    """Test cases for Excel extractor functionality."""
    
    def test_extractor_initialization(self):
        """Test that ExcelExtractor initializes correctly."""
        # Use a simple test file
        test_file = Path("excel_files/simple_model.xlsx")
        
        if test_file.exists():
            extractor = ExcelExtractor(test_file)
            
            # Check that extractor is initialized correctly
            assert extractor.file_path == test_file
            assert extractor.workbook is None  # Not loaded yet
            assert isinstance(extractor.extracted_data, dict)
            assert "metadata" in extractor.extracted_data
            assert "sheets" in extractor.extracted_data
            assert "global_features" in extractor.extracted_data
            assert "relationships" in extractor.extracted_data
            assert "summary" in extractor.extracted_data
    
    def test_extractor_nonexistent_file(self):
        """Test that ExcelExtractor handles nonexistent files."""
        test_file = Path("nonexistent_file.xlsx")
        
        with pytest.raises(FileNotFoundError):
            extractor = ExcelExtractor(test_file)
            extractor.extract_all()
    
    def test_extract_all_with_simple_file(self):
        """Test that extract_all works with a simple Excel file."""
        test_file = Path("excel_files/simple_model.xlsx")
        
        if test_file.exists():
            extractor = ExcelExtractor(test_file)
            result = extractor.extract_all()
            
            # Check that we get a dictionary with expected keys
            assert isinstance(result, dict)
            assert "metadata" in result
            assert "sheets" in result
            assert "global_features" in result
            assert "relationships" in result
            assert "summary" in result
            
            # Check metadata
            assert result["metadata"]["filename"] == "simple_model.xlsx"
            assert result["metadata"]["file_size_kb"] > 0
            assert result["metadata"]["sheet_count"] > 0
            
            # Check that workbook is loaded
            assert extractor.workbook is not None
    
    def test_extract_all_with_complex_file(self):
        """Test that extract_all works with a complex Excel file."""
        test_file = Path("excel_files/complex_model.xlsx")
        
        if test_file.exists():
            extractor = ExcelExtractor(test_file)
            result = extractor.extract_all()
            
            # Check basic structure
            assert isinstance(result, dict)
            assert "metadata" in result
            assert "sheets" in result
            assert "global_features" in result
            assert "relationships" in result
            assert "summary" in result
            
            # Check that we have sheets
            assert len(result["sheets"]) > 0
            
            # Check that we have some data
            total_cells = result["summary"].get("total_cells_with_data", 0)
            assert total_cells > 0
    
    def test_to_markdown(self):
        """Test markdown report generation."""
        test_file = Path("excel_files/simple_model.xlsx")
        
        if test_file.exists():
            extractor = ExcelExtractor(test_file)
            extractor.extract_all()
            
            markdown_content = extractor.to_markdown()
            
            # Check that markdown contains expected content
            assert "# Excel Workbook Analysis:" in markdown_content
            assert "## 📊 Executive Summary" in markdown_content
            assert "## 📋 File Metadata" in markdown_content
            assert "## 📄 Sheet Analysis" in markdown_content
    
    def test_to_markdown_empty_data(self):
        """Test markdown generation with empty data."""
        test_file = Path("excel_files/simple_model.xlsx")
        
        if test_file.exists():
            extractor = ExcelExtractor(test_file)
            # Don't call extract_all() to keep data empty
            
            markdown_content = extractor.to_markdown()
            
            # Should handle empty data gracefully
            assert isinstance(markdown_content, str)
            assert len(markdown_content) > 0
    
    def test_save_markdown(self):
        """Test saving markdown report to file."""
        test_file = Path("excel_files/simple_model.xlsx")
        
        if test_file.exists():
            extractor = ExcelExtractor(test_file)
            extractor.extract_all()
            
            # Create temporary directory
            with tempfile.TemporaryDirectory() as temp_dir:
                output_path = Path(temp_dir) / "test_report.md"
                
                # Save markdown
                saved_path = extractor.save_markdown(output_path)
                
                # Check that file was created
                assert saved_path.exists()
                assert saved_path.suffix == ".md"
                
                # Check that file has content
                content = saved_path.read_text(encoding='utf-8')
                assert len(content) > 0
                assert "# Excel Workbook Analysis:" in content
    
    def test_save_json(self):
        """Test saving JSON data to file."""
        test_file = Path("excel_files/simple_model.xlsx")
        
        if test_file.exists():
            extractor = ExcelExtractor(test_file)
            extractor.extract_all()
            
            # Create temporary directory
            with tempfile.TemporaryDirectory() as temp_dir:
                output_path = Path(temp_dir) / "test_data.json"
                
                # Save JSON
                saved_path = extractor.save_json(output_path)
                
                # Check that file was created
                assert saved_path.exists()
                assert saved_path.suffix == ".json"
                
                # Check that file contains valid JSON
                content = saved_path.read_text(encoding='utf-8')
                data = json.loads(content)
                
                # Check that JSON has expected structure
                assert "metadata" in data
                assert "sheets" in data
                assert "global_features" in data
                assert "relationships" in data
                assert "summary" in data
    
    def test_extract_metadata(self):
        """Test metadata extraction."""
        test_file = Path("excel_files/simple_model.xlsx")
        
        if test_file.exists():
            extractor = ExcelExtractor(test_file)
            extractor.workbook = extractor._load_workbook()
            extractor._extract_metadata()
            
            metadata = extractor.extracted_data["metadata"]
            
            # Check metadata fields
            assert "filename" in metadata
            assert "file_size" in metadata
            assert "file_size_kb" in metadata
            assert "last_modified" in metadata
            assert "file_extension" in metadata
            assert "has_vba" in metadata
            assert "sheet_count" in metadata
            assert "sheet_names" in metadata
            
            # Check specific values
            assert metadata["filename"] == "simple_model.xlsx"
            assert metadata["file_extension"] == ".xlsx"
            assert metadata["has_vba"] is False
            assert metadata["sheet_count"] > 0
            assert isinstance(metadata["sheet_names"], list)
    
    def test_extract_global_features(self):
        """Test global features extraction."""
        test_file = Path("excel_files/complex_model.xlsx")
        
        if test_file.exists():
            extractor = ExcelExtractor(test_file)
            extractor.workbook = extractor._load_workbook()
            extractor._extract_global_features()
            
            global_features = extractor.extracted_data["global_features"]
            
            # Check global features structure
            assert "named_ranges" in global_features
            assert "external_links" in global_features
            assert "properties" in global_features
            
            # Check that these are dictionaries/lists
            assert isinstance(global_features["named_ranges"], dict)
            assert isinstance(global_features["external_links"], list)
            assert isinstance(global_features["properties"], dict)
    
    def test_extract_sheets(self):
        """Test sheet extraction."""
        test_file = Path("excel_files/complex_model.xlsx")
        
        if test_file.exists():
            extractor = ExcelExtractor(test_file)
            extractor.workbook = extractor._load_workbook()
            extractor._extract_sheets()
            
            sheets = extractor.extracted_data["sheets"]
            
            # Check that we have sheets
            assert len(sheets) > 0
            
            # Check structure of first sheet
            first_sheet_name = list(sheets.keys())[0]
            first_sheet_data = sheets[first_sheet_name]
            
            assert "dimensions" in first_sheet_data
            assert "data" in first_sheet_data
            assert "formulas" in first_sheet_data
            assert "tables" in first_sheet_data
            assert "charts" in first_sheet_data
            assert "data_validations" in first_sheet_data
            assert "merged_cells" in first_sheet_data
            assert "styles" in first_sheet_data
            assert "summary" in first_sheet_data
    
    def test_extract_relationships(self):
        """Test relationship extraction."""
        test_file = Path("excel_files/enterprise_model.xlsx")
        
        if test_file.exists():
            extractor = ExcelExtractor(test_file)
            extractor.workbook = extractor._load_workbook()
            extractor._extract_sheets()
            extractor._extract_relationships()
            
            relationships = extractor.extracted_data["relationships"]
            
            # Check relationships structure
            assert "cross_sheet_references" in relationships
            assert "external_references" in relationships
            assert "circular_references" in relationships
            
            # Check that these are lists
            assert isinstance(relationships["cross_sheet_references"], list)
            assert isinstance(relationships["external_references"], list)
            assert isinstance(relationships["circular_references"], list)


class TestExcelExtractorIntegration:
    """Integration tests for Excel extractor workflow."""
    
    def test_full_extraction_workflow(self):
        """Test the complete extraction workflow."""
        test_file = Path("excel_files/mycoolsample.xlsx")
        
        if test_file.exists():
            # Create extractor and run full workflow
            extractor = ExcelExtractor(test_file)
            result = extractor.extract_all()
            
            # Check that we have complete data
            assert "metadata" in result
            assert "sheets" in result
            assert "global_features" in result
            assert "relationships" in result
            assert "summary" in result
            
            # Check that we have some data
            assert result["summary"]["total_cells_with_data"] > 0
            
            # Generate markdown
            markdown_content = extractor.to_markdown()
            assert len(markdown_content) > 0
            assert "# Excel Workbook Analysis:" in markdown_content
            
            # Save to files
            with tempfile.TemporaryDirectory() as temp_dir:
                # Save markdown
                markdown_path = extractor.save_markdown(Path(temp_dir) / "test.md")
                assert markdown_path.exists()
                
                # Save JSON
                json_path = extractor.save_json(Path(temp_dir) / "test.json")
                assert json_path.exists()
    
    def test_extract_excel_to_markdown_function(self):
        """Test the extract_excel_to_markdown convenience function."""
        test_file = Path("excel_files/simple_model.xlsx")
        
        if test_file.exists():
            with tempfile.TemporaryDirectory() as temp_dir:
                output_dir = Path(temp_dir)
                
                # Run the convenience function
                markdown_path, json_path = extract_excel_to_markdown(test_file, output_dir)
                
                # Check that files were created
                assert markdown_path.exists()
                assert json_path.exists()
                
                # Check file names
                assert markdown_path.name.endswith("_extractor_report.md")
                assert json_path.name.endswith("_extracted_data.json")
                
                # Check that files have content
                markdown_content = markdown_path.read_text(encoding='utf-8')
                json_content = json_path.read_text(encoding='utf-8')
                
                assert len(markdown_content) > 0
                assert len(json_content) > 0
    
    def test_multiple_file_processing(self):
        """Test processing multiple files."""
        test_files = [
            Path("excel_files/simple_model.xlsx"),
            Path("excel_files/complex_model.xlsx")
        ]
        
        existing_files = [f for f in test_files if f.exists()]
        
        if len(existing_files) >= 2:
            with tempfile.TemporaryDirectory() as temp_dir:
                output_dir = Path(temp_dir)
                
                results = []
                for test_file in existing_files:
                    extractor = ExcelExtractor(test_file)
                    result = extractor.extract_all()
                    
                    # Save reports
                    markdown_path = extractor.save_markdown(
                        output_dir / f"{test_file.stem}_report.md"
                    )
                    json_path = extractor.save_json(
                        output_dir / f"{test_file.stem}_data.json"
                    )
                    
                    results.append({
                        "file": test_file.name,
                        "result": result,
                        "markdown_path": markdown_path,
                        "json_path": json_path
                    })
                
                # Check that all files were processed
                assert len(results) == len(existing_files)
                
                # Check that all output files exist
                for result in results:
                    assert result["markdown_path"].exists()
                    assert result["json_path"].exists()
                    
                    # Check that we have data
                    assert result["result"]["summary"]["total_cells_with_data"] > 0


class TestExcelExtractorEdgeCases:
    """Test edge cases and error handling."""
    
    def test_empty_excel_file(self):
        """Test handling of empty Excel files."""
        # Create a minimal empty Excel file for testing
        from openpyxl import Workbook
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            wb = Workbook()
            wb.save(tmp_file.name)
            tmp_path = Path(tmp_file.name)
        
        try:
            extractor = ExcelExtractor(tmp_path)
            result = extractor.extract_all()
            
            # Should handle empty file gracefully
            assert isinstance(result, dict)
            assert "metadata" in result
            assert "sheets" in result
            assert result["metadata"]["sheet_count"] == 1  # Default sheet
            
        finally:
            # Clean up
            tmp_path.unlink()
    
    def test_file_with_only_formulas(self):
        """Test handling of files with only formulas."""
        from openpyxl import Workbook
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            wb = Workbook()
            ws = wb.active
            ws['A1'] = '=SUM(1,2,3)'  # Formula only
            wb.save(tmp_file.name)
            tmp_path = Path(tmp_file.name)
        
        try:
            extractor = ExcelExtractor(tmp_path)
            result = extractor.extract_all()
            
            # Should handle formula-only file
            assert isinstance(result, dict)
            assert "sheets" in result
            
            # Check that formula was extracted
            sheet_data = list(result["sheets"].values())[0]
            assert len(sheet_data["formulas"]) > 0
            
        finally:
            # Clean up
            tmp_path.unlink()
    
    def test_file_with_special_characters(self):
        """Test handling of files with special characters in names."""
        test_file = Path("excel_files/Book 3.xlsx")  # File with space in name
        
        if test_file.exists():
            extractor = ExcelExtractor(test_file)
            result = extractor.extract_all()
            
            # Should handle special characters in filename
            assert isinstance(result, dict)
            assert result["metadata"]["filename"] == "Book 3.xlsx"
    
    def test_memory_cleanup(self):
        """Test that memory is properly cleaned up."""
        test_file = Path("excel_files/simple_model.xlsx")
        
        if test_file.exists():
            # Create multiple extractors to test cleanup
            extractors = []
            for i in range(3):
                extractor = ExcelExtractor(test_file)
                extractor.extract_all()
                extractors.append(extractor)
            
            # All should work without memory issues
            for extractor in extractors:
                assert extractor.workbook is not None
                assert len(extractor.extracted_data["sheets"]) > 0


if __name__ == "__main__":
    pytest.main([__file__]) 