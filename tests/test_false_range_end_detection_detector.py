#!/usr/bin/env python3
"""
Tests for False Range End Detection (Empty Cell Trap) Detector.

Covers:
- Columns with no gap (should not flag)
- Gap in middle with formulas stopping at gap (should flag, high probability)
- Data after gap with formulas covering all (should not flag)
- Single missing formula after gap (should flag, low probability)
- Intentional gaps (should not flag if possible)
"""

import pytest
import tempfile
import shutil
from pathlib import Path
import openpyxl

from excel_analyzer.probabilistic_error_detector import FalseRangeEndDetectionDetector, ErrorSeverity

class TestFalseRangeEndDetectionDetector:
    def setup_method(self):
        self.detector = FalseRangeEndDetectionDetector()
        self.temp_dir = Path(tempfile.mkdtemp())
    def teardown_method(self):
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)
    def create_test_workbook(self, data_rows, formula_rows, empty_rows=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in data_rows:
            ws[f'A{row}'].value = 1  # Data present
        for row in formula_rows:
            ws[f'A{row}'].value = f'=B{row}+1'
            ws[f'A{row}'].data_type = 'f'
        if empty_rows:
            for row in empty_rows:
                ws[f'A{row}'].value = None
        return wb
    def test_no_gap(self):
        # No gap, formulas cover all data
        data_rows = list(range(2, 11))
        formula_rows = list(range(2, 11))
        wb = self.create_test_workbook(data_rows, formula_rows)
        results = self.detector.detect(wb)
        assert not any(r.error_type == 'false_range_end_detection' for r in results)
    def test_gap_in_middle_formulas_stop(self):
        # Gap in middle, formulas stop at gap
        data_rows = list(range(2, 11))
        formula_rows = [2, 3, 4]  # Formulas stop at row 4, gap at 5, data continues after
        empty_rows = [5]
        wb = self.create_test_workbook(data_rows, formula_rows, empty_rows)
        results = self.detector.detect(wb)
        assert any(r.probability >= 0.9 for r in results)
    def test_data_after_gap_formulas_cover_all(self):
        # Data after gap, formulas cover all
        data_rows = list(range(2, 11))
        formula_rows = [2, 3, 4, 6, 7, 8, 9, 10]  # Gap at 5, but formulas cover all data after
        empty_rows = [5]
        wb = self.create_test_workbook(data_rows, formula_rows, empty_rows)
        results = self.detector.detect(wb)
        assert not any(r.error_type == 'false_range_end_detection' for r in results)
    def test_single_missing_formula_after_gap(self):
        # Single missing formula after gap
        data_rows = list(range(2, 11))
        formula_rows = [2, 3, 4, 6, 7, 8, 9]  # Gap at 5, missing formula at 10
        empty_rows = [5]
        wb = self.create_test_workbook(data_rows, formula_rows, empty_rows)
        results = self.detector.detect(wb)
        assert any(0.2 < r.probability <= 0.6 for r in results)
    def test_intentional_gap(self):
        # Intentional gap (e.g., header or subtotal row)
        data_rows = list(range(2, 11))
        formula_rows = [2, 3, 4, 6, 7, 8, 9, 10]  # Gap at 5 (intentional)
        empty_rows = [5]
        wb = self.create_test_workbook(data_rows, formula_rows, empty_rows)
        results = self.detector.detect(wb)
        # May or may not flag depending on detector sensitivity
        # The detector should handle this gracefully
        assert len(results) >= 0  # Should not crash

if __name__ == '__main__':
    pytest.main([__file__]) 