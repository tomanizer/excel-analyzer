import pytest
import openpyxl
from src.excel_analyzer.probabilistic_error_detector import CopyPasteFormulaGapsDetector, ErrorSeverity

def create_sheet_with_formula_gaps(formula_rows, gap_rows=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    gap_rows = set(gap_rows or [])
    for row in range(1, 21):
        cell = ws.cell(row=row, column=1)
        if row in formula_rows:
            cell.value = f"=A{row}+1"
            cell.data_type = 'f'
        elif row in gap_rows:
            cell.value = 42  # hardcoded value
        else:
            cell.value = None  # empty
    return wb

def test_no_gaps():
    wb = create_sheet_with_formula_gaps([1, 2, 3, 4, 5])
    detector = CopyPasteFormulaGapsDetector()
    results = detector.detect(wb)
    assert not results

def test_small_gap():
    wb = create_sheet_with_formula_gaps([1, 2, 4, 5], gap_rows=[3])
    detector = CopyPasteFormulaGapsDetector()
    results = detector.detect(wb)
    assert results
    r = results[0]
    assert r.details['gap_cells'] == [3]
    assert r.probability == 0.8
    assert r.severity == ErrorSeverity.HIGH

def test_large_gap():
    wb = create_sheet_with_formula_gaps([1, 2, 8, 9], gap_rows=[3, 4, 5, 6, 7])
    detector = CopyPasteFormulaGapsDetector()
    results = detector.detect(wb)
    assert results
    r = results[0]
    assert r.details['gap_cells'] == [3, 4, 5, 6, 7]
    assert r.probability == 0.6
    assert r.severity == ErrorSeverity.MEDIUM

def test_dissimilar_formulas():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1).value = "=A1+1"
    ws.cell(row=1, column=1).data_type = 'f'
    ws.cell(row=2, column=1).value = 42  # gap
    ws.cell(row=3, column=1).value = "=B1*2"  # different formula
    ws.cell(row=3, column=1).data_type = 'f'
    detector = CopyPasteFormulaGapsDetector()
    results = detector.detect(wb)
    assert not results

def test_insufficient_formulas():
    wb = create_sheet_with_formula_gaps([1, 3], gap_rows=[2])
    detector = CopyPasteFormulaGapsDetector()
    results = detector.detect(wb)
    assert not results 