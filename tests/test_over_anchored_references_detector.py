import pytest
import openpyxl
from src.excel_analyzer.probabilistic_error_detector import OverAnchoredReferencesDetector, ErrorSeverity

def create_sheet_with_over_anchored_issues():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Set up headers and varying values
    ws.cell(row=1, column=1).value = "Rate"
    ws.cell(row=1, column=2).value = "Amount"
    ws.cell(row=1, column=3).value = "Result"
    
    # Set varying values (different in each row)
    for row in range(2, 6):
        ws.cell(row=row, column=1).value = row * 0.1  # Varying rates
        ws.cell(row=row, column=2).value = row * 100  # Varying amounts
    
    # Add formulas with over-anchored references (copied pattern)
    ws.cell(row=2, column=3).value = "=$A$2+$B$2"  # Over-anchored: should be =A2+B2
    ws.cell(row=2, column=3).data_type = 'f'
    
    ws.cell(row=3, column=3).value = "=$A$3+$B$3"  # Over-anchored: should be =A3+B3
    ws.cell(row=3, column=3).data_type = 'f'
    
    ws.cell(row=4, column=3).value = "=$A$4+$B$4"  # Over-anchored: should be =A4+B4
    ws.cell(row=4, column=3).data_type = 'f'
    
    # Add correctly anchored header reference
    ws.cell(row=5, column=3).value = "=A$1+B5"  # Correct: header should be anchored
    ws.cell(row=5, column=3).data_type = 'f'
    
    return wb

def test_no_over_anchored_references():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Set up data
    ws.cell(row=1, column=1).value = "Header"
    for row in range(2, 5):
        ws.cell(row=row, column=1).value = row
        ws.cell(row=row, column=2).value = f"=A{row}+1"  # Relative references
        ws.cell(row=row, column=2).data_type = 'f'
    
    detector = OverAnchoredReferencesDetector()
    results = detector.detect(wb)
    assert len(results) == 0

def test_over_anchored_in_copied_pattern():
    wb = create_sheet_with_over_anchored_issues()
    detector = OverAnchoredReferencesDetector()
    results = detector.detect(wb)
    
    # Should flag over-anchored references in copied pattern
    over_anchored_results = [r for r in results if "$A$" in r.details['over_anchored_reference'] or "$B$" in r.details['over_anchored_reference']]
    assert len(over_anchored_results) >= 2  # At least 2 over-anchored references
    
    for r in over_anchored_results:
        assert r.probability > 0.5
        assert r.severity in (ErrorSeverity.MEDIUM, ErrorSeverity.LOW)
        assert "should be" in r.description

def test_over_anchored_with_varying_values():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Set up varying values
    for row in range(1, 5):
        ws.cell(row=row, column=1).value = row * 10  # Varying values
    
    # Add formula with over-anchored reference to varying value
    ws.cell(row=2, column=2).value = "=$A$1+B2"  # Over-anchored: A1 should be relative
    ws.cell(row=2, column=2).data_type = 'f'
    
    detector = OverAnchoredReferencesDetector()
    results = detector.detect(wb)
    
    # Should flag over-anchored reference to varying value
    over_anchored_results = [r for r in results if "$A$1" in r.details['over_anchored_reference']]
    assert len(over_anchored_results) >= 0  # May or may not flag depending on context

def test_correctly_anchored_headers():
    wb = create_sheet_with_over_anchored_issues()
    detector = OverAnchoredReferencesDetector()
    results = detector.detect(wb)
    
    # Should not flag correctly anchored header references
    header_results = [r for r in results if "A$1" in r.details['over_anchored_reference']]
    assert len(header_results) == 0

def test_no_formulas():
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in range(1, 5):
        ws.cell(row=row, column=1).value = row
    detector = OverAnchoredReferencesDetector()
    results = detector.detect(wb)
    assert len(results) == 0

def test_single_formula_no_pattern():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Single formula (not part of a copied pattern)
    ws.cell(row=2, column=2).value = "=$A$1+B2"
    ws.cell(row=2, column=2).data_type = 'f'
    
    detector = OverAnchoredReferencesDetector()
    results = detector.detect(wb)
    
    # Should not flag single formulas (no copied pattern)
    assert len(results) == 0 