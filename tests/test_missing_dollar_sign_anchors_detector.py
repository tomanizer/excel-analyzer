import pytest
import openpyxl
from src.excel_analyzer.probabilistic_error_detector import MissingDollarSignAnchorsDetector, ErrorSeverity

def create_sheet_with_formulas():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Set up headers and constants
    ws.cell(row=1, column=1).value = "Rate"
    ws.cell(row=1, column=2).value = "Amount"
    ws.cell(row=1, column=3).value = "Result"
    
    # Set constant values
    for row in range(2, 6):
        ws.cell(row=row, column=1).value = 0.05  # Constant rate
    
    # Set varying amounts
    for row in range(2, 6):
        ws.cell(row=row, column=2).value = row * 100
    
    # Add formulas
    ws.cell(row=2, column=3).value = "=A2*B2"  # Missing anchor for A2 (should be $A$2)
    ws.cell(row=2, column=3).data_type = 'f'
    
    ws.cell(row=3, column=3).value = "=A3*B3"  # Missing anchor for A3 (should be $A$3)
    ws.cell(row=3, column=3).data_type = 'f'
    
    ws.cell(row=4, column=3).value = "=$A$4*B4"  # Properly anchored
    ws.cell(row=4, column=3).data_type = 'f'
    
    ws.cell(row=5, column=3).value = "=A5+B5"  # Relative references (should stay relative)
    ws.cell(row=5, column=3).data_type = 'f'
    
    # Set A5 to a unique value (not constant)
    ws.cell(row=5, column=1).value = 999  # Unique value
    
    return wb

def test_properly_anchored_formula():
    wb = create_sheet_with_formulas()
    detector = MissingDollarSignAnchorsDetector()
    results = detector.detect(wb)
    # Should not flag the properly anchored formula in row 4
    anchored_results = [r for r in results if "A4" in r.details['reference']]
    assert len(anchored_results) == 0

def test_missing_anchor_for_constant():
    wb = create_sheet_with_formulas()
    detector = MissingDollarSignAnchorsDetector()
    results = detector.detect(wb)
    # Should flag missing anchors for A2 and A3 (constant values)
    constant_results = [r for r in results if r.details['reference'] in ['A2', 'A3']]
    assert len(constant_results) >= 1
    for r in constant_results:
        assert r.probability > 0.5
        assert r.severity in (ErrorSeverity.HIGH, ErrorSeverity.MEDIUM)

def test_relative_reference_stays_relative():
    wb = create_sheet_with_formulas()
    detector = MissingDollarSignAnchorsDetector()
    results = detector.detect(wb)
    # Should not flag A5 and B5 as they are different values (not constants)
    relative_results = [r for r in results if r.details['reference'] in ['A5', 'B5']]
    assert len(relative_results) == 0

def test_header_reference():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1).value = "Header"
    ws.cell(row=2, column=2).value = "=A1"  # Missing anchor for header
    ws.cell(row=2, column=2).data_type = 'f'
    detector = MissingDollarSignAnchorsDetector()
    results = detector.detect(wb)
    assert len(results) == 1
    assert results[0].details['reference'] == 'A1'
    assert results[0].probability == 0.9
    assert results[0].severity == ErrorSeverity.HIGH

def test_no_formulas():
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in range(1, 5):
        ws.cell(row=row, column=1).value = row
    detector = MissingDollarSignAnchorsDetector()
    results = detector.detect(wb)
    assert len(results) == 0 