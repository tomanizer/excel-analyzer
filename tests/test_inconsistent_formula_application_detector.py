import pytest
import openpyxl
from src.excel_analyzer.probabilistic_error_detector import InconsistentFormulaApplicationDetector, ErrorSeverity

def create_sheet_with_mixed_content(formula_rows, hardcoded_rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in range(1, 21):
        cell = ws.cell(row=row, column=1)
        if row in formula_rows:
            cell.value = f"=A{row}+1"
            cell.data_type = 'f'
        elif row in hardcoded_rows:
            cell.value = 42
        else:
            cell.value = None  # empty
    return wb

def test_all_formulas():
    wb = create_sheet_with_mixed_content([1, 2, 3, 4, 5], [])
    detector = InconsistentFormulaApplicationDetector()
    results = detector.detect(wb)
    assert not results

def test_all_hardcoded():
    wb = create_sheet_with_mixed_content([], [1, 2, 3, 4, 5])
    detector = InconsistentFormulaApplicationDetector()
    results = detector.detect(wb)
    assert not results

def test_balanced_mix():
    wb = create_sheet_with_mixed_content([1, 2, 3], [4, 5, 6])
    detector = InconsistentFormulaApplicationDetector()
    results = detector.detect(wb)
    assert results
    r = results[0]
    assert r.probability == 0.9
    assert r.severity == ErrorSeverity.HIGH
    assert r.details['formula_ratio'] == 0.5
    assert r.details['hardcoded_ratio'] == 0.5

def test_mostly_formulas():
    wb = create_sheet_with_mixed_content([1, 2, 3, 4, 5, 6, 7], [8, 9])
    detector = InconsistentFormulaApplicationDetector()
    results = detector.detect(wb)
    assert results
    r = results[0]
    assert r.probability == 0.5
    assert r.severity == ErrorSeverity.MEDIUM
    assert r.details['formula_ratio'] > 0.7
    assert r.details['hardcoded_ratio'] < 0.3

def test_mostly_hardcoded():
    wb = create_sheet_with_mixed_content([1, 2], [3, 4, 5, 6, 7, 8, 9])
    detector = InconsistentFormulaApplicationDetector()
    results = detector.detect(wb)
    assert results
    r = results[0]
    assert r.probability == 0.5
    assert r.severity == ErrorSeverity.MEDIUM
    assert r.details['formula_ratio'] < 0.3
    assert r.details['hardcoded_ratio'] > 0.7

def test_small_range():
    wb = create_sheet_with_mixed_content([1], [2])
    detector = InconsistentFormulaApplicationDetector()
    results = detector.detect(wb)
    assert not results

def test_insufficient_mix():
    wb = create_sheet_with_mixed_content([1, 2, 3, 4, 5, 6, 7, 8], [9])
    detector = InconsistentFormulaApplicationDetector()
    results = detector.detect(wb)
    assert not results  # Only 11% hardcoded, below 20% threshold 