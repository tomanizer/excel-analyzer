#!/usr/bin/env python3
"""
Tests for Conditional Formatting Overlap Conflicts Detector.

Covers:
- Non-overlapping conditional formatting rules (should not flag)
- Overlapping rules with different formats (should flag, high probability)
- Overlapping rules of different types (should flag, medium probability)
- Overlapping rules with compatible formats (should flag, low probability)
"""

import pytest
import tempfile
import shutil
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule, FormulaRule

from excel_analyzer.probabilistic_error_detector import ConditionalFormattingOverlapConflictsDetector

class TestConditionalFormattingOverlapConflictsDetector:
    def setup_method(self):
        self.detector = ConditionalFormattingOverlapConflictsDetector()
        self.temp_dir = Path(tempfile.mkdtemp())
    def teardown_method(self):
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)
    def create_test_workbook(self, setup_func):
        wb = openpyxl.Workbook()
        ws = wb.active
        setup_func(ws)
        return wb
    def test_non_overlapping_rules(self):
        # Two rules on non-overlapping ranges
        def setup(ws):
            ws['A1'].value = 1
            ws['B1'].value = 2
            ws.conditional_formatting.add('A1', CellIsRule(operator='equal', formula=['1'], fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')))
            ws.conditional_formatting.add('B1', CellIsRule(operator='equal', formula=['2'], fill=PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')))
        wb = self.create_test_workbook(setup)
        results = self.detector.detect(wb)
        assert not any(r.error_type == 'conditional_formatting_overlap_conflicts' for r in results)
    def test_overlapping_rules_different_formats(self):
        # Two rules on overlapping range, different fill colors
        def setup(ws):
            ws['A1'].value = 1
            ws['A2'].value = 2
            ws.conditional_formatting.add('A1:A2', CellIsRule(operator='equal', formula=['1'], fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')))
            ws.conditional_formatting.add('A2', CellIsRule(operator='equal', formula=['2'], fill=PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')))
        wb = self.create_test_workbook(setup)
        # Debug: print extracted rules
        sheet = wb.active
        detector = self.detector
        cf_rules = detector._extract_conditional_formatting_rules(sheet)
        print('Extracted rules:')
        for rule in cf_rules:
            print(rule)
        results = self.detector.detect(wb)
        assert any(r.error_type == 'conditional_formatting_overlap_conflicts' for r in results)
        for r in results:
            assert r.probability >= 0.9
    def test_overlapping_rules_different_types(self):
        # Overlapping rules of different types (formula vs cell is)
        def setup(ws):
            ws['A1'].value = 1
            ws['A2'].value = 2
            ws.conditional_formatting.add('A1:A2', CellIsRule(operator='equal', formula=['1'], fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')))
            ws.conditional_formatting.add('A1:A2', FormulaRule(formula=['A1=2'], fill=PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')))
        wb = self.create_test_workbook(setup)
        results = self.detector.detect(wb)
        assert any(r.error_type == 'conditional_formatting_overlap_conflicts' for r in results)
        for r in results:
            assert r.probability >= 0.6
    def test_overlapping_rules_compatible_formats(self):
        # Overlapping rules, both set font but same font
        def setup(ws):
            ws['A1'].value = 1
            ws['A2'].value = 2
            font = Font(bold=True)
            ws.conditional_formatting.add('A1:A2', CellIsRule(operator='equal', formula=['1'], font=font))
            ws.conditional_formatting.add('A1:A2', CellIsRule(operator='equal', formula=['2'], font=font))
        wb = self.create_test_workbook(setup)
        results = self.detector.detect(wb)
        assert any(r.error_type == 'conditional_formatting_overlap_conflicts' for r in results)
        for r in results:
            assert r.probability <= 0.3

if __name__ == '__main__':
    pytest.main([__file__]) 