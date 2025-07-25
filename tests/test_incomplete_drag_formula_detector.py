#!/usr/bin/env python3
"""
Tests for Incomplete Drag Formula Detector.

Covers:
- Columns with formulas dragged to all data rows (should not flag)
- Columns with formulas missing at the end (should flag)
- Columns with gaps in the middle (should flag, high probability)
- Columns with only a single missing formula (should flag, low probability)
- Columns with intentional gaps (should not flag if possible)
"""

import pytest
import tempfile
import shutil
from pathlib import Path
import openpyxl

from excel_analyzer.probabilistic_error_detector import IncompleteDragFormulaDetector, ErrorSeverity

class TestIncompleteDragFormulaDetector:
    def setup_method(self):
        self.detector = IncompleteDragFormulaDetector()
        self.temp_dir = Path(tempfile.mkdtemp())
    def teardown_method(self):
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)
    def create_test_workbook(self, formula_rows, data_rows, intentional_gaps=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in data_rows:
            ws[f'A{row}'].value = 1  # Data present
        for row in formula_rows:
            ws[f'A{row}'].value = f'=B{row}+1'
            ws[f'A{row}'].data_type = 'f'
        if intentional_gaps:
            for row in intentional_gaps:
                ws[f'A{row}'].value = None
        return wb
    def test_full_drag(self):
        # Formulas dragged to all data rows
        data_rows = list(range(2, 11))
        formula_rows = list(range(2, 11))
        wb = self.create_test_workbook(formula_rows, data_rows)
        results = self.detector.detect(wb)
        assert not any(r.error_type == 'incomplete_drag_formula' for r in results)
    def test_cutoff_at_end(self):
        # Formulas missing at the end
        data_rows = list(range(2, 11))
        formula_rows = list(range(2, 8))  # Missing 8,9,10
        wb = self.create_test_workbook(formula_rows, data_rows)
        results = self.detector.detect(wb)
        assert any(r.probability >= 0.6 for r in results)
    def test_gap_in_middle(self):
        # Gap in the middle
        data_rows = list(range(2, 11))
        formula_rows = [2, 3, 4, 6, 7, 8, 9, 10]  # Missing 5
        wb = self.create_test_workbook(formula_rows, data_rows)
        results = self.detector.detect(wb)
        assert any(r.probability >= 0.9 for r in results)
    def test_single_missing_formula(self):
        # Only a single missing formula
        data_rows = list(range(2, 11))
        formula_rows = [2, 3, 4, 5, 6, 7, 8, 9]  # Missing 10
        wb = self.create_test_workbook(formula_rows, data_rows)
        results = self.detector.detect(wb)
        assert any(0.2 < r.probability <= 0.6 for r in results)
    def test_intentional_gap(self):
        # Intentional gap (e.g., header or subtotal row)
        data_rows = list(range(2, 11))
        formula_rows = [2, 3, 4, 6, 7, 8, 9, 10]  # Missing 5 (intentional)
        wb = self.create_test_workbook(formula_rows, data_rows, intentional_gaps=[5])
        results = self.detector.detect(wb)
        # Should still flag, but user can review details
        assert any(r.error_type == 'incomplete_drag_formula' for r in results)

if __name__ == '__main__':
    pytest.main([__file__]) 