#!/usr/bin/env python3
"""
Tests for Volatile Functions Detector.

Covers:
- Workbooks with no volatile functions
- Workbooks with many volatile functions
- Workbooks with high-impact volatile functions (many dependencies)
- Different types of volatile functions
- Large models vs small models
"""

import pytest
import tempfile
import shutil
from pathlib import Path
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from excel_analyzer.probabilistic_error_detector import VolatileFunctionsDetector, ErrorSeverity

class TestVolatileFunctionsDetector:
    def setup_method(self):
        self.detector = VolatileFunctionsDetector()
        self.temp_dir = Path(tempfile.mkdtemp())
    def teardown_method(self):
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)
    def create_test_workbook(self, setup_func):
        wb = openpyxl.Workbook()
        ws = wb.active
        setup_func(ws)
        return wb
    def test_no_volatile_functions(self):
        # Workbook with no volatile functions
        def setup(ws: Worksheet):
            ws['A1'].value = '=SUM(A2:A10)'
            ws['A1'].data_type = 'f'
            ws['B1'].value = '=AVERAGE(B2:B10)'
            ws['B1'].data_type = 'f'
        wb = self.create_test_workbook(setup)
        results = self.detector.detect(wb)
        assert not any(r.error_type == 'volatile_functions' for r in results)
    def test_few_volatile_functions(self):
        # Workbook with few volatile functions
        def setup(ws: Worksheet):
            ws['A1'].value = '=NOW()'
            ws['A1'].data_type = 'f'
            ws['B1'].value = '=SUM(B2:B10)'
            ws['B1'].data_type = 'f'
        wb = self.create_test_workbook(setup)
        results = self.detector.detect(wb)
        assert any(r.error_type == 'volatile_functions' for r in results)
        for r in results:
            assert r.probability < 0.5  # Low probability for few functions
    def test_many_volatile_functions(self):
        # Workbook with many volatile functions
        def setup(ws: Worksheet):
            for i in range(10):
                ws[f'A{i+1}'].value = '=NOW()'
                ws[f'A{i+1}'].data_type = 'f'
                ws[f'B{i+1}'].value = '=RAND()'
                ws[f'B{i+1}'].data_type = 'f'
                ws[f'C{i+1}'].value = '=OFFSET(A1,0,0)'
                ws[f'C{i+1}'].data_type = 'f'
        wb = self.create_test_workbook(setup)
        results = self.detector.detect(wb)
        assert any(r.error_type == 'volatile_functions' for r in results)
        for r in results:
            assert r.probability >= 0.7  # High probability for many functions
            assert r.details['total_volatile_functions'] >= 30
    def test_high_impact_volatile_functions(self):
        # Volatile function with many dependencies
        def setup(ws: Worksheet):
            # Create a volatile function that many other cells depend on
            ws['A1'].value = '=NOW()'
            ws['A1'].data_type = 'f'
            # Create many cells that reference A1
            for i in range(10):
                ws[f'B{i+1}'].value = f'=A1+{i}'
                ws[f'B{i+1}'].data_type = 'f'
                ws[f'C{i+1}'].value = f'=A1*{i}'
                ws[f'C{i+1}'].data_type = 'f'
        wb = self.create_test_workbook(setup)
        results = self.detector.detect(wb)
        assert any(r.error_type == 'volatile_functions' for r in results)
        for r in results:
            assert r.probability >= 0.8  # Very high probability due to high impact
            assert r.details['high_impact_cells'] >= 1
    def test_different_volatile_function_types(self):
        # Test different types of volatile functions
        def setup(ws: Worksheet):
            ws['A1'].value = '=TODAY()'
            ws['A1'].data_type = 'f'
            ws['A2'].value = '=RANDBETWEEN(1,100)'
            ws['A2'].data_type = 'f'
            ws['A3'].value = '=INDIRECT("A1")'
            ws['A3'].data_type = 'f'
            ws['A4'].value = '=CELL("address",A1)'
            ws['A4'].data_type = 'f'
        wb = self.create_test_workbook(setup)
        results = self.detector.detect(wb)
        assert any(r.error_type == 'volatile_functions' for r in results)
        for r in results:
            assert r.details['total_volatile_functions'] >= 4
    def test_large_model_with_volatile_functions(self):
        # Large model with volatile functions (more sensitive)
        def setup(ws: Worksheet):
            # Create many formulas (large model)
            for i in range(100):
                ws[f'A{i+1}'].value = f'=SUM(A{i+2}:A{i+11})'
                ws[f'A{i+1}'].data_type = 'f'
            # Add a few volatile functions
            ws['B1'].value = '=NOW()'
            ws['B1'].data_type = 'f'
            ws['B2'].value = '=RAND()'
            ws['B2'].data_type = 'f'
        wb = self.create_test_workbook(setup)
        results = self.detector.detect(wb)
        assert any(r.error_type == 'volatile_functions' for r in results)
        for r in results:
            # Should have higher probability due to large model size
            assert r.probability >= 0.6
            assert r.details['total_formulas'] >= 100
    def test_volatile_functions_in_named_ranges(self):
        # Test volatile functions in named ranges (if detectable)
        def setup(ws: Worksheet):
            ws['A1'].value = '=NOW()'
            ws['A1'].data_type = 'f'
            # Create a named range reference (simplified test)
            ws['B1'].value = '=A1'
            ws['B1'].data_type = 'f'
        wb = self.create_test_workbook(setup)
        results = self.detector.detect(wb)
        assert any(r.error_type == 'volatile_functions' for r in results)

if __name__ == '__main__':
    pytest.main([__file__]) 