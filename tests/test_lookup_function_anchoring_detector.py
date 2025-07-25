import pytest
import openpyxl
from src.excel_analyzer.probabilistic_error_detector import LookupFunctionAnchoringDetector, ErrorSeverity

def create_sheet_with_lookup_errors():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Set up lookup data
    for row in range(1, 6):
        ws.cell(row=row, column=1).value = f"Key{row}"  # Lookup keys
        ws.cell(row=row, column=2).value = row * 10     # Values
        ws.cell(row=row, column=3).value = row * 100    # More values
    
    # Add VLOOKUP with wrong anchoring (copying across)
    ws.cell(row=1, column=4).value = "=VLOOKUP(A1,$B$1:$C$5,2)"  # Wrong: A1 should be $A1
    ws.cell(row=1, column=4).data_type = 'f'
    
    ws.cell(row=1, column=5).value = "=VLOOKUP($A1,B1:C5,2)"  # Wrong: B1:C5 should be $B$1:$C$5
    ws.cell(row=1, column=5).data_type = 'f'
    
    # Add HLOOKUP with wrong anchoring (copying down)
    ws.cell(row=2, column=4).value = "=HLOOKUP(A1,$B$1:$C$5,2)"  # Wrong: A1 should be A$1
    ws.cell(row=2, column=4).data_type = 'f'
    
    # Add INDEX/MATCH with wrong anchoring
    ws.cell(row=3, column=4).value = "=INDEX($A$1:$A$5,MATCH(A1,$B$1:$B$5,0))"  # Wrong: A1 should be $A1
    ws.cell(row=3, column=4).data_type = 'f'
    
    # Add correct VLOOKUP
    ws.cell(row=4, column=4).value = "=VLOOKUP($A1,$B$1:$C$5,2)"  # Correct
    ws.cell(row=4, column=4).data_type = 'f'
    
    return wb

def test_correct_vlookup_anchoring():
    wb = create_sheet_with_lookup_errors()
    detector = LookupFunctionAnchoringDetector()
    results = detector.detect(wb)
    
    # Should not flag correctly anchored VLOOKUP
    correct_results = [r for r in results if "VLOOKUP" in r.details['formula'] and "$A1" in r.details['formula'] and "$B$1:$C$5" in r.details['formula']]
    assert len(correct_results) == 0

def test_vlookup_wrong_lookup_value_anchoring():
    wb = create_sheet_with_lookup_errors()
    detector = LookupFunctionAnchoringDetector()
    results = detector.detect(wb)
    
    # Should flag VLOOKUP with wrong lookup value anchoring
    lookup_value_errors = [r for r in results if r.details['function_type'] == 'VLOOKUP' and r.details['parameter'] == 'lookup_value']
    assert len(lookup_value_errors) >= 1
    
    for r in lookup_value_errors:
        assert r.probability > 0.7  # High probability for lookup functions
        assert r.severity == ErrorSeverity.HIGH
        assert "should be column-locked" in r.description

def test_vlookup_wrong_table_array_anchoring():
    wb = create_sheet_with_lookup_errors()
    detector = LookupFunctionAnchoringDetector()
    results = detector.detect(wb)
    
    # Should flag VLOOKUP with wrong table array anchoring
    table_array_errors = [r for r in results if r.details['function_type'] == 'VLOOKUP' and r.details['parameter'] == 'table_array']
    assert len(table_array_errors) >= 1
    
    for r in table_array_errors:
        assert r.probability > 0.8  # Very high probability for table arrays
        assert r.severity == ErrorSeverity.HIGH
        assert "should be fully locked" in r.description

def test_hlookup_anchoring_errors():
    wb = create_sheet_with_lookup_errors()
    detector = LookupFunctionAnchoringDetector()
    results = detector.detect(wb)
    
    # Should flag HLOOKUP anchoring errors
    hlookup_errors = [r for r in results if r.details['function_type'] == 'HLOOKUP']
    assert len(hlookup_errors) >= 1
    
    for r in hlookup_errors:
        assert r.probability > 0.7
        assert r.severity == ErrorSeverity.HIGH

def test_index_match_anchoring_errors():
    wb = create_sheet_with_lookup_errors()
    detector = LookupFunctionAnchoringDetector()
    results = detector.detect(wb)
    
    # Should flag INDEX/MATCH anchoring errors
    index_match_errors = [r for r in results if r.details['function_type'] == 'INDEX/MATCH']
    assert len(index_match_errors) >= 1
    
    for r in index_match_errors:
        assert r.probability > 0.7
        assert r.severity == ErrorSeverity.HIGH

def test_no_lookup_functions():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Add non-lookup formulas
    ws.cell(row=1, column=2).value = "=A1+B1"
    ws.cell(row=1, column=2).data_type = 'f'
    
    ws.cell(row=2, column=2).value = "=SUM(A1:A10)"
    ws.cell(row=2, column=2).data_type = 'f'
    
    detector = LookupFunctionAnchoringDetector()
    results = detector.detect(wb)
    assert len(results) == 0

def test_copy_direction_detection():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Set up data
    for row in range(1, 4):
        ws.cell(row=row, column=1).value = f"Key{row}"
        ws.cell(row=row, column=2).value = row * 10
    
    # Add VLOOKUP formulas in a row (copying across)
    ws.cell(row=1, column=3).value = "=VLOOKUP(A1,$B$1:$B$3,1)"  # Wrong: A1 should be $A1
    ws.cell(row=1, column=3).data_type = 'f'
    
    ws.cell(row=1, column=4).value = "=VLOOKUP(A1,$B$1:$B$3,1)"  # Wrong: A1 should be $A1
    ws.cell(row=1, column=4).data_type = 'f'
    
    detector = LookupFunctionAnchoringDetector()
    results = detector.detect(wb)
    
    # Should detect copying across and flag lookup value anchoring
    across_errors = [r for r in results if r.details['copy_direction'] == 'across']
    assert len(across_errors) >= 1

def test_critical_parameters():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Set up data
    for row in range(1, 4):
        ws.cell(row=row, column=1).value = f"Key{row}"
        ws.cell(row=row, column=2).value = row * 10
    
    # Add VLOOKUP with wrong table array (critical parameter)
    ws.cell(row=1, column=3).value = "=VLOOKUP($A1,B1:B3,1)"  # Wrong: B1:B3 should be $B$1:$B$3
    ws.cell(row=1, column=3).data_type = 'f'
    
    detector = LookupFunctionAnchoringDetector()
    results = detector.detect(wb)
    
    # Should flag table array error with high probability
    table_array_errors = [r for r in results if r.details['parameter'] == 'table_array']
    assert len(table_array_errors) == 1
    assert table_array_errors[0].probability > 0.8  # Very high for critical parameters 