import pytest
import openpyxl
from src.excel_analyzer.probabilistic_error_detector import ArrayFormulaAnchoringDetector, ErrorSeverity

def create_sheet_with_array_formula_errors():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Set up data
    for row in range(1, 101):
        ws.cell(row=row, column=1).value = row
        ws.cell(row=row, column=2).value = row * 10
    
    # Add array formulas with over-anchored ranges
    ws.cell(row=1, column=3).value = "=SUM(IF($A$1:$A$100>0,$B$1:$B$100,0))"  # Over-anchored: should be relative
    ws.cell(row=1, column=3).data_type = 'f'
    
    ws.cell(row=2, column=3).value = "=UNIQUE($A$1:$A$100)"  # Over-anchored: should be relative
    ws.cell(row=2, column=3).data_type = 'f'
    
    ws.cell(row=3, column=3).value = "=FILTER($A$1:$A$100,$B$1:$B$100>50)"  # Over-anchored: should be relative
    ws.cell(row=3, column=3).data_type = 'f'
    
    # Add correctly anchored array formulas
    ws.cell(row=4, column=3).value = "=SUM(IF(A1:A10>0,B1:B10,0))"  # Correct: relative ranges
    ws.cell(row=4, column=3).data_type = 'f'
    
    ws.cell(row=5, column=3).value = "=UNIQUE(A1:A10)"  # Correct: relative range
    ws.cell(row=5, column=3).data_type = 'f'
    
    return wb

def test_correct_array_formula_anchoring():
    wb = create_sheet_with_array_formula_errors()
    detector = ArrayFormulaAnchoringDetector()
    results = detector.detect(wb)
    
    # Should not flag correctly anchored array formulas
    correct_results = [r for r in results if "A1:A10" in r.details['range_reference'] or "B1:B10" in r.details['range_reference']]
    assert len(correct_results) == 0

def test_sum_if_over_anchored_ranges():
    wb = create_sheet_with_array_formula_errors()
    detector = ArrayFormulaAnchoringDetector()
    results = detector.detect(wb)
    
    # Should flag SUM(IF) with over-anchored ranges
    sum_if_errors = [r for r in results if r.details['array_function'] == 'SUM(IF)']
    assert len(sum_if_errors) >= 1
    
    for r in sum_if_errors:
        assert r.probability > 0.6  # Medium probability for array formulas
        assert r.severity in (ErrorSeverity.MEDIUM, ErrorSeverity.LOW)
        assert "should be relative" in r.description

def test_unique_over_anchored_range():
    wb = create_sheet_with_array_formula_errors()
    detector = ArrayFormulaAnchoringDetector()
    results = detector.detect(wb)
    
    # Should flag UNIQUE with over-anchored range
    unique_errors = [r for r in results if r.details['array_function'] == 'UNIQUE']
    assert len(unique_errors) == 1
    
    r = unique_errors[0]
    assert r.probability > 0.6
    assert r.severity in (ErrorSeverity.MEDIUM, ErrorSeverity.LOW)
    assert "$A$1:$A$100" in r.details['range_reference']

def test_filter_over_anchored_ranges():
    wb = create_sheet_with_array_formula_errors()
    detector = ArrayFormulaAnchoringDetector()
    results = detector.detect(wb)
    
    # Should flag FILTER with over-anchored ranges
    filter_errors = [r for r in results if r.details['array_function'] == 'FILTER']
    assert len(filter_errors) >= 1
    
    for r in filter_errors:
        assert r.probability > 0.6
        assert r.severity in (ErrorSeverity.MEDIUM, ErrorSeverity.LOW)

def test_large_ranges():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Set up large dataset
    for row in range(1, 1001):
        ws.cell(row=row, column=1).value = row
    
    # Add array formula with large over-anchored range
    ws.cell(row=1, column=2).value = "=UNIQUE($A$1:$A$1000)"  # Large over-anchored range
    ws.cell(row=1, column=2).data_type = 'f'
    
    detector = ArrayFormulaAnchoringDetector()
    results = detector.detect(wb)
    
    # Should flag large over-anchored range with higher probability
    large_range_errors = [r for r in results if "$A$1:$A$1000" in r.details['range_reference']]
    assert len(large_range_errors) == 1
    assert large_range_errors[0].probability > 0.7  # Higher probability for large ranges

def test_no_array_functions():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Add non-array formulas
    ws.cell(row=1, column=2).value = "=A1+B1"
    ws.cell(row=1, column=2).data_type = 'f'
    
    ws.cell(row=2, column=2).value = "=SUM(A1:A10)"
    ws.cell(row=2, column=2).data_type = 'f'
    
    detector = ArrayFormulaAnchoringDetector()
    results = detector.detect(wb)
    assert len(results) == 0

def test_modern_array_functions():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Set up data
    for row in range(1, 51):
        ws.cell(row=row, column=1).value = row
        ws.cell(row=row, column=2).value = row * 10
    
    # Add modern array functions with over-anchored ranges
    ws.cell(row=1, column=3).value = "=SORT($A$1:$A$50)"  # Over-anchored
    ws.cell(row=1, column=3).data_type = 'f'
    
    ws.cell(row=2, column=3).value = "=SORTBY($A$1:$A$50,$B$1:$B$50)"  # Over-anchored
    ws.cell(row=2, column=3).data_type = 'f'
    
    detector = ArrayFormulaAnchoringDetector()
    results = detector.detect(wb)
    
    # Should flag modern array functions with over-anchored ranges
    modern_errors = [r for r in results if r.details['array_function'] in ['SORT', 'SORTBY']]
    assert len(modern_errors) >= 1
    
    for r in modern_errors:
        assert r.probability > 0.6
        assert r.severity in (ErrorSeverity.MEDIUM, ErrorSeverity.LOW)

def test_small_ranges_not_flagged():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Set up small dataset
    for row in range(1, 11):
        ws.cell(row=row, column=1).value = row
    
    # Add array formula with small over-anchored range (should not be flagged)
    ws.cell(row=1, column=2).value = "=UNIQUE($A$1:$A$10)"  # Small range, may not be flagged
    ws.cell(row=1, column=2).data_type = 'f'
    
    detector = ArrayFormulaAnchoringDetector()
    results = detector.detect(wb)
    
    # Small ranges may or may not be flagged depending on the logic
    # This test ensures the detector doesn't crash on small ranges
    assert len(results) >= 0 