#!/usr/bin/env python3
"""
Tests for Precision Errors in Financial Calculations Detector.

Covers:
- Formulas with explicit rounding (should not flag)
- Formulas with decimal arithmetic and no rounding (should flag)
- Formulas with chained arithmetic (should flag, high probability)
- Formulas with subtraction of nearly equal numbers (should flag, high probability)
- Integer-only calculations (should not flag)
"""

import pytest
import tempfile
import shutil
from pathlib import Path
import openpyxl

from excel_analyzer.probabilistic_error_detector import PrecisionErrorsInFinancialCalculationsDetector, ErrorSeverity

class TestPrecisionErrorsInFinancialCalculationsDetector:
    def setup_method(self):
        self.detector = PrecisionErrorsInFinancialCalculationsDetector()
        self.temp_dir = Path(tempfile.mkdtemp())
    def teardown_method(self):
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)
    def create_test_workbook(self, formula):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'].value = 1.234
        ws['A2'].value = 1.233
        ws['A3'].value = 100
        ws['A4'].value = 3
        ws['B1'].value = formula
        ws['B1'].data_type = 'f'
        return wb
    def test_explicit_rounding(self):
        # Formula with explicit rounding
        formula = '=ROUND(A1/A3, 2)'
        wb = self.create_test_workbook(formula)
        results = self.detector.detect(wb)
        assert not any(r.error_type == 'precision_errors_in_financial_calculations' for r in results)
    def test_decimal_arithmetic_no_rounding(self):
        # Formula with decimal arithmetic and no rounding
        formula = '=A1/A3'
        wb = self.create_test_workbook(formula)
        results = self.detector.detect(wb)
        assert any(r.error_type == 'precision_errors_in_financial_calculations' for r in results)
        for r in results:
            assert r.probability >= 0.6
    def test_chained_arithmetic(self):
        # Formula with chained arithmetic and no rounding
        formula = '=A1/A3+A2/A4-A1*A2/A3'
        wb = self.create_test_workbook(formula)
        results = self.detector.detect(wb)
        assert any(r.error_type == 'precision_errors_in_financial_calculations' for r in results)
        for r in results:
            assert r.probability >= 0.8
    def test_subtraction_nearly_equal(self):
        # Formula with subtraction of nearly equal numbers
        formula = '=A1-A2'
        wb = self.create_test_workbook(formula)
        results = self.detector.detect(wb)
        assert any(r.error_type == 'precision_errors_in_financial_calculations' for r in results)
        for r in results:
            assert r.probability >= 0.8
    def test_integer_only_calculation(self):
        # Integer-only calculation (should not flag)
        formula = '=A3-A4'
        wb = self.create_test_workbook(formula)
        results = self.detector.detect(wb)
        assert not any(r.error_type == 'precision_errors_in_financial_calculations' for r in results)

if __name__ == '__main__':
    pytest.main([__file__]) 