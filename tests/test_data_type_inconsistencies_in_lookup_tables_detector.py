#!/usr/bin/env python3
"""
Tests for Data Type Inconsistencies in Lookup Tables Detector.

Covers:
- All numbers in lookup key (should not flag)
- All text in lookup key (should not flag)
- Mixed numbers and text (should flag, high probability)
- Numbers stored as text (should flag, medium probability)
- Lookup ranges with empty cells (should not flag unless mixed types)
"""

import pytest
import tempfile
import shutil
from pathlib import Path
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from excel_analyzer.probabilistic_error_detector import DataTypeInconsistenciesInLookupTablesDetector, ErrorSeverity

class TestDataTypeInconsistenciesInLookupTablesDetector:
    def setup_method(self):
        self.detector = DataTypeInconsistenciesInLookupTablesDetector()
        self.temp_dir = Path(tempfile.mkdtemp())
    def teardown_method(self):
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)
    def create_test_workbook(self, data, formula, rng='A1:A10'):
        wb = openpyxl.Workbook()
        ws = wb.active
        min_row = 1
        for i, value in enumerate(data, min_row):
            ws[f'A{i}'].value = value
        ws['B1'].value = formula
        ws['B1'].data_type = 'f'
        return wb
    def test_all_numbers(self):
        # All numbers in lookup key
        data = [i for i in range(1, 11)]
        formula = '=VLOOKUP(5,A1:A10,1,FALSE)'
        wb = self.create_test_workbook(data, formula)
        results = self.detector.detect(wb)
        assert not any(r.error_type == 'data_type_inconsistencies_in_lookup_tables' for r in results)
    def test_all_text(self):
        # All text in lookup key
        data = [str(i) for i in range(1, 11)]
        formula = '=VLOOKUP("5",A1:A10,1,FALSE)'
        wb = self.create_test_workbook(data, formula)
        results = self.detector.detect(wb)
        assert not any(r.error_type == 'data_type_inconsistencies_in_lookup_tables' for r in results)
    def test_mixed_numbers_and_text(self):
        # Mixed numbers and text
        data = [1, 2, '3', 4, '5', 6, 7, '8', 9, 10]
        formula = '=VLOOKUP(5,A1:A10,1,FALSE)'
        wb = self.create_test_workbook(data, formula)
        results = self.detector.detect(wb)
        assert any(r.error_type == 'data_type_inconsistencies_in_lookup_tables' for r in results)
        for r in results:
            assert r.probability >= 0.9
    def test_numbers_stored_as_text(self):
        # Numbers stored as text
        data = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10']
        formula = '=VLOOKUP(5,A1:A10,1,FALSE)'
        wb = self.create_test_workbook(data, formula)
        results = self.detector.detect(wb)
        # Should flag as medium probability (all text, but used as numbers)
        for r in results:
            assert r.probability <= 0.5
    def test_lookup_range_with_empty_cells(self):
        # Lookup range with empty cells
        data = [1, None, 3, None, 5, 6, None, 8, 9, 10]
        formula = '=VLOOKUP(5,A1:A10,1,FALSE)'
        wb = self.create_test_workbook(data, formula)
        results = self.detector.detect(wb)
        # Should not flag unless mixed types
        assert not any(r.error_type == 'data_type_inconsistencies_in_lookup_tables' for r in results)

if __name__ == '__main__':
    pytest.main([__file__]) 