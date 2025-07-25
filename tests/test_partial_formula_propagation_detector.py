import pytest
import openpyxl
from openpyxl.utils import get_column_letter
from src.excel_analyzer.probabilistic_error_detector import PartialFormulaPropagationDetector, ErrorSeverity


def create_sheet_with_partial_formulas(rows, missing_rows=None, edge_missing=False):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    missing_rows = set(missing_rows or [])
    for row in range(1, rows + 1):
        cell = ws.cell(row=row, column=1)
        if edge_missing and (row == 1 or row == rows):
            continue
        if row not in missing_rows:
            cell.value = f"=A{row}+1"
            cell.data_type = 'f'
        else:
            cell.value = 42  # hardcoded value
    return wb

def test_all_formulas():
    wb = create_sheet_with_partial_formulas(20)
    detector = PartialFormulaPropagationDetector()
    results = detector.detect(wb)
    assert not results

def test_one_missing_in_middle():
    wb = create_sheet_with_partial_formulas(20, missing_rows=[10])
    detector = PartialFormulaPropagationDetector()
    results = detector.detect(wb)
    assert results
    assert results[0].probability >= 0.7
    assert results[0].details['row'] == 10
    assert results[0].severity == ErrorSeverity.HIGH

def test_one_missing_at_edge():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Set row 1 to a hardcoded value (missing formula), rest have formulas
    ws.cell(row=1, column=1).value = 42
    for row in range(2, 21):
        cell = ws.cell(row=row, column=1)
        cell.value = f"=A{row}+1"
        cell.data_type = 'f'
    detector = PartialFormulaPropagationDetector()
    results = detector.detect(wb)
    assert results
    assert results[0].probability == 0.5
    assert results[0].details['row'] == 1
    assert results[0].severity == ErrorSeverity.MEDIUM

def test_multiple_missing():
    wb = create_sheet_with_partial_formulas(20, missing_rows=[5, 10, 15])
    detector = PartialFormulaPropagationDetector()
    results = detector.detect(wb)
    assert len(results) == 3
    for r in results:
        assert r.probability == 0.6
        assert r.severity == ErrorSeverity.MEDIUM

def test_small_range():
    wb = create_sheet_with_partial_formulas(3, missing_rows=[2])
    detector = PartialFormulaPropagationDetector()
    results = detector.detect(wb)
    assert not results

def test_all_non_formulas():
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in range(1, 20):
        ws.cell(row=row, column=1).value = 42
    detector = PartialFormulaPropagationDetector()
    results = detector.detect(wb)
    assert not results 

# New test file for FormulaBoundaryMismatchDetector
import pytest
import openpyxl
from src.excel_analyzer.probabilistic_error_detector import FormulaBoundaryMismatchDetector, ErrorSeverity

def create_sheet_with_sum_formula(data_rows, formula_end, extra_data_rows=0):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Fill data
    for row in range(1, data_rows + 1):
        ws.cell(row=row, column=1).value = row
    # Add extra data if needed
    for row in range(data_rows + 1, data_rows + 1 + extra_data_rows):
        ws.cell(row=row, column=1).value = row
    # Place SUM formula in row 1, column 2
    ws.cell(row=1, column=2).value = f"=SUM(A1:A{formula_end})"
    ws.cell(row=1, column=2).data_type = 'f'
    return wb

def test_formula_covers_all_data():
    wb = create_sheet_with_sum_formula(50, 50)
    detector = FormulaBoundaryMismatchDetector()
    results = detector.detect(wb)
    assert not results

def test_formula_misses_some_data():
    wb = create_sheet_with_sum_formula(100, 50)
    detector = FormulaBoundaryMismatchDetector()
    results = detector.detect(wb)
    assert results
    r = results[0]
    assert r.details['referenced_range'] == 'A1:A50'
    assert r.details['max_data_row'] == 100
    assert r.probability > 0.5
    assert r.severity in (ErrorSeverity.HIGH, ErrorSeverity.MEDIUM)

def test_formula_covers_more_than_data():
    wb = create_sheet_with_sum_formula(50, 100)
    detector = FormulaBoundaryMismatchDetector()
    results = detector.detect(wb)
    assert not results

def test_non_aggregation_formula():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in range(1, 20):
        ws.cell(row=row, column=1).value = row
    ws.cell(row=1, column=2).value = "=A1+A2"
    ws.cell(row=1, column=2).data_type = 'f'
    detector = FormulaBoundaryMismatchDetector()
    results = detector.detect(wb)
    assert not results

def test_multi_column_range():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in range(1, 20):
        ws.cell(row=row, column=1).value = row
        ws.cell(row=row, column=2).value = row
    ws.cell(row=1, column=3).value = "=SUM(A1:B10)"
    ws.cell(row=1, column=3).data_type = 'f'
    detector = FormulaBoundaryMismatchDetector()
    results = detector.detect(wb)
    assert not results 