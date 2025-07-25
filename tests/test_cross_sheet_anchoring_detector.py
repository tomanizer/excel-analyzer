#!/usr/bin/env python3
"""
Test suite for CrossSheetAnchoringDetector.

Tests various cross-sheet anchoring scenarios including:
- Wrong anchoring when copying across
- Wrong anchoring when copying down
- Wrong anchoring for fixed references
- Complex cross-sheet formulas
- Edge cases and error conditions
"""

import pytest
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Add the src directory to the path
import sys
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from excel_analyzer.probabilistic_error_detector import (
    CrossSheetAnchoringDetector, 
    ErrorSeverity,
    ProbabilisticErrorSniffer
)


class TestCrossSheetAnchoringDetector:
    """Test cases for CrossSheetAnchoringDetector."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.detector = CrossSheetAnchoringDetector()
        self.workbook = Workbook()
        
        # Create test sheets
        self.sheet1 = self.workbook.active
        self.sheet1.title = "Sheet1"
        self.sheet2 = self.workbook.create_sheet("Sheet2")
        self.sheet3 = self.workbook.create_sheet("Data")
    
    def teardown_method(self):
        """Clean up test fixtures."""
        self.workbook.close()
    
    def test_detect_wrong_anchoring_when_copying_across(self):
        """Test detection of wrong anchoring when copying across."""
        # Set up data in Sheet2
        self.sheet2['A1'] = "Header"
        self.sheet2['A2'] = 100
        self.sheet2['A3'] = 200
        self.sheet2['A4'] = 300
        
        # Set up formulas in Sheet1 that should be column-locked when copying across
        # Wrong: fully locked when should be column-locked
        self.sheet1['B2'] = "=Sheet2!$A$2"  # Should be =Sheet2!$A2
        self.sheet1['C2'] = "=Sheet2!$A$3"  # Should be =Sheet2!$A3
        self.sheet1['D2'] = "=Sheet2!$A$4"  # Should be =Sheet2!$A4
        
        # Create a pattern to indicate copying across
        self.sheet1['B1'] = "Jan"
        self.sheet1['C1'] = "Feb"
        self.sheet1['D1'] = "Mar"
        
        results = self.detector.detect(self.workbook)
        
        assert len(results) >= 3
        for result in results:
            assert result.error_type == "cross_sheet_anchoring_errors"
            assert "Wrong cross-sheet anchoring" in result.description
            assert result.probability > 0.5
            assert result.severity in [ErrorSeverity.MEDIUM, ErrorSeverity.LOW]
            assert "Sheet2!$A" in result.details['expected_reference']
    
    def test_detect_wrong_anchoring_when_copying_down(self):
        """Test detection of wrong anchoring when copying down."""
        # Set up data in Sheet2
        self.sheet2['A1'] = "Header"
        self.sheet2['B1'] = "Value1"
        self.sheet2['C1'] = "Value2"
        self.sheet2['D1'] = "Value3"
        
        # Set up formulas in Sheet1 that should be row-locked when copying down
        # Wrong: fully locked when should be row-locked
        self.sheet1['A2'] = "=Sheet2!$B$1"  # Should be =Sheet2!B$1
        self.sheet1['A3'] = "=Sheet2!$C$1"  # Should be =Sheet2!C$1
        self.sheet1['A4'] = "=Sheet2!$D$1"  # Should be =Sheet2!D$1
        
        # Create a pattern to indicate copying down
        self.sheet1['A1'] = "Row1"
        self.sheet1['B1'] = "Row2"
        self.sheet1['C1'] = "Row3"
        
        results = self.detector.detect(self.workbook)
        
        assert len(results) >= 3
        for result in results:
            assert result.error_type == "cross_sheet_anchoring_errors"
            assert "Wrong cross-sheet anchoring" in result.description
            assert result.probability > 0.5
            assert "$1" in result.details['expected_reference']
    
    def test_detect_wrong_anchoring_for_fixed_references(self):
        """Test detection of wrong anchoring for fixed references."""
        # Set up a header row in Sheet2 (should be fully locked)
        self.sheet2['A1'] = "Product"
        self.sheet2['B1'] = "Price"
        self.sheet2['C1'] = "Category"
        
        # Set up formulas that reference headers but are not fully locked
        self.sheet1['A2'] = "=Sheet2!A1"  # Should be =Sheet2!$A$1
        self.sheet1['B2'] = "=Sheet2!B1"  # Should be =Sheet2!$B$1
        self.sheet1['C2'] = "=Sheet2!C1"  # Should be =Sheet2!$C$1
        
        results = self.detector.detect(self.workbook)
        
        assert len(results) >= 3
        for result in results:
            assert result.error_type == "cross_sheet_anchoring_errors"
            assert "Wrong cross-sheet anchoring" in result.description
            assert result.probability > 0.5
            assert "$" in result.details['expected_reference']
    
    def test_detect_complex_cross_sheet_formulas(self):
        """Test detection in complex cross-sheet formulas."""
        # Set up lookup data in Sheet2
        self.sheet2['A1'] = "ID"
        self.sheet2['B1'] = "Name"
        self.sheet2['A2'] = 1
        self.sheet2['B2'] = "Alice"
        self.sheet2['A3'] = 2
        self.sheet2['B3'] = "Bob"
        
        # Complex VLOOKUP with wrong anchoring
        self.sheet1['A2'] = '=VLOOKUP(A1,Sheet2!$A$1:$B$3,2,FALSE)'  # Should be =VLOOKUP(A1,Sheet2!A1:B3,2,FALSE)
        
        results = self.detector.detect(self.workbook)
        
        assert len(results) >= 1
        for result in results:
            assert result.error_type == "cross_sheet_anchoring_errors"
            assert "VLOOKUP" in result.details['formula']
            assert result.probability > 0.6  # Higher probability for critical functions
    
    def test_detect_quoted_sheet_names(self):
        """Test detection with quoted sheet names."""
        # Create sheet with space in name
        sheet_with_space = self.workbook.create_sheet("My Data")
        sheet_with_space['A1'] = "Value"
        
        # Reference with quoted sheet name and wrong anchoring
        self.sheet1['A2'] = "='My Data'!$A$1"  # Should be ='My Data'!A1
        
        results = self.detector.detect(self.workbook)
        
        assert len(results) >= 1
        for result in results:
            assert result.error_type == "cross_sheet_anchoring_errors"
            assert "'My Data'" in result.details['cross_sheet_reference']
    
    def test_no_errors_for_correct_anchoring(self):
        """Test that no errors are detected for correct anchoring."""
        # Set up data in Sheet2
        self.sheet2['A1'] = "Header"
        self.sheet2['A2'] = 100
        
        # Correct anchoring for copying across
        self.sheet1['B2'] = "=Sheet2!$A2"  # Correct: column-locked
        self.sheet1['C2'] = "=Sheet2!$A2"  # Correct: column-locked
        
        # Correct anchoring for copying down
        self.sheet1['A3'] = "=Sheet2!A$1"  # Correct: row-locked
        self.sheet1['A4'] = "=Sheet2!A$1"  # Correct: row-locked
        
        # Correct anchoring for fixed reference
        self.sheet1['A5'] = "=Sheet2!$A$1"  # Correct: fully locked
        
        results = self.detector.detect(self.workbook)
        
        # Should not detect errors for correct anchoring
        cross_sheet_errors = [r for r in results if r.error_type == "cross_sheet_anchoring_errors"]
        assert len(cross_sheet_errors) == 0
    
    def test_handle_nonexistent_sheet_references(self):
        """Test handling of references to non-existent sheets."""
        # Reference to non-existent sheet
        self.sheet1['A1'] = "=NonExistentSheet!A1"
        
        results = self.detector.detect(self.workbook)
        
        # Should not crash and should not detect errors for non-existent sheets
        cross_sheet_errors = [r for r in results if r.error_type == "cross_sheet_anchoring_errors"]
        assert len(cross_sheet_errors) == 0
    
    def test_probability_calculation(self):
        """Test probability calculation for different scenarios."""
        # Set up critical calculation (VLOOKUP)
        self.sheet2['A1'] = "ID"
        self.sheet2['B1'] = "Value"
        self.sheet2['A2'] = 1
        self.sheet2['B2'] = 100
        
        # Critical function with wrong anchoring
        self.sheet1['A2'] = '=VLOOKUP(A1,Sheet2!$A$1:$B$2,2,FALSE)'  # Should be =VLOOKUP(A1,Sheet2!A1:B2,2,FALSE)
        
        results = self.detector.detect(self.workbook)
        
        assert len(results) >= 1
        # Critical functions should have higher probability
        assert results[0].probability > 0.7
    
    def test_suggested_fixes(self):
        """Test that suggested fixes are provided."""
        # Set up simple scenario
        self.sheet2['A1'] = "Header"
        self.sheet2['A2'] = 100
        
        # Wrong anchoring
        self.sheet1['B2'] = "=Sheet2!$A$2"  # Should be =Sheet2!$A2
        
        results = self.detector.detect(self.workbook)
        
        assert len(results) >= 1
        assert results[0].suggested_fix is not None
        assert "Update cross-sheet reference" in results[0].suggested_fix
        assert "=Sheet2!A$2" in results[0].details['expected_formula']
    
    def test_multiple_cross_sheet_references(self):
        """Test detection with multiple cross-sheet references in one formula."""
        # Set up data in multiple sheets
        self.sheet2['A1'] = "Value1"
        self.sheet3['A1'] = "Value2"
        
        # Formula with multiple cross-sheet references
        self.sheet1['A2'] = "=Sheet2!$A$1 + Data!$A$1"  # Should be =Sheet2!A1 + Data!A1
        
        results = self.detector.detect(self.workbook)
        
        assert len(results) >= 1
        # Complex references should have higher probability
        assert results[0].probability > 0.6
    
    def test_edge_case_empty_formulas(self):
        """Test handling of empty formulas."""
        # Empty formula cell
        self.sheet1['A1'] = ""
        
        results = self.detector.detect(self.workbook)
        
        # Should not crash and should not detect errors
        cross_sheet_errors = [r for r in results if r.error_type == "cross_sheet_anchoring_errors"]
        assert len(cross_sheet_errors) == 0
    
    def test_edge_case_malformed_references(self):
        """Test handling of malformed cross-sheet references."""
        # Malformed reference
        self.sheet1['A1'] = "=Sheet2!A"  # Missing row number
        
        results = self.detector.detect(self.workbook)
        
        # Should not crash
        assert isinstance(results, list)
    
    def test_integration_with_probabilistic_sniffer(self):
        """Test integration with the main ProbabilisticErrorSniffer."""
        # Create a temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            self.workbook.save(tmp.name)
            tmp_path = tmp.name
        
        try:
            # Set up test data
            self.sheet2['A1'] = "Header"
            self.sheet2['A2'] = 100
            self.sheet1['B2'] = "=Sheet2!$A$2"  # Wrong anchoring
            
            self.workbook.save(tmp_path)
            
            # Test with ProbabilisticErrorSniffer
            sniffer = ProbabilisticErrorSniffer(Path(tmp_path))
            sniffer.register_detector(self.detector)
            results = sniffer.detect_all_errors()
            
            # Check that cross-sheet anchoring errors are detected
            cross_sheet_errors = results.get("cross_sheet_anchoring_errors", [])
            assert len(cross_sheet_errors) >= 1
            
        finally:
            # Clean up
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)


class TestCrossSheetAnchoringDetectorHelperMethods:
    """Test helper methods of CrossSheetAnchoringDetector."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.detector = CrossSheetAnchoringDetector()
        self.workbook = Workbook()
        self.sheet1 = self.workbook.active
        self.sheet2 = self.workbook.create_sheet("Sheet2")
    
    def teardown_method(self):
        """Clean up test fixtures."""
        self.workbook.close()
    
    def test_extract_cross_sheet_references(self):
        """Test extraction of cross-sheet references."""
        formula = "=Sheet2!A1 + 'My Sheet'!B2 + Sheet3!C3"
        
        refs = self.detector._extract_cross_sheet_references(formula)
        
        assert len(refs) == 3
        # Check that all expected references are present (order may vary)
        sheet_names = [ref['sheet_name'] for ref in refs]
        cell_refs = [ref['cell_ref'] for ref in refs]
        assert "Sheet2" in sheet_names
        assert "My Sheet" in sheet_names
        assert "Sheet3" in sheet_names
        assert "A1" in cell_refs
        assert "B2" in cell_refs
        assert "C3" in cell_refs
    
    def test_get_anchoring_type_from_ref(self):
        """Test anchoring type detection."""
        assert self.detector._get_anchoring_type_from_ref("A1") == "relative"
        assert self.detector._get_anchoring_type_from_ref("$A1") == "column_locked"
        assert self.detector._get_anchoring_type_from_ref("A$1") == "row_locked"
        assert self.detector._get_anchoring_type_from_ref("$A$1") == "fully_locked"
    
    def test_parse_cell_reference(self):
        """Test cell reference parsing."""
        col, row = self.detector._parse_cell_reference("A1")
        assert col == 1
        assert row == 1
        
        col, row = self.detector._parse_cell_reference("B5")
        assert col == 2
        assert row == 5
        
        col, row = self.detector._parse_cell_reference("$C$10")
        assert col == 3
        assert row == 10
    
    def test_has_cross_sheet_reference(self):
        """Test cross-sheet reference detection."""
        assert self.detector._has_cross_sheet_reference("=Sheet2!A1")
        assert self.detector._has_cross_sheet_reference("='My Sheet'!B2")
        assert not self.detector._has_cross_sheet_reference("=A1+B2")
        assert not self.detector._has_cross_sheet_reference("=SUM(A1:A10)")
    
    def test_suggest_correct_cross_sheet_reference(self):
        """Test correct reference suggestion."""
        assert self.detector._suggest_correct_cross_sheet_reference("A1", "fully_locked") == "$A$1"
        assert self.detector._suggest_correct_cross_sheet_reference("A1", "column_locked") == "$A1"
        assert self.detector._suggest_correct_cross_sheet_reference("A1", "row_locked") == "A$1"
        assert self.detector._suggest_correct_cross_sheet_reference("A1", "relative") == "A1"
    
    def test_is_critical_cross_sheet_calculation(self):
        """Test critical calculation detection."""
        self.sheet1['A1'] = "=VLOOKUP(A1,Sheet2!A1:B10,2,FALSE)"
        assert self.detector._is_critical_cross_sheet_calculation(self.sheet1, 1, 1)
        
        self.sheet1['A2'] = "=Sheet2!A1 + Sheet2!B1"
        assert not self.detector._is_critical_cross_sheet_calculation(self.sheet1, 2, 1)
    
    def test_is_complex_cross_sheet_reference(self):
        """Test complex reference detection."""
        formula = "=Sheet2!A1 + Sheet3!B1 + Sheet4!C1"
        assert self.detector._is_complex_cross_sheet_reference(formula)
        
        formula = "=Sheet2!A1"
        assert not self.detector._is_complex_cross_sheet_reference(formula)


if __name__ == "__main__":
    pytest.main([__file__]) 